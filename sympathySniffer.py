import pandas as pd

import yfinance as yf
import bs4 as bs
import pickle
import requests
import os
import csv
import glob
import pandas as pdlib
import operator
import openpyxl

import xlrd


def compare_tickers(master_file):
    print("-> Comparing tickers to see which one's saw gains on the same day\n")

    # list = pd.read_csv(master_file)
    # print(list.head())
    stock_list = pdlib.read_csv(master_file, delimiter = ',',
                                )
    # print(stock_list.head())
    pd.to_datetime(stock_list['Date'],format='%Y-%m-%d')
    # print(stock_list.sort_values(by='Date'))
    return 0

def get_test_ticker_data(ticker):
    for x in ticker:
        if not os.path.exists('test'):
            os.makedirs('test')

        if not os.path.exists('test/{}.csv'.format(ticker)):
            try:
                print('Getting data for: ' + ticker)
                df = yf.download(ticker, period='3y', interval='1d')
                df.dropna(inplace=True)
                df["Symbol"] = ticker
                df["PercentIncrease_High_Open"] = df['High']*100 / df['Open'] -100
                df["Float"] = yf.Ticker(ticker).info['floatShares']
                df.to_csv('test/{}.csv'.format(ticker))

            except Exception as ex:
                print('Error:', ex)
        else:
            print('-> Already have {}.csv file so skipping'.format(ticker))
        # print('Getting data for: ' + ticker)
        # df = yf.download(ticker, period='1y', interval='1d')
        # df.dropna(inplace=True)
        # df.to_csv('oil/{}.csv'.format(ticker))
def create_test_master_list(tickers,file_out):
    # Produce a single CSV after combining all files

    result_obj = pdlib.concat([pdlib.read_csv('test/{}.csv'.format(ticker)) for ticker in tickers])
    # Convert the above object into a csv file and export
    result_obj.to_csv(file_out, index=False, encoding="utf-8")
    print("\n\n-> Master file called %s  has been generated\n" % file_out)

def save_sp500_tickers():
    resp = requests.get('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')
    soup = bs.BeautifulSoup(resp.text, "lxml")
    table = soup.find('table', {'class': 'wikitable sortable'})
    tickers = []
    for row in table.findAll('tr')[1:]:
        ticker = row.findAll('td')[0].text
        ticker = ticker[:-1]
        tickers.append(ticker + '\n')
    # print(tickers)
    with open('sp500tickers.pickle.text', 'wb') as f:
        pickle.dump(tickers, f)
    return tickers




#
def get_ticker_data(tickers,directory):
    for x in tickers:
        if not os.path.exists('sector'):
            os.makedirs('sector')
        if not os.path.exists('sector/{}'.format(directory)):
            os.makedirs('sector/{}'.format(directory))

        if not os.path.exists('sector/{}/{}.csv'.format(directory,x)):
            try:
                print('Getting data for: ' + x)
                df = yf.download(x , period='3y', interval='1d')
                df.dropna(inplace=True)
                df["Symbol"] = x
                df["PercentIncrease_High_Open"] = df['High']*100 / df['Open'] -100
                df["Sector"] = '{}'.format(directory)
                df.to_csv('sector/{}/{}.csv'.format(directory,x))

            except Exception as ex:
                print('Error:', ex)
        else:
            print('-> Already have {} file so skipping'.format(x))

def create_master_list(tickers,file_out,master_list_directory):
    print('Creating master file: {}'.format(file_out))
    # Produce a single CSV after combining all files

    result_obj = pdlib.concat([pdlib.read_csv('sector/{}/{}.csv'.format(master_list_directory,ticker)) for ticker in tickers])
    # Convert the above object into a csv file and export
    result_obj.to_csv('sector/{}'.format(file_out), index=False, encoding="utf-8")
    print("\n\n-> Master file called %s  has been generated\n" % file_out)

def purge_low_gain_volume(master_file):
    print ('Purging low volume and low gain entries')
    lines = list()
    lines.append(['Date','Open','High','Low','Close','Adj Close','Volume','Symbol','PercentIncrease_High_Open',
                  'Sector'])
    with open('sector/{}'.format(master_file),'r') as readFile:
        reader = csv.reader(readFile)
        next(reader)
        for row in reader:
            if row[8] =='':
                continue
            else:
                if float(row[8]) >= 10.0 and float(row[6]) >= 10000000:
                    lines.append(row)
    print("-> Sorting the list by Date and Volume\n")
    sortedlist = sorted(lines, key=operator.itemgetter(0,6), reverse=True)

    with open('sector/{}'.format(master_file),'w') as writeFile:
        writer = csv.writer(writeFile)
        writer.writerows(sortedlist)
    # print ("Getting floats")

    print("-> Removed anything below 10%% percent gain between Open and High from: %s\n" % master_file)
    print("-> Removed anything with volume below 10 million for that day from: %s\n" % master_file)
    return 0
def get_industry_ticker_data(tickers,directory):
    for x in tickers:
        if not os.path.exists('industry'):
            os.makedirs('industry')
        if not os.path.exists('industry/{}'.format(directory)):
            os.makedirs('industry/{}'.format(directory))

        if not os.path.exists('industry/{}/{}.csv'.format(directory,x)):
            try:
                print('Getting data for: ' + x)
                df = yf.download(x , period='3y', interval='1d')
                df.dropna(inplace=True)
                df["Symbol"] = x
                df["PercentIncrease_High_Open"] = df['High']*100 / df['Open'] -100
                df["Industry"] = '{}'.format(directory)
                df["Float"] = '{}'.format(directory)
                df.to_csv('industry/{}/{}.csv'.format(directory,x))

            except Exception as ex:
                print('Error:', ex)
        else:
            print('-> Already have {} file so skipping'.format(x))

def create_industry_master_list(tickers,file_out,master_list_directory):
    print('Creating master file: {}'.format(file_out))
    # Produce a single CSV after combining all files

    result_obj = pdlib.concat([pdlib.read_csv('industry/{}/{}.csv'.format(master_list_directory,ticker)) for ticker in tickers])
    # Convert the above object into a csv file and export
    result_obj.to_csv('industry/{}'.format(file_out), index=False, encoding="utf-8")
    print("\n\n-> Master file called %s  has been generated\n" % file_out)

def purge_industry_low_gain_volume(master_file):
    print ('Purging low volume and low gain entries')
    lines = list()
    lines.append(['Date','Open','High','Low','Close','Adj Close','Volume','Symbol','PercentIncrease_High_Open',
                  'Industry','Float'])
    with open('industry/{}'.format(master_file),'r') as readFile:
        reader = csv.reader(readFile)
        next(reader)
        for row in reader:
            if row[8] =='':
                continue
            else:
                if float(row[8]) >= 10.0 and float(row[6]) >= 10000000:
                    lines.append(row)
    print("-> Sorting the list by Date and Volume\n")
    sortedlist = sorted(lines, key=operator.itemgetter(0,6), reverse=True)

    with open('industry/{}'.format(master_file),'w') as writeFile:
        writer = csv.writer(writeFile)
        writer.writerows(sortedlist)
    # print ("Getting floats")

    print("-> Removed anything below 10%% percent gain between Open and High from: %s\n" % master_file)
    print("-> Removed anything with volume below 10 million for that day from: %s\n" % master_file)
    return 0

def combine_masters_into_one(directory,name):
    with open('{}/main.csv'.format(directory),"w") as empty_file:
        pass
    lines = list()
    lines.append(['Link'])
    with open('{}/main.csv'.format(directory),'w') as writeFile:
        writer = csv.writer(writeFile)
        writer.writerows(lines)

    all_files = glob.glob(os.path.join(directory, "*.csv"))
    sector_writer = pd.ExcelWriter(name, engine='xlsxwriter')
    df2 = pd.DataFrame()
    for f in all_files:
        df = pd.read_csv(f)
        df.to_excel(sector_writer,sheet_name = os.path.splitext(os.path.basename(f))[0], index=False)
    sector_writer.save()

    wb = openpyxl.load_workbook('{}'.format(name))
    worksheetnames = wb.sheetnames
    sheet_index = worksheetnames.index('main')
    wb.active=sheet_index
    for sheet in wb:
        if sheet.title == 'main':
            sheet.sheet_view.tabSelected = True
        else:
            sheet.sheet_view.tabSelected=False
    wb.save(name)



    return 0




#sectors

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
consumer_durables =['ARLO', 'ASPU', 'ATVI', 'BC', 'BHAT', 'BNSO', 'BSET', 'BZH', 'CCS', 'CHCI', 'CLAR', 'CRWS', 'CSPR', 'CTB', 'CTHR', 'CVCO', 'DHI', 'DOOO', 'DORM', 'DRTT', 'DXYN', 'EA', 'ELY', 'ESCA', 'ETH', 'F','FCAU','FIT', 'FLXS', 'FNKO', 'FORD', 'FOSL', 'FOXF', 'FTDR', 'FUV', 'GIGM', 'GM', 'GNSS', 'GOLF', 'GPRO', 'GRBK', 'GT', 'HAS', 'HBB', 'HEAR', 'HELE', 'HGSH', 'HMC', 'HMI', 'HOFT', 'HOG', 'HOV', 'HZN', 'IRBT', 'JAKK', 'JOUT', 'KBAL', 'KBH', 'KGJI', 'KNDI', 'KOSS', 'LCUT', 'LEG', 'LEGH', 'LEN', 'LEN.B', 'LGIH', 'LKQ', 'LZB', 'MAT', 'MBUU', 'MCFT', 'MDC', 'MHK', 'MHO', 'MOV', 'MPX', 'MSA', 'MSN', 'MTH', 'NIO', 'NIU', 'NLS', 'NPK', 'NTZ', 'NVFY', 'NVR', 'NWHM', 'OC', 'PATK', 'PHM', 'PII', 'PTE', 'RACE', 'REVG', 'RGR', 'ROKU', 'SCX', 'SKY', 'SLGG', 'SMP', 'SNA', 'SNBR', 'SNE', 'SOLO', 'SPB', 'SWK', 'THO', 'TLF','TM', 'TMHC', 'TOL', 'TPH', 'TPX', 'TSLA', 'TTM', 'TTWO', 'UEIC', 'VIOT', 'VSTO', 'WGO', 'WHR', 'XPEL']
consumer_durables_master_file='consumer_durables_mf.csv'
consumer_durables_directory='consumer_durables'
get_ticker_data(consumer_durables,consumer_durables_directory)
create_master_list(consumer_durables,consumer_durables_master_file,consumer_durables_directory)
print("-> Master file created, filename: %s\n" % consumer_durables_master_file)
purge_low_gain_volume(consumer_durables_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
consumer_non_durables = ['ABEV', 'AKO.A', 'AKO.B', 'APEX', 'BF.A', 'BF.B', 'BGS', 'BREW', 'BRFS', 'BRID', 'BTI', 'BUD', 'BYND', 'CAG', 'CAL', 'CALM', 'CCEP', 'CCU', 'CELH', 'CENT', 'CENTA', 'CHD', 'CL', 'CLX', 'COKE', 'COLM', 'COTY', 'CPB', 'CROX', 'CVGW', 'DECK', 'DEO', 'DLA', 'DOGZ', 'DTEA', 'EL', 'ELF', 'ENR', 'EPC', 'EVK', 'FARM', 'FDP', 'FIZZ', 'FLO', 'FMX', 'FRPT', 'FTFT', 'GES', 'GIII', 'GIL', 'GIS', 'GOOS', 'HAIN', 'HBI', 'HRL', 'HSY', 'ICON', 'IFF', 'IFFT', 'IPAR', 'JBSS', 'JJSF', 'JRSH', 'JVA', 'K', 'KBSF', 'KDP', 'KHC', 'KMB', 'KO', 'KOF', 'KTB', 'LANC', 'LEVI', 'LK', 'LNDC', 'LW', 'LWAY', 'MDLZ', 'MKC', 'MKC.V', 'MNST', 'MO', 'NAKD', 'NATR', 'NBEV', 'NHTC', 'NKE', 'NOMD', 'NTCO', 'NUS', 'OBCI', 'OXM', 'PEP', 'PETZ', 'PG', 'PLAG', 'PLIN', 'PM', 'POST', 'PPC', 'PRMW', 'PVH', 'PYX', 'RCKY', 'REED', 'RELV', 'REV', 'RIBT', 'RMCF', 'RUHN', 'SAFM', 'SAM', 'SDI', 'SEB', 'SENEA', 'SGC', 'SHOO', 'SJM', 'SKX', 'SMPL', 'SQBG', 'STKL', 'STZ', 'STZ.B', 'TAP', 'TAP.A', 'TBLT', 'THS', 'TPB', 'TR', 'TSN', 'TWNK', 'UA', 'UAA', 'UG', 'UL', 'UN', 'USNA', 'UVV', 'VFC', 'VGR', 'VNCE', 'VRA', 'WTER', 'WVVI', 'WWW', 'XELB']
consumer_non_durables_master_file='consumer_non_durables_mf.csv'
consumer_non_durables_directory='consumer_non_durables'
get_ticker_data(consumer_non_durables,consumer_non_durables_directory)
create_master_list(consumer_non_durables,consumer_non_durables_master_file,consumer_non_durables_directory)
print("-> Master file created, filename: %s\n" % consumer_non_durables_master_file)
purge_low_gain_volume(consumer_non_durables_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
consumer_services = ['ACEL', 'AFYA', 'AHC', 'AMBO', 'AMC', 'AMCX', 'APEI', 'ARCO', 'ARKR', 'ARMK', 'ATGE', 'ATUS', 'AYTU', 'BATRA', 'BATRK', 'BBGI', 'BBQ', 'BDL', 'BH', 'BH.A', 'BJRI', 'BKNG', 'BLMN', 'BLNK', 'BTN', 'BWL.A', 'BXG', 'BYD', 'CABO', 'CAKE', 'CBRL', 'CCL', 'CETV', 'CHDN', 'CHH', 'CHTR', 'CHUY', 'CLCT', 'CLUB', 'CMCSA', 'CMG', 'CMLS', 'CNK', 'CNTY', 'CPHC', 'CSSE', 'CSV', 'CTAS', 'CUK', 'CVEO', 'CZR', 'DENN', 'DESP', 'DIN', 'DIS', 'DISCA', 'DISCB', 'DISCK', 'DISH', 'DJCO', 'DLB', 'DLPN', 'DNKN', 'DPZ', 'DRI', 'DS', 'DV', 'EAT', 'EBAY', 'EDU', 'EDUC', 'ERI', 'EROS', 'ETM', 'EVC', 'EVRI', 'EXPE', 'FAT', 'FENG', 'FLL', 'FOX', 'FOXA', 'FRGI', 'FUN', 'FWONA', 'FWONK', 'GCI', 'GDEN', 'GHC', 'GHG', 'GMBL', 'GNUS', 'GRIL', 'GRVY', 'GTIM', 'GTN', 'GTN.A', 'H', 'HLG', 'HLT', 'HMHC', 'HMTV', 'HRB', 'HTHT', 'IGT', 'IHG', 'IHRT', 'IMAX', 'JACK', 'JAX', 'JW.A', 'JW.B', 'KRUS', 'LBRDA', 'LBRDK', 'LBTYA', 'LBTYB', 'LBTYK', 'LEE', 'LGF.A', 'LGF.B', 'LILA', 'LILAK', 'LIND', 'LOCO', 'LOV', 'LQDT', 'LRN', 'LSXMA', 'LSXMB', 'LSXMK', 'LTRPA', 'LTRPB', 'LUB', 'LVS', 'LYV', 'MANU', 'MAR', 'MCD', 'MCRI', 'MCS', 'MDIA', 'MDP', 'MGM', 'MIN', 'MLCO', 'MMYT', 'MRKR', 'MSC', 'MSGE', 'MSGN', 'MSGS', 'MTN', 'NATH', 'NCLH', 'NDLS', 'NFLX', 'NTN', 'NWGI', 'NWS', 'NWSA', 'NXST', 'NYT', 'OSW', 'PBPB', 'PENN', 'PLAY', 'PLNT', 'PLYA', 'PRDO', 'PSO', 'PTON', 'PZZA', 'QSR', 'RAVE', 'RCL', 'RGS', 'RICK', 'RLH', 'ROL', 'RRGB', 'RRR', 'RST', 'RUTH', 'SALM', 'SBGI', 'SBUX', 'SCHL', 'SCI', 'SEAS', 'SERV', 'SGA', 'SGMS', 'SHAK', 'SIRI', 'SIX', 'SJR', 'SSP', 'STAY', 'STKS', 'STON', 'STRA', 'TA', 'TACO', 'TAL', 'TAST', 'TCOM', 'TGNA', 'TH', 'TOUR', 'TPCO', 'TRCH', 'TRIP', 'TRWH', 'TSQ', 'TV', 'TXRH', 'UNF', 'UONE', 'UONEK', 'VAC', 'VIAC', 'VIACA', 'WEN', 'WH', 'WING', 'WSG', 'WW', 'WWE', 'WYND', 'WYNN', 'XSPA', 'YCBD', 'YTRA', 'YUM', 'YUMC', 'ZVOD']
consumer_services_master_file='consumer_services_mf.csv'
consumer_services_directory='consumer_services'
get_ticker_data(consumer_services,consumer_services_directory)
create_master_list(consumer_services,consumer_services_master_file,consumer_services_directory)
print("-> Master file created, filename: %s\n" % consumer_services_master_file)
purge_low_gain_volume(consumer_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
electronic_technology = ['AAOI', 'AAPL', 'AAXN', 'ACIA', 'ACLS', 'ADI', 'ADTN', 'AEHR', 'AEIS', 'AIR', 'AIRG', 'AIRI', 'AJRD', 'AKTS', 'ALLT', 'ALOT', 'AMBA', 'AMD', 'AMKR', 'AMOT', 'ANET', 'AOSL', 'APH', 'APWC', 'ASTC', 'ASX', 'ATRO', 'AUDC', 'AVAV', 'AVGO', 'AVNW', 'AVYA', 'AXTI', 'BA', 'BDR', 'BELFA', 'BELFB', 'BHE', 'BKTI', 'BOSC', 'BRKS', 'BW', 'CAJ', 'CALX', 'CAMP', 'CAMT', 'CAN', 'CASA', 'CCMP', 'CETX', 'CEVA', 'CGNX', 'CIEN', 'CLFD', 'CLIR', 'CLRO', 'CLS', 'CLWT', 'CMBM', 'CMTL', 'CODA', 'COHR', 'COHU', 'COMM', 'CREE', 'CRNT', 'CRUS', 'CTS', 'CUB', 'CVU', 'CVV', 'CW', 'CYBE', 'DAIO', 'DAKT', 'DBD', 'DCO', 'DELL', 'DGII', 'DGLY', 'DIOD', 'DSPG', 'DZSI', 'ELTK', 'EMAN', 'EMKR', 'ENPH', 'ENTG', 'ERIC', 'ERJ', 'ESE', 'ESLT', 'ESP', 'EXFO', 'EXTR', 'FARO', 'FCEL', 'FEIM', 'FFIV', 'FLEX', 'FLIR', 'FN', 'FORM', 'FSLR', 'FTNT', 'FTV', 'GD', 'GE', 'GILT', 'GLW', 'GRMN', 'GSIT', 'HEI', 'HEI.A', 'HII', 'HIMX', 'HLIT', 'HOLI', 'HPE', 'HPQ', 'HXL', 'ICHR', 'IDCC', 'IDN', 'IEC', 'IIVI', 'IMMR', 'IMOS', 'IMTE', 'INFN', 'INSG', 'INTC', 'INTT', 'INVE', 'IOTS', 'IPGP', 'IPHI', 'ISNS', 'ISSC', 'ITI', 'ITRI', 'ITRN', 'IVAC', 'JBL', 'JCS', 'KAMN', 'KEM', 'KEYS', 'KLAC', 'KLIC', 'KN', 'KODK', 'KOPN', 'KTCC', 'KTOS', 'KVHI', 'LASR', 'LEDS', 'LGL', 'LHX', 'LMT', 'LOGI', 'LPL', 'LPTH', 'LRCX', 'LSCC', 'LUNA', 'MAGS', 'MAXR', 'MCHP', 'MEI', 'MICT', 'MKSI', 'MOG.A', 'MOG.B', 'MOSY', 'MPWR', 'MRAM', 'MRCY', 'MRVL', 'MSI', 'MTSC', 'MTSI', 'MU', 'MVIS', 'MX', 'MXIM', 'MXL', 'NCR', 'NOC', 'NOK', 'NOVT', 'NPTN', 'NSYS', 'NTAP', 'NTP', 'NVDA', 'NVMI', 'NVT', 'NXPI', 'OBAS', 'OBLG', 'OCC', 'OIIM', 'OLED', 'ON', 'ONTO', 'OSIS', 'OSS', 'PANW', 'PAR', 'PI', 'PKE', 'PLAB', 'PLT', 'PLUG', 'PLXS', 'POLA', 'POWI', 'PRCP', 'PSTG', 'PWFL', 'PXLW', 'QCOM', 'QMCO', 'QRVO', 'QUIK', 'QUMU', 'RADA', 'RBBN', 'RBCN', 'RDWR', 'REFR', 'RELL', 'RESN', 'RFIL', 'RMBS', 'ROP', 'RTX', 'SANM', 'SATS', 'SCKT', 'SCON', 'SEAC', 'SENS', 'SGH', 'SGMA', 'SGOC', 'SIF', 'SILC', 'SIMO', 'SITM', 'SLAB', 'SMCI', 'SMTC', 'SMTX', 'SOL', 'SONM', 'SONO', 'SPCB', 'SPCE', 'SPI', 'SPR', 'SQNS', 'SSTI', 'SSYS', 'ST', 'STM', 'STX', 'SWBI', 'SWIR', 'SWKS', 'TACT', 'TATT', 'TCCO', 'TDC', 'TDG', 'TDY', 'TEL', 'TER', 'TESS', 'TGI', 'TRMB', 'TRNS', 'TRT', 'TSEM', 'TSM', 'TTMI', 'TXN', 'TXT', 'UAVS', 'UCTT', 'UI', 'UMC', 'UTSI', 'UUU', 'VCRA', 'VECO', 'VERT.U', 'VIAV', 'VICR', 'VISL', 'VOXX', 'VPG', 'VRT', 'VSAT', 'VSH', 'VUZI', 'WATT', 'WDC', 'WISA', 'WSTL', 'WTT', 'XLNX', 'XPER', 'XRX', 'ZBRA']
electronic_technology_master_file='electronic_technology_mf.csv'
electronic_technology_directory='electronic_technology'
get_ticker_data(electronic_technology,electronic_technology_directory)
create_master_list(electronic_technology,electronic_technology_master_file,electronic_technology_directory)
print("-> Master file created, filename: %s\n" % electronic_technology_master_file)
purge_low_gain_volume(electronic_technology_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
energy_minerals =['AMPY', 'APA', 'AR', 'ARCH', 'AREC', 'ARLP', 'AXAS', 'BATL', 'BCEI', 'BP', 'BRN', 'BRY', 'BSM', 'BTE', 'BTU', 'CCR', 'CDEV', 'CEI', 'CEIX', 'CEO', 'CHAP', 'CHK', 'CLR', 'CNQ', 'CNX', 'COG', 'COP', 'CPE', 'CPG', 'CRC', 'CRK', 'CRT', 'CTRA', 'CVE', 'CVI', 'CVX', 'CXO', 'CZZ', 'DK', 'DMLP', 'DNR', 'DVN', 'E', 'EC', 'ECT', 'EGY', 'ENSV', 'EOG', 'EPM', 'EPSN', 'EQNR', 'EQT', 'ERF', 'ESTE', 'FANG', 'FLMN', 'FLNG', 'GBR', 'GDP', 'GPOR', 'GPRK', 'GTE', 'HCC', 'HES', 'HESM', 'HFC', 'HNRG', 'HPR', 'HUSA', 'IMO', 'INDO', 'KOS', 'LLEX', 'LONE', 'LPI', 'MARPS', 'MCEP', 'MCF', 'METC', 'MGY', 'MPC', 'MR', 'MRO', 'MTDR', 'MTR', 'MUR', 'MXC', 'NBL', 'NC', 'NEXT', 'NFG', 'NOG', 'NRP', 'NRT', 'OAS', 'OVV', 'OXY', 'PARR', 'PBF', 'PBR', 'PBR.A', 'PDCE', 'PE', 'PER', 'PHX', 'PNRG', 'PRT', 'PSX', 'PTR', 'PVAC', 'PVL', 'PXD', 'QEP', 'RDS.A', 'RDS.B', 'REI', 'ROYT', 'RRC', 'SBOW', 'SBR', 'SD', 'SM', 'SNDE', 'SNMP', 'SNP', 'SU', 'SWN', 'SXC', 'TALO', 'TAT', 'TELL', 'TGA', 'TGC', 'TGS', 'TOT', 'TREC', 'TRGP', 'UGP', 'USEG', 'VET', 'VIST', 'VLO', 'VNOM', 'VOC', 'WES', 'WLL', 'WPX', 'WTI', 'XEC', 'XOG', 'XOM', 'YPF', 'ZN']
energy_minerals_master_file='energy_minerals_mf.csv'
energy_minerals_directory='energy_minerals'
get_ticker_data(energy_minerals,energy_minerals_directory)
create_master_list(energy_minerals,energy_minerals_master_file,energy_minerals_directory)
print("-> Master file created, filename: %s\n" % energy_minerals_master_file)
purge_low_gain_volume(energy_minerals_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
finance = ['AAMC', 'AAME', 'AAN', 'AAT', 'AB', 'ABCB', 'ABR', 'ABTX', 'AC', 'ACAM', 'ACAMU', 'ACBI', 'ACC', 'ACGL', 'ACNB', 'ACRE', 'ACTT', 'ACTTU', 'ACY', 'ADC', 'AEB', 'AEG', 'AEL', 'AER', 'AFG', 'AFH', 'AFIN', 'AFL', 'AGBA', 'AGM', 'AGM.A', 'AGMH', 'AGNC', 'AGO', 'AHCO', 'AHH', 'AHT', 'AI', 'AIG', 'AIHS', 'AINC', 'AINV', 'AIV', 'AIZ', 'AJG', 'AJX', 'AJXA', 'AKR', 'AL', 'ALAC', 'ALACU', 'ALEX', 'ALL', 'ALLY', 'ALRS', 'ALUS', 'ALUS.U', 'ALX', 'AMAL', 'AMBC', 'AMCI', 'AMCIU', 'AMG', 'AMH', 'AMHC', 'AMHCU', 'AMK', 'AMNB', 'AMP', 'AMRB', 'AMRK', 'AMSF', 'AMT', 'AMTB', 'AMTBB', 'AMTD', 'ANAT', 'ANDA', 'ANH', 'AON', 'APAM', 'APG', 'APLE', 'APO', 'APTS', 'APXT', 'APXTU', 'ARCC', 'ARE', 'ARES', 'ARGD', 'ARGO', 'ARI', 'ARL', 'AROW', 'ARR', 'ARYA', 'ARYAU', 'ASB', 'ASFI', 'ASPS', 'ASRV', 'ATAX', 'ATCO', 'ATCX', 'ATH', 'ATIF', 'ATLC', 'ATLO', 'AUB', 'AUBN', 'AVAL', 'AVB', 'AVCT', 'AX', 'AXP', 'AXR', 'AXS', 'BAC', 'BAM', 'BANC', 'BANF', 'BANR', 'BAP', 'BBAR', 'BBCP', 'BBD', 'BBDC', 'BBDO', 'BBVA', 'BBX', 'BCBP', 'BCH', 'BCML', 'BCOW', 'BCS', 'BCTF', 'BDGE', 'BDN', 'BEN', 'BFC', 'BFIN', 'BFS', 'BFST', 'BFYT', 'BGCP', 'BHB', 'BHF', 'BHFAL', 'BHLB', 'BHR', 'BIOX', 'BK', 'BKCC', 'BKSC', 'BKU', 'BLK', 'BLX', 'BMA', 'BMO', 'BMRC', 'BMRG', 'BMRG.U', 'BMTC', 'BNS', 'BOCH', 'BOH', 'BOKF', 'BOTJ', 'BPFH', 'BPOP', 'BPRN', 'BPY', 'BPYU', 'BRBS', 'BRG', 'BRK.A', 'BRK.B', 'BRKL', 'BRMK', 'BRO', 'BROG', 'BRP', 'BRT', 'BRX', 'BSAC', 'BSBK', 'BSBR', 'BSIG', 'BSMX', 'BSRR', 'BSVN', 'BUSE', 'BWB', 'BWFG', 'BX', 'BXMT', 'BXP', 'BXS', 'BY', 'BYFC', 'C', 'CAC', 'CACC', 'CADE', 'CAI', 'CALB', 'CAR', 'CARE', 'CARV', 'CASH', 'CATC', 'CATY', 'CB', 'CBAN', 'CBFV', 'CBL', 'CBMB', 'CBNK', 'CBOE', 'CBRE', 'CBSH', 'CBTX', 'CBU', 'CCAC', 'CCAC.U', 'CCAP', 'CCB', 'CCBG', 'CCH', 'CCH.U', 'CCI', 'CCNE', 'CCX', 'CCX.U', 'CCXX', 'CCXX.U', 'CDOR', 'CDR', 'CFB', 'CFBI', 'CFBK', 'CFFA', 'CFFAU', 'CFFI', 'CFFN', 'CFG', 'CFR', 'CG', 'CGBD', 'CGROU', 'CHAQ.U', 'CHCO', 'CHCT', 'CHMG', 'CHMI', 'CHPM', 'CHPMU', 'CIA', 'CIB', 'CIGI', 'CIIC', 'CIICU', 'CIM', 'CINF', 'CIO', 'CIT', 'CIVB', 'CIZN', 'CLBK', 'CLDB', 'CLDT', 'CLI', 'CLNC', 'CLNY', 'CLPR', 'CM', 'CMA', 'CMCT', 'CME', 'CMO', 'CNA', 'CNBKA', 'CNF', 'CNFR', 'CNFRL', 'CNNB', 'CNNE', 'CNO', 'CNOB', 'CNS', 'CODI', 'COF', 'COFS', 'COHN', 'COLB', 'COLD', 'CONE', 'COOP', 'COR', 'CORR', 'COWN', 'CPAA', 'CPAAU', 'CPF', 'CPLG', 'CPSS', 'CPT', 'CPTA', 'CRD.A', 'CRD.B', 'CRESY', 'CRSA', 'CRSAU', 'CRVL', 'CS', 'CSFL', 'CSTR', 'CSWC', 'CTBI', 'CTO', 'CTRE', 'CTT', 'CUBE', 'CUBI', 'CURO', 'CUZ', 'CVBF', 'CVCY', 'CVLY', 'CWBC', 'CWK', 'CXP', 'CXW', 'CZNC', 'CZWI', 'DB', 'DBCP', 'DCOM', 'DEA', 'DEI', 'DFNS', 'DFNS.U', 'DFPH', 'DFPHU', 'DFS', 'DGICA', 'DGICB', 'DHC', 'DHIL', 'DLR', 'DMYT', 'DMYT.U', 'DNJR', 'DOC', 'DPHC', 'DPHCU', 'DRE', 'DRH', 'DX', 'DXF', 'EARN', 'EBIX', 'EBMT', 'EBSB', 'EBTC', 'ECPG', 'EFC', 'EFSC', 'EGBN', 'EGP', 'EHTH', 'EIG', 'ELS', 'ELVT', 'EMCF', 'ENOB', 'ENVA', 'EPR', 'EPRT', 'EQBK', 'EQC', 'EQH', 'EQIX', 'EQR', 'EQS', 'ERIE', 'ESBA', 'ESBK', 'ESGR', 'ESNT', 'ESQ', 'ESRT', 'ESS', 'ESSA', 'ESSC', 'ESSCU', 'ESXB', 'ETFC', 'EV', 'EVBN', 'EVR', 'EWBC', 'EXPC', 'EXPCU', 'EXPI', 'EXR', 'EZPW', 'FAF', 'FANH', 'FBC', 'FBIZ', 'FBK', 'FBMS', 'FBNC', 'FBP', 'FBSS', 'FCAP', 'FCBC', 'FCBP', 'FCCO', 'FCCY', 'FCF', 'FCFS', 'FCNCA', 'FCPT', 'FDBC', 'FDEF', 'FDUS', 'FEAC', 'FEAC.U', 'FFBC', 'FFBW', 'FFG', 'FFIC', 'FFIN', 'FFNW', 'FFWM', 'FGBI', 'FHB', 'FHI', 'FHN', 'FIBK', 'FINV', 'FISI', 'FISK', 'FITB', 'FLIC', 'FLY', 'FMAO', 'FMBH', 'FMBI', 'FMCI', 'FMCIU', 'FMNB', 'FNB', 'FNCB', 'FNF', 'FNHC', 'FNLC', 'FNWB', 'FOCS', 'FOR', 'FPAC', 'FPAC.U', 'FPAY', 'FPH', 'FPI', 'FR', 'FRAF', 'FRBA', 'FRBK', 'FRC', 'FRHC', 'FRME', 'FRPH', 'FRSX', 'FRT', 'FSB', 'FSBW', 'FSEA', 'FSFG', 'FSK', 'FSP', 'FSRV', 'FSRVU', 'FSV', 'FTAC', 'FTACU', 'FTAI', 'FULT', 'FUNC', 'FUSB', 'FVAC.U', 'FVCB', 'FXNC', 'GABC', 'GAIN', 'GATX', 'GBCI', 'GBDC', 'GBL', 'GBLI', 'GCAP', 'GCBC', 'GDOT', 'GDYN', 'GECC', 'GEO', 'GFED', 'GGAL', 'GHIV', 'GHIVU', 'GHL', 'GIK.U', 'GIX', 'GIX.U', 'GL', 'GLAD', 'GLADD', 'GLBZ', 'GLEO', 'GLEO.U', 'GLG', 'GLPI', 'GLRE', 'GMHI', 'GMHIU', 'GMRE', 'GMTA', 'GNL', 'GNRS', 'GNRSU', 'GNTY', 'GNW', 'GOOD', 'GPAQ', 'GPMT', 'GRAF', 'GRAF.U', 'GRIF', 'GRNQ', 'GRNV', 'GROW', 'GRP.U', 'GS', 'GSBC', 'GSBD', 'GSHD', 'GSKY', 'GSMG', 'GTY', 'GWB', 'GWGH', 'GXGX', 'GXGXU', 'GYC', 'GYRO', 'HAFC', 'HALL', 'HASI', 'HBAN', 'HBCP', 'HBMD', 'HBNC', 'HBT', 'HCAC', 'HCCH', 'HCCO', 'HCCOU', 'HCFT', 'HCI', 'HCXY', 'HCXZ', 'HDB', 'HEES', 'HFBL', 'HFWA', 'HGV', 'HHC', 'HIFS', 'HIG', 'HIW', 'HKIB', 'HLI', 'HLNE', 'HMG', 'HMN', 'HMNF', 'HMST', 'HNNA', 'HOMB', 'HONE', 'HOPE', 'HPP', 'HQY', 'HR', 'HRI', 'HRTG', 'HSBC',  'HST', 'HT', 'HTA', 'HTBI', 'HTBK', 'HTGC', 'HTH', 'HTLF', 'HTZ', 'HUIZ', 'HUSN', 'HVBC', 'HWBK', 'HWC', 'HX', 'HYAC', 'HYACU', 'HYRE', 'IBCP', 'IBKC', 'IBKR', 'IBN', 'IBOC', 'IBTX', 'ICBK', 'ICCH', 'ICE', 'ICMB', 'IFS', 'IGIC', 'IHC', 'IHT', 'IIPR', 'ILPT', 'IMH', 'IMVT', 'IMVTU', 'IMXI', 'INBK', 'INDB', 'INFO', 'ING', 'INN', 'INSU', 'INSUU', 'INTG', 'INTL', 'INVH', 'IOR', 'IPOB.U', 'IPOC.U', 'IPV', 'IPV.U', 'IRCP', 'IRET', 'IRM', 'IROQ', 'IRS', 'IRT', 'ISBC', 'ISTR', 'ITCB', 'ITIC', 'ITUB', 'IVR', 'IVZ', 'IX', 'JBGS', 'JCAP', 'JEF', 'JFIN', 'JFK', 'JHG', 'JIH', 'JIH.U', 'JLL', 'JMP', 'JMPNL', 'JOE', 'JP', 'JPM', 'JRVR', 'JT', 'JWS.U', 'KB', 'KEY', 'K', 'KFFB', 'KFS', 'KIM', 'KINS', 'KKR', 'KMPR', 'KNSL', 'KRC', 'KREF', 'KRG', 'KRNY', 'KW', 'KYN', 'L', 'LACQ', 'LADR', 'LAMR', 'LAND', 'LARK', 'LATN', 'LATNU', 'LAZ', 'LBAI', 'LBC', 'LC', 'LCA', 'LCAHU', 'LCNB', 'LEJU', 'LEVL', 'LFAC', 'LFC', 'LGC', 'LGVW.U', 'LHC', 'LIVKU', 'LKFN', 'LM', 'LMHA', 'LMHB', 'LMRK', 'LMST', 'LNC', 'LOAC', 'LOACU', 'LOAK', 'LOAK.U', 'LOAN', 'LOB', 'LPLA', 'LSAC', 'LSACU', 'LSBK', 'LSI', 'LTC', 'LX', 'LXP', 'LYG', 'LYL', 'MA', 'MAA', 'MAC', 'MAIN', 'MAYS', 'MBCN', 'MBI', 'MBIN', 'MBWM', 'MC', 'MCB', 'MCBC', 'MCBS', 'MCI', 'MCMJ', 'MCY', 'MDJH', 'MDLQ', 'MDLX', 'MDLY', 'MDRR', 'MET', 'MFA', 'MFAC', 'MFAC.U', 'MFC', 'MFG', 'MFIN', 'MFNC', 'MGP', 'MGRC', 'MGYR', 'MHLA', 'MHLD', 'MIC', 'MINI', 'MITT', 'MJCO', 'M', 'MKL', 'MKTX', 'MLP', 'MLVF', 'MMC', 'MMI', 'MN', 'MNCL', 'MNR', 'MNSB', 'MOFG', 'MOGO', 'MORN', 'MPB', 'MPW', 'MRBK', 'MRCC', 'M', 'MRCCL', 'MRLN', 'MS', 'MSBF', 'MSBI', 'MSCI', 'MSVB', 'MTB', 'MTG', 'MUFG', 'MVBF', 'MVC', 'MVCD', 'MYFW', 'NAVI', 'NBAC', 'NBACU', 'NBHC', 'NBN', 'NBTB', 'NCBS', 'NDAQ', 'NEBU', 'NEBUU', 'NEN', 'NFBK', 'NFH', 'NFIN', 'NFINU', 'NGHC', 'NHI', 'NHLD', 'NICK', 'NKLA', 'NKSH', 'NLY', 'NMFC', 'NMIH', 'NMR', 'NMRK', 'NNI', 'NNN', 'NOAH', 'NODK', 'NOVSU', 'NPA', 'NPAUU', 'NREF', 'NRIM', 'NRZ', 'NSA', 'NSCO', 'NSEC', 'NTB', 'NTRS', 'NWBI', 'NWFL', 'NWLI', 'NXRT', 'NYCB',  'NYMT', 'O', 'OAC', 'OAC.U', 'OBNK', 'OCFC', 'OCFT', 'OCN', 'OCSL', 'OFC', 'OFED', 'OFG', 'OFS', 'OFSSZ', 'OGCP', 'OHI', 'OLP', 'OMF', 'ONB', 'ONDK', 'OPBK', 'OPES', 'OPHC', 'OPI', 'OPOF', 'OPRT', 'OPY', 'ORC', 'ORCC', 'ORGO', 'ORI', 'ORRF', 'ORSN', 'ORSNU', 'OSBC', 'OTTW', 'OUT', 'OVBC', 'OVLY', 'OXBR', 'OXSQ', 'OZK', 'PAAC', 'PAACU', 'PACQ', 'PACW', 'PB', 'PBB', 'PBCT', 'PBFS', 'PBHC', 'PBIP', 'PBY', 'PCB', 'PCH', 'PCPL.U', 'PCSB', 'PDLB', 'PDM', 'PEAK', 'PEB', 'PEBK', 'PEBO', 'PEI', 'PFBC', 'PFBI', 'PFG', 'PFHD', 'PFIS', 'PFLT', 'PFS', 'PFSI', 'PGC', 'PGR', 'PGRE', 'PHCF', 'PIC', 'PIC.U', 'PICO', 'PIH', 'PINE', 'PIPR', 'PJT', 'PK', 'PKBK', 'PLBC', 'PLD', 'PLMR', 'PLYM', 'PMBC', 'PMT', 'PNBK', 'PNC', 'PNFP', 'PNNT', 'PPBI', 'PPHI', 'PRA', 'PRI', 'PRK', 'PROS', 'PROV', 'PRS', 'PRU', 'PSA', 'PSB', 'PSEC', 'PSTL', 'PTAC', 'PTACU', 'PTMN', 'PTVCA', 'PTVCB', 'PUB', 'PUK', 'PUYI', 'PVBC', 'PW', 'PWOD', 'PYS', 'PZN', 'QCRH', 'QFIN', 'QK', 'QTS', 'R', 'RAND', 'RBB', 'RBCAA', 'RBKB', 'RBNC', 'RBS', 'RC', 'RCII', 'RDFN', 'RDI', 'RDIB', 'RDN', 'RE', 'REG', 'RESI', 'REXR', 'RF', 'RFL', 'RFM', 'RGA', 'RHE', 'RHP', 'RILY', 'RILYI', 'RIOT', 'RIVE', 'RJF', 'RLGY', 'RLI', 'RLJ', 'RM', 'RMAX', 'RMBI', 'RMG', 'RMI', 'RMR', 'RNDB', 'RNR', 'RNST', 'ROCHU', 'ROIC', 'RPAI', 'RPLA', 'RPLA.U', 'RPT', 'RRBI', 'RVI', 'RVSB', 'RWT', 'RY', 'RYN', 'RZB', 'SACH', 'SAF', 'SAFE', 'SAFT', 'SAL', 'SAMA', 'SAMAU', 'SAMG', 'SAN', 'SAQN', 'SAR', 'SASR', 'SBAC', 'SBBX', 'SBCF', 'SBE', 'SBFG', 'SBNY', 'SBRA', 'SBSI', 'SBT', 'SC', 'SCA', 'SCHW', 'S', 'SCM', 'SCPE', 'SCU', 'SCVX', 'SCVX.U', 'SEIC', 'SELF', 'SF', 'SFBC', 'SFBS', 'SFE', 'SFNC', 'SFST', 'SFTW', 'SFTW.U', 'SG', 'SHBI', 'SHG', 'SHLL', 'SHLL.U', 'SHO', 'SI', 'SIEB', 'SIGI', 'SITC', 'SIVB', 'SKT', 'SLCT', 'SLF', 'SLG', 'SLM', 'SLQT', 'SLRC', 'SMBC', 'SMBK', 'SMFG', 'SMMC', 'SMMCU', 'SMMF', 'SNFCA', 'SNR', 'SNV', 'SOAC.U', 'SOHO', 'SONA', 'SPAQ', 'SPAQ.U', 'SPFI', 'SPG', 'SRAC', 'SRC', 'SRCE', 'SRG', 'SRL', 'SSB', 'SSBI', 'SSPK', 'SSPKU', 'SSSS', 'STAG', 'STAR', 'STBA', 'STC', 'STFC', 'STL', 'STND', 'STOR', 'STRS', 'STT', 'STWD', 'STXB', 'SUI', 'SUNS', 'SUPV', 'SVBI', 'SVC', 'SVVC', 'SYBT', 'SYF', 'TBBK', 'TBK', 'TBNK', 'TCBI', 'TCBK', 'TCF', 'TCFC', 'TCI', 'TCO', 'TCRD', 'TCRZ', 'TD', 'TDAC', 'TDACU', 'TFC', 'TFSL', 'TGH', 'THBR', 'THBRU', 'THCA', 'THCAU', 'THCB', 'THFF', 'THG', 'TIGR', 'TIPT', 'TMP', 'TOTA', 'TOWN', 'TPHS', 'TPRE', 'TPVG', 'TPVY', 'TRC', 'TREE', 'TRMK', 'TRMT', 'TRNE', 'TRNE.U', 'TRNO', 'TROW', 'TRST', 'TRTN', 'TRTX', 'TRUP', 'TRV', 'TSBK', 'TSC', 'TSLX', 'TURN', 'TWO', 'TZAC', 'UBA', 'UBCP', 'UBFO', 'UBOH', 'UBP', 'UBS', 'UBSI', 'UCBI', 'UDR', 'UE', 'UFCS', 'UHAL', 'U', 'UHT', 'UIHC', 'UMBF', 'UMH', 'UMPQ', 'UNAM', 'UNB', 'UNIT', 'UNM', 'UNMA', 'UNTY', 'URI', 'USB', 'USWS', 'UVE', 'UVSP', 'V', 'VBFC', 'VBTX', 'VCTR', 'VEL', 'VER', 'VERY', 'VICI', 'VIRT', 'VLY', 'VNO', 'VOYA', 'VRSK', 'VRTS', 'VTR', 'WABC', 'WAFD', 'WAL', 'WALA', 'WASH', 'WBK', 'WBS', 'WD', 'WDR', 'WELL', 'WETF', 'WF', 'WFC', 'WHF', 'WHFBZ', 'WHG', 'WHLR', 'WINA', 'WINS', 'WLFC', 'WLTW', 'WMC', 'WNEB', 'WPC', 'WPF.U', 'WPG', 'WRB', 'W', 'WRE', 'WRI', 'WRLD', 'WSBC', 'WSBF', 'WSC', 'WSFS', 'WSR', 'WTBA', 'WTFC', 'WTM', 'WTRE', 'WU', 'WVFC', 'WY', 'XAN', 'XHR', 'XIN', 'XYF', 'Y', 'YIN', 'ZGYH', 'ZGYHU', 'ZION', 'ZIONL']
finance_master_file='finance_mf.csv'
finance_directory='finance'
get_ticker_data(finance,finance_directory)
create_master_list(finance,finance_master_file,finance_directory)
print("-> Master file created, filename: %s\n" % finance_master_file)
purge_low_gain_volume(finance_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
health_services = ['ACHC', 'ADUS', 'AIH', 'AME', 'AMEH', 'AMS', 'ANTM', 'APTX', 'ARA', 'BASI', 'BIOC', 'BK', 'CATS', 'CCM', 'CHE', 'CI', 'CNC', 'CO', 'CRDF', 'CRHM', 'CSTL', 'CSU', 'CYH', 'DGX', 'DVA', 'EHC', 'ENSG', 'FLGT', 'FMS', 'FVE', 'GEN', 'GH', 'GHSI', 'GTS', 'HCA', 'HNGR', 'HUM', 'IDXG', 'IMAC', 'IQV', 'JYNT', 'LH', 'LHCG', 'MGLN', 'MOH', 'NEO', 'NHC', 'NRC', 'NTRA', 'NVTA', 'ONCS', 'ONEM', 'OPCH', 'OPTN', 'PDEX', 'PINC', 'PLX', 'PM', 'PNTG', 'RDNT', 'SDC', 'SEM', 'SGRY', 'SSY', 'SYNH', 'TDOC', 'THC', 'TVTY', 'UHS', 'UNH', 'USPH', 'VM', 'XGN']
health_services_master_file='health_services_mf.csv'
health_services_directory='health_services'
get_ticker_data(health_services,health_services_directory)
create_master_list(health_services,health_services_master_file,health_services_directory)
print("-> Master file created, filename: %s\n" % health_services_master_file)
purge_low_gain_volume(health_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
health_technology = ['A', 'ABBV', 'ABEO', 'ABIO', 'ABMD', 'ABT', 'ABUS', 'ACAD', 'ACER', 'ACHV', 'ACIU', 'ACOR', 'ACRS', 'ACRX', 'ACST', 'ADAP', 'ADCT', 'ADIL', 'ADMA', 'ADMP', 'ADMS', 'ADRO', 'ADVM', 'ADXS', 'AEMD', 'AERI', 'AEZS', 'AFMD', 'AGE', 'AGEN', 'AGIO', 'AGLE', 'AGRX', 'AGTC', 'AHPI', 'AIKI', 'AIM', 'AIMT', 'AKBA', 'AKCA', 'AKER', 'AKRO', 'AKTX', 'ALBO', 'ALC', 'ALDX', 'ALEC', 'ALGN', 'ALIM', 'ALKS', 'ALLK', 'ALLO', 'ALNA', 'ALNY', 'ALPN', 'ALRN', 'ALXN', 'AMAG', 'AMGN', 'AMPE', 'AMPH', 'AMRN', 'AMRX', 'ANAB', 'ANCN', 'ANGO', 'ANIK', 'ANIP', 'ANIX', 'ANPC', 'ANVS', 'APEN', 'APLS', 'APLT', 'APM', 'APOP', 'APRE', 'APT', 'APTO', 'APVO', 'APYX', 'AQB', 'AQST', 'ARAV', 'ARAY', 'ARCT', 'ARDS', 'ARDX', 'ARGX', 'ARMP', 'ARNA', 'ARPO', 'ARQT', 'ARVN', 'ARWR', 'ASLN', 'ASMB', 'ASND', 'ASRT', 'ATEC', 'ATHE', 'ATHX', 'ATNM', 'ATNX', 'ATOS', 'ATRA', 'ATRC', 'ATRI', 'ATRS', 'ATXI', 'AUPH', 'AUTL', 'AVCO', 'AVDL', 'AVEO', 'AVGR', 'AVNS', 'AVRO', 'AVXL', 'AXDX', 'AXGN', 'AXGT', 'AXLA', 'AXNX', 'AXSM', 'AYLA', 'AZN', 'AZRX', 'BAX', 'BBI', 'BBIO', 'BCDA', 'BCLI', 'BCRX', 'BCYC', 'BDSI', 'BDTX', 'BDX', 'BEAM', 'BEAT', 'BFRA', 'BGNE', 'BHC', 'BHVN', 'BIIB', 'BIO', 'BIOL', 'BLCM', 'BLFS', 'BLPH', 'BLRX', 'BLU', 'BLUE', 'BMRA', 'BMRN', 'BMY', 'BNGO', 'BNTC', 'BNTX', 'BPMC', 'BPTH', 'BRBR', 'BRKR', 'BSGM', 'BSTC', 'BSX', 'BTAI', 'BVXV', 'BWAY', 'BXRX', 'BYSI', 'CABA', 'CALA', 'CANF', 'CAPR', 'CARA', 'CASI', 'CATB', 'CBAY', 'CBIO', 'CBLI', 'CBMG', 'CBPO', 'CCXI', 'CDMO', 'CDNA', 'CDTX', 'CEMI', 'CERC', 'CERS', 'CFMS', 'CFRX', 'CGEN', 'CGIX', 'CHEK', 'CHFS', 'CHMA', 'CHRS', 'CKPT', 'CLDX', 'CLGN', 'CLLS', 'CLPT', 'CLRB', 'CLSD', 'CLSN', 'CLVS', 'CMD', 'CMRX', 'CNCE', 'CNMD', 'CNSP', 'CNST', 'COCP', 'CODX', 'COLL', 'COO', 'CORT', 'CPHI', 'CPIX', 'CPRX', 'CRBP', 'CRIS', 'CRMD', 'CRNX', 'CRSP', 'CRTX', 'CRVS', 'CRY', 'CSBR', 'CSII', 'CTIC', 'CTLT', 'CTMX', 'CTSO', 'CTXR', 'CUE', 'CUTR', 'CVM', 'CWBR', 'CYAD', 'CYAN', 'CYCC', 'CYCN', 'CYTK', 'DARE', 'DBVT', 'DCPH', 'DCTH', 'DFFN', 'DHR', 'DMAC', 'DMTK', 'DNLI', 'DRAD', 'DRIO', 'DRNA', 'DRRX', 'DTIL', 'DVAX', 'DXCM', 'DXR', 'DYAI', 'DYNT', 'EARS', 'EBS', 'ECOR', 'EDAP', 'EDIT', 'EDSA', 'EGRX', 'EIDX', 'EIGR', 'EKSO', 'ELGX', 'ELMD', 'ELOX', 'ENDP', 'ENLV', 'ENTA', 'ENTX', 'ENZ', 'EOLS', 'EPIX', 'EPZM', 'EQ', 'ERYP', 'ESPR', 'ESTA', 'ETNB', 'ETON', 'ETTX', 'EVFM', 'EVGN', 'EVLO', 'EVOK', 'EW', 'EXAS', 'EXEL', 'EYEG', 'EYEN', 'EYES', 'EYPT', 'FATE', 'FBIO', 'FENC', 'FGEN', 'FIXX', 'FLDM', 'FLXN', 'FOLD', 'FONR', 'FPRX', 'FREQ', 'FULC', 'FWP', 'GALT', 'GBT', 'GENE', 'GERN', 'GILD', 'GKOS', 'GLMD', 'GLPG', 'GLYC', 'GMAB', 'GMDA', 'GMED', 'GNCA', 'GNFT', 'GNMK', 'GNPX', 'GOSS', 'GRFS', 'GRTS', 'GRTX', 'GSK', 'GTHX', 'GWPH', 'HAE', 'HALO', 'HAPP', 'HARP', 'HBIO', 'HCM', 'HEPA', 'HJLI', 'HOLX', 'HOOK', 'HOTH', 'HRC', 'HROW', 'HRTX', 'HSDT', 'HSKA', 'HSTO', 'HTBX', 'HTGM', 'HZNP', 'IART', 'IBIO', 'ICCC', 'ICLR', 'ICPT', 'ICUI', 'IDRA', 'IDXX', 'IDYA', 'IFRX', 'IGC', 'IGMS', 'IIN', 'ILMN', 'IMAB', 'IMGN', 'IMMP', 'IMMU', 'IMRA', 'IMRN', 'IMUX', 'IMV', 'INCY', 'INFI', 'INFU', 'INGN', 'INMB', 'INMD', 'INO', 'INSM', 'INSP', 'INVA', 'IONS', 'IOVA', 'IPHA', 'IRIX', 'IRMD', 'IRTC', 'IRWD', 'ISEE', 'ISR', 'ISRG', 'ITCI', 'ITMR', 'ITRM', 'IVC', 'JAGX', 'JAZZ', 'JNCE', 'JNJ', 'KALA', 'KALV', 'KDMN', 'KIDS', 'KIN', 'KLDO', 'KMDA', 'KAMA', 'KNSA', 'KOD', 'KPTI', 'KRMD', 'KRTX', 'KRYS', 'KTOV', 'KURA', 'KZIA', 'KZR', 'LCI', 'LCTX', 'LFVN', 'LGND', 'LIFE', 'LIVN', 'LJPC', 'LLIT', 'LLY', 'LMAT', 'LMNL', 'LMNX', 'LNTH', 'LOGC', 'LPCN', 'LPTX', 'LQDA', 'LUMO', 'LXRX', 'LYRA', 'MACK', 'MASI', 'MBIO', 'MBOT', 'MBRX', 'MCRB', 'MDGL', 'MDGS', 'MDT', 'MDWD', 'MEIP', 'MESO', 'MGEN', 'MGNX', 'MGTA', 'MGTX', 'MIRM', 'MIST', 'MITO', 'MLAB', 'MLND', 'MLSS', 'MMSI', 'MNK', 'MNKD', 'MNLO', 'MNOV', 'MNPR', 'MNTA', 'MOR', 'MORF', 'MOTS', 'MREO', 'MRK', 'MRNA', 'MRNS', 'MRSN', 'MRTX', 'MRUS', 'MSON', 'MTD', 'MTEM', 'MTEX', 'MTNB', 'MTP', 'MYGN', 'MYL', 'MYOK', 'MYOS', 'MYOV', 'NAII', 'NAOV', 'NARI', 'NAVB', 'NBIX', 'NBRV', 'NBSE', 'NBY', 'NCNA', 'NDRA', 'NEOG', 'NEOS', 'NEPH', 'NEPT', 'NERV', 'NGM', 'NK', 'NKTR', 'NLTX', 'NMRD', 'NMTR', 'NNVC', 'NOVN', 'NRBO', 'NSPR', 'NSTG', 'NTEC', 'NTLA', 'NTRP', 'NTUS', 'NURO', 'NUVA', 'NVAX', 'NVCN', 'NVCR', 'NVIV', 'NVO', 'NVRO', 'NVS', 'NVST', 'NVUS', 'NXTC', 'NYMX', 'OBLN', 'OBSV', 'OCGN', 'OCUL', 'OCX', 'ODT', 'OFIX', 'OGEN', 'OMER', 'ONCT', 'ONCY', 'ONTX', 'ONVO', 'OPK', 'OPNT', 'ORGS', 'ORIC', 'ORMP', 'ORTX', 'OSMT', 'OSUR', 'OTIC', 'OTLK', 'OVID', 'OXFD', 'OYST', 'PACB', 'PAHC', 'PASG', 'PAVM', 'PBH', 'PBYI', 'PCRX', 'PDLI', 'PDSB', 'PEN', 'PETQ', 'PFE', 'PFNX', 'PGEN', 'PGNX', 'PHAS', 'PHAT', 'PHG', 'PHGE', 'PHGE.U', 'PHIO', 'PIRS', 'PKI', 'PLRX', 'PLSE', 'PLXP', 'POAI', 'PODD', 'PRGO', 'PRNB', 'PROF', 'PRPH', 'PRPO', 'PRQR', 'PRTA', 'PRTK', 'PRVB', 'PRVL', 'PSNL', 'PSTI', 'PSTV', 'PTCT', 'PTGX', 'PTI', 'PTLA', 'PTN', 'PULM', 'QDEL', 'QGEN', 'QLGN', 'QTNT', 'QTRX', 'QURE', 'RAPT', 'RARE', 'RCEL', 'RCKT', 'RCUS', 'RDHL', 'RDUS', 'RDY', 'REGN', 'REPH', 'REPL', 'RETA', 'REXN', 'RGEN', 'RGLS', 'RGNX', 'RIGL', 'RLMD', 'RMD', 'RMED', 'RMTI', 'RTIX', 'RTRX', 'RUBY', 'RVMD', 'RVNC', 'RVP', 'RWLK', 'RYTM', 'SAGE', 'SAVA', 'SBBP', 'SBPH', 'SCPH', 'SCYX', 'SEEL', 'SELB', 'SESN', 'SGEN', 'SGMO', 'SIBN', 'SIEN', 'SIGA', 'SILK', 'SINT', 'SLDB', 'SLGL', 'SLNO', 'SLS', 'SMMT', 'SNCA', 'SNDX', 'SNES', 'SNGX', 'SNN', 'SNOA', 'SNSS', 'SNY', 'SOLY', 'SONN', 'SPNE', 'SPPI', 'SPRO', 'SRDX', 'SRNE', 'SRPT', 'SRRA', 'SRRK', 'SRTS', 'SSKN', 'STAA', 'STE', 'STIM', 'STML', 'STOK', 'STRO', 'STSA', 'STXS', 'SUPN', 'SURF', 'SVRA', 'SWAV', 'SWTX', 'SXTC', 'SYBX', 'SYK', 'SYN', 'SYRS', 'TAK', 'TARA', 'TARO', 'TBIO', 'TBPH', 'TCDA', 'TCMD', 'TCON', 'TCRR', 'TECH', 'TELA', 'TENX', 'TEVA', 'TFFP', 'TFX', 'TGTX', 'THMO', 'THTX', 'TLC', 'TLGT', 'TLSA', 'TMBR', 'TMDI', 'TMDX', 'TMO', 'TNDM', 'TNXP', 'TOCA', 'TORC', 'TPTX', 'TRIB', 'TRIL', 'TRPX', 'TRVI', 'TRVN', 'TTNP', 'TTOO', 'TTPH', 'TWST', 'TXG', 'TXMD', 'TYME', 'UBX', 'UMRX', 'URGN', 'UROV', 'UTHR', 'UTMD', 'VAPO', 'VAR', 'VBIV', 'VBLT', 'VCEL', 'VCNX', 'VCYT', 'VERO', 'VERU', 'VIE', 'VIR', 'VIVE', 'VIVO', 'VKTX', 'VNDA', 'VNRX', 'VRAY', 'VRCA', 'VREX', 'VRML', 'VRNA', 'VRTX', 'VSTM', 'VTGN', 'VTVT', 'VVUS', 'VXRT', 'VYGR', 'WAT', 'WINT', 'WMGI', 'WST', 'WVE', 'XBIO', 'XBIT', 'XENE', 'XENT', 'XERS', 'XFOR', 'XLRN', 'XNCR', 'XOMA', 'XRAY', 'XTLB', 'XTNT', 'XXII', 'YMAB', 'ZBH', 'ZEAL', 'ZGNX', 'ZIOP', 'ZLAB', 'ZNTL', 'ZSAN', 'ZTS', 'ZYME', 'ZYNE', 'ZYXI']
health_technology_master_file='health_technology_mf.csv'
health_technology_directory='health_technology'
get_ticker_data(health_technology,health_technology_directory)
create_master_list(health_technology,health_technology_master_file,health_technology_directory)
print("-> Master file created, filename: %s\n" % health_technology_master_file)
purge_low_gain_volume(health_technology_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
industrial_services = ['ACM', 'ADSW', 'AEGN', 'AGX', 'ALTM', 'AM', 'AMRC', 'AQUA', 'AROC', 'AWX', 'BKEP', 'BKR', 'BORR', 'BPMP', 'CCLP', 'CEQP', 'CHRA', 'CKH', 'CLB', 'CLH', 'CNXM', 'CQP', 'CVA', 'CWST', 'DCP', 'DKL', 'DOV', 'DRQ', 'DY', 'ECOL', 'EME', 'ENB', 'ENBA', 'ENBL', 'ENG', 'ENLC', 'EPD', 'EQM', 'ET', 'ETRN', 'EXPO', 'EXTN', 'FET', 'FI', 'FIX', 'FLR', 'FTEK', 'FTI', 'FTK', 'FTSI', 'GEL', 'GEOS', 'GLDD', 'GRAM', 'GV', 'GVA', 'HAL', 'HCCI', 'HDSN', 'HEP', 'HIL', 'HLX', 'HP', 'ICD', 'IEA', 'IESC', 'J', 'JAN', 'KBR', 'KLXE', 'KMI', 'LBRT', 'LNG', 'MG', 'MIND', 'MMLP', 'MMP', 'MPLX', 'MTRX', 'MTZ', 'MYRG', 'NBLX', 'NBR', 'NCSM', 'NE', 'NESR', 'NEX', 'NFE', 'NGL', 'NGS', 'NINE', 'NOA', 'NOV', 'NR', 'OII', 'OIS', 'OKE', 'OMP', 'ORN', 'PAA', 'PACD', 'PAGP', 'PBA', 'PBFX', 'PDS', 'PECK', 'PED', 'PESI', 'PRIM', 'PSXP', 'PTEN', 'PUMP', 'PWR', 'QES', 'QRHC', 'RCMT', 'RES', 'RIG', 'RNGR', 'ROAD', 'RSG', 'RTLR', 'SAEX', 'SDPI', 'SDRL', 'SHLX', 'SLB', 'SMED', 'SOI', 'SPN', 'SRCL', 'SRLP', 'STN', 'STRL', 'TCP', 'TDW', 'TPC', 'TRP', 'TTEK', 'TTI', 'TUSK', 'USAC', 'USDP', 'VAL', 'VSEC', 'VTNR', 'WCN', 'WHD', 'WM', 'WMB', 'WTTR']
industrial_services_master_file='industrial_services_mf.csv'
industrial_services_directory='industrial_services'
get_ticker_data(industrial_services,industrial_services_directory)
create_master_list(industrial_services,industrial_services_master_file,industrial_services_directory)
print("-> Master file created, filename: %s\n" % industrial_services_master_file)
purge_low_gain_volume(industrial_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
miscellaneous = ['AAAU', 'AADR', 'AAXJ', 'ABEQ', 'ACES', 'ACIO', 'ACP', 'ACSG', 'ACSI', 'ACT', 'ACV', 'ACWF', 'ACWI', 'ACWV', 'ACWX', 'ADME', 'ADRE', 'ADX', 'AEF', 'AESR', 'AFB', 'AFIF', 'AFK', 'AFLG', 'AFMC', 'AFSM', 'AFT', 'AFTY', 'AGD', 'AGG', 'AGGP', 'AGGY', 'AGQ', 'AGT', 'AGZ', 'AGZD', 'AIA', 'AIEQ', 'AIF', 'AIIQ', 'AIO', 'AIQ', 'AIRR', 'ALFA', 'ALTS', 'ALTY', 'AMCA', 'AMJ', 'AMLP', 'AMOM', 'AMU', 'AMUB', 'AMZA', 'ANGL', 'AOA', 'AOD', 'AOK', 'AOM', 'AOR', 'ARB', 'ARCM', 'ARDC', 'ARGT', 'ARKF', 'ARKG', 'ARKK', 'ARKQ', 'ARKW', 'ARMR', 'ASA', 'ASEA', 'ASET', 'ASG', 'ASHR', 'ASHS', 'ASHX', 'ATMP', 'AUSF', 'AVDE', 'AVDV', 'AVEM', 'AVK', 'AVUS', 'AVUV', 'AWAY', 'AWF', 'AWP', 'AWTM', 'AZAA', 'AZBA', 'BAB', 'BAF', 'BAL', 'BANX', 'BAPR', 'BAR', 'BATT', 'BAUG', 'BBAX', 'BBC', 'BBCA', 'BBEU', 'BBF', 'BBH', 'BBIN', 'BBJP', 'BBK', 'BBMC', 'BBN', 'BBP', 'BBRE', 'BBSA', 'BBUS', 'BCD', 'BCI', 'BCM', 'BCSF', 'BCV', 'BCX', 'BDCS', 'BDCX', 'BDCY', 'BDCZ', 'BDEC', 'BDJ', 'BDRY', 'BFEB', 'BFIT', 'BFK', 'BFO', 'BFOR', 'BFY', 'BFZ', 'BGB', 'BGH', 'BGIO', 'BGR', 'BGRN', 'BGT', 'BGX', 'BGY', 'BHK', 'BHV', 'BIB', 'BIBL', 'BICK', 'BIF', 'BIL', 'BIS', 'BIT', 'BIV', 'BIZD', 'BJAN', 'BJK', 'BJUL', 'BJUN', 'BKAG', 'BKEM', 'BKF', 'BKHY', 'BKIE', 'BKK', 'BKLC', 'BKLN', 'BKMC', 'BKN', 'BKSB', 'BKSE', 'BKT', 'BLCN', 'BLE', 'BLES', 'BLHY', 'BLOK', 'BLV', 'BLW', 'BMAR', 'BMAY', 'BME', 'BMEZ', 'BMLP', 'BND', 'BNDC', 'BNDW', 'BNDX', 'BNKD', 'BNKO', 'BNKU', 'BNKZ', 'BNO', 'BNOV', 'BNY', 'BOCT', 'BOE', 'BOIL', 'BOND', 'BOSS', 'BOTZ', 'BOUT', 'BPT', 'BQH', 'BRF', 'BRZU', 'BSAE', 'BSBE', 'BSCE', 'BSCK', 'BSCL', 'BSCM', 'BSCN', 'BSCO', 'BSCP', 'BSCQ', 'BSCR', 'BSCS', 'BSCT', 'BSD', 'BSDE', 'BSE', 'BSEP', 'BSJK', 'BSJL', 'BSJM', 'BSJN', 'BSJO', 'BSJP', 'BSJQ', 'BSJR', 'BSL', 'BSML', 'BSMM', 'BSMN', 'BSMO', 'BSMP', 'BSMQ', 'BSMR', 'BSMS', 'BSMT', 'BST', 'BSTZ', 'BSV', 'BTA', 'BTAL', 'BTEC', 'BTO', 'BTT', 'BTYS', 'BTZ', 'BUG', 'BUI', 'BUL', 'BUY', 'BUYZ', 'BVAL', 'BWG', 'BWX', 'BWZ', 'BXMX', 'BYLD', 'BYM', 'BZM', 'BZQ', 'CACG', 'CAF', 'CALF', 'CANE', 'CAPE', 'CARZ', 'CATH', 'CBH', 'CBON', 'CCD', 'CCNC', 'CCOR', 'CDC', 'CDL', 'CEE', 'CEF', 'CEFD', 'CEFS', 'CEM', 'CEMB', 'CEN', 'CET', 'CEV', 'CEW', 'CEY', 'CEZ', 'CFA', 'CFCV', 'CFO', 'CGO', 'CGW', 'CHAD', 'CHAU', 'CHEP', 'CHGX', 'CHI', 'CHIC', 'CHIE', 'CHIH', 'CHII', 'CHIK', 'CHIL', 'CHIM', 'CHIQ', 'CHIR', 'CHIS', 'CHIU', 'CHIX', 'CHN', 'CHNA', 'CHW', 'CHY', 'CIBR', 'CID', 'CIF', 'CII', 'CIK', 'CIL', 'CIZ', 'CKX', 'CLIX', 'CLM', 'CLOU', 'CLRG', 'CLSK', 'CLTL', 'CMBS', 'CMDY', 'CMF', 'CMU', 'CN', 'CNBS', 'CNCR', 'CNRG', 'CNXT', 'CNYA', 'COM', 'COMB', 'COMT', 'COPX', 'CORN', 'CORP', 'COW', 'COWZ', 'CPER', 'CPI', 'CPZ', 'CQQQ', 'CRAK', 'CRBN', 'CRF', 'CROC', 'CROP', 'CSA', 'CSB', 'CSD', 'CSF', 'CSM', 'CSML', 'CSQ', 'CTR', 'CUBA', 'CURE', 'CUT', 'CVY', 'CWB', 'CWEB', 'CWI', 'CWS', 'CXE', 'CXH', 'CXSE', 'CYB', 'CZA', 'DALI', 'DALT', 'DAUD', 'DAUG', 'DAX', 'DBA', 'DBAW', 'DBB', 'DBC', 'DBE', 'DBEF', 'DBEH', 'DBEM', 'DBEU', 'DBEZ', 'DBGR', 'DBJP', 'DBL', 'DBLV', 'DBMF', 'DBO', 'DBP', 'DBS', 'DBV', 'DCF', 'DCHF', 'DDF', 'DDG', 'DDIV', 'DDLS', 'DDM', 'DDWM', 'DEED', 'DEEF', 'DEF', 'DEFA', 'DEM', 'DES', 'DEUR', 'DEUS', 'DEW', 'DEX', 'DFE', 'DFEB', 'DFEN', 'DFJ', 'DFND', 'DFNL', 'DFP', 'DFVL', 'DFVS', 'DGAZ', 'DGBP', 'DGL', 'DGLD', 'DGP', 'DGRE', 'DGRO', 'DGRS', 'DGRW', 'DGS', 'DGT', 'DGZ', 'DHF', 'DHS', 'DHY', 'DIA', 'DIAL', 'DIAX', 'DIG', 'DIM', 'DINT', 'DIV', 'DIVA', 'DIVB', 'DIVC', 'DIVO', 'DIVY', 'DJCB', 'DJCI', 'DJD', 'DJP', 'DJPY', 'DLN', 'DLS', 'DLY', 'DMAY', 'DMB', 'DMDV', 'DMF', 'DMO', 'DMRE', 'DMRI', 'DMRL', 'DMRM', 'DMRS', 'DNI', 'DNL', 'DNOV', 'DNP', 'DOG', 'DOL', 'DON', 'DOO', 'DPG', 'DPST', 'DRIP', 'DRIV', 'DRN', 'DRSK', 'DRV', 'DRW', 'DSE', 'DSI', 'DSL', 'DSLV', 'DSM', 'DSTL', 'DSU', 'DTD', 'DTEC', 'DTF', 'DTH', 'DTN', 'DTUL', 'DTUS', 'DTYL', 'DUC', 'DUG', 'DURA', 'DUSA', 'DUSL', 'DUST', 'DVLU', 'DVOL', 'DVP', 'DVY', 'DVYA', 'DVYE', 'DWAS', 'DWAT', 'DWAW', 'DWCR', 'DWEQ', 'DWFI', 'DWLD', 'DWM', 'DWMC', 'DWMF', 'DWPP', 'DWSH', 'DWUS', 'DWX', 'DXD', 'DXGE', 'DXJ', 'DXJS', 'DYNF', 'DZK', 'DZZ', 'EAD', 'EAGG', 'EASG', 'EASI', 'EBIZ', 'EBND', 'ECC', 'ECCX', 'ECCY', 'ECF', 'ECH', 'ECLN', 'ECNS', 'ECON', 'ECOW', 'ECOZ', 'EDC', 'EDD', 'EDEN', 'EDF', 'EDI', 'EDIV', 'EDOG', 'EDOW', 'EDV', 'EDZ', 'EEA', 'EEH', 'EELV', 'EEM', 'EEMA', 'EEMD', 'EEMO', 'EEMS', 'EEMV', 'EEMX', 'EES', 'EET', 'EEV', 'EFA', 'EFAD', 'EFAS', 'EFAV', 'EFAX', 'EFF', 'EFG', 'EFL', 'EFNL', 'EFO', 'EFR', 'EFT', 'EFU', 'EFV', 'EFZ', 'EGF', 'EGIF', 'EGPT', 'EHI', 'EHT', 'EIC', 'EIDO', 'EIM', 'EINC', 'EIRL', 'EIS', 'EJAN', 'EJUL', 'EKAR', 'ELD', 'EMAG', 'EMB', 'EMBD', 'EMBH', 'EMCB', 'EMD', 'EMDV', 'EMF', 'EMFM', 'EMGF', 'EMHY', 'EMIF', 'EMLC', 'EMLP', 'EMMF', 'EMNT', 'EMO', 'EMQQ', 'EMSG', 'EMSH', 'EMTL', 'EMTY', 'EMXC', 'ENFR', 'ENOR', 'ENTR', 'ENX', 'ENZL', 'EOD', 'EOI', 'EOS', 'EOT', 'EPHE', 'EPI', 'EPOL', 'EPP', 'EPRF', 'EPS', 'EPU', 'EPV', 'EQAL', 'EQL', 'EQRR', 'EQWL', 'ERC', 'ERH', 'ERM', 'ERSX', 'ERUS', 'ERX', 'ERY', 'ESCR', 'ESEB', 'ESG', 'ESGD', 'ESGE', 'ESGG', 'ESGN', 'ESGS', 'ESGU', 'ESGV', 'ESHY', 'ESML', 'ESNG', 'ESPO', 'ETB', 'ETG', 'ETHO', 'ETJ', 'ETO', 'ETV', 'ETW', 'ETX', 'ETY', 'EUDG', 'EUDV', 'EUFN', 'EUFX', 'EUM', 'EUMV', 'EUO', 'EURL', 'EURZ', 'EUSA', 'EUSC', 'EVF', 'EVG', 'EVM', 'EVN', 'EVSTC', 'EVT', 'EVV', 'EVX', 'EVY', 'EWA', 'EWC', 'EWCO', 'EWD', 'EWG', 'EWGS', 'EWH', 'EWI', 'EWJ', 'EWJE', 'EWJV', 'EWK', 'EWL', 'EWM', 'EWMC', 'EWN', 'EWO', 'EWP', 'EWQ', 'EWRE', 'EWS', 'EWSC', 'EWT', 'EWU', 'EWUS', 'EWV', 'EWW', 'EWX', 'EWY', 'EWZ', 'EWZS', 'EXD', 'EXG', 'EXI', 'EYLD', 'EZA', 'EZJ', 'EZM', 'EZU', 'FAAR', 'FAB', 'FAD', 'FALN', 'FAM', 'FAN', 'FAS', 'FAUG', 'FAUS', 'FAX', 'FAZ', 'FBGX', 'FBND', 'FBT', 'FBZ', 'FCA', 'FCAL', 'FCAN', 'FCEF', 'FCG', 'FCO', 'FCOM', 'FCOR', 'FCPI', 'FCT', 'FCTR', 'FCVT', 'FDD', 'FDEM', 'FDEU', 'FDEV', 'FDG', 'FDHY', 'FDIS', 'FDIV', 'FDL', 'FDLO', 'FDM', 'FDMO', 'FDN', 'FDNI', 'FDRR', 'FDT', 'FDTS', 'FDVV', 'FEI', 'FEM', 'FEMB', 'FEMS', 'FEN', 'FENY', 'FEO', 'FEP', 'FEUL', 'FEUZ', 'FEX', 'FEZ', 'FFA', 'FFC', 'FFEB', 'FFEU', 'FFHG', 'FFIU', 'FFR', 'FFSG', 'FFTG', 'FFTI', 'FFTY', 'FGB', 'FGD', 'FGM', 'FHK', 'FHLC', 'FIBR', 'FID', 'FIDI', 'FIDU', 'FIEE', 'FIF', 'FIHD', 'FILL', 'FINX', 'FISR', 'FITE', 'FIV', 'FIVA', 'FIVG', 'FIW', 'FIXD', 'FIYY', 'FJNK', 'FJP', 'FKO', 'FKU', 'FLAT', 'FLAU', 'FLAX', 'FLBL', 'FLBR', 'FLC', 'FLCA', 'FLCB', 'FLCH', 'FLCO', 'FLDR', 'FLEE', 'FLEH', 'FLEU', 'FLFR', 'FLGB', 'FLGE', 'FLGR', 'FLHK', 'FLHY', 'FLIA', 'FLIN', 'FLIY', 'FLJH', 'FLJP', 'FLKR', 'FLLA', 'FLLV', 'FLM', 'FLMB', 'FLMI', 'FLMX', 'FLN', 'FLOT', 'FLQD', 'FLQE', 'FLQG', 'FLQH', 'FLQL', 'FLQM', 'FLQS', 'FLRN', 'FLRT', 'FLRU', 'FLSA', 'FLSP', 'FLSW', 'FLTB', 'FLTR', 'FLTW', 'FLV', 'FLYT', 'FLZA', 'FM', 'FMAT', 'FMAY', 'FMB', 'FMF', 'FMHI', 'FMK', 'FMN', 'FMO', 'FMY', 'FNCL', 'FNDA', 'FNDB', 'FNDC', 'FNDE', 'FNDF', 'FNDX', 'FNGD', 'FNGO', 'FNGS', 'FNGU', 'FNGZ', 'FNI', 'FNK', 'FNOV', 'FNX', 'FNY', 'FOF', 'FOVL', 'FPA', 'FPE', 'FPEI', 'FPF', 'FPL', 'FPX', 'FPXE', 'FPXI', 'FQAL', 'FRA', 'FRAK', 'FRDM', 'FREL', 'FRI', 'FRLG', 'FSD', 'FSLF', 'FSMB', 'FSMD', 'FSTA', 'FSZ', 'FT', 'FTA', 'FTAG', 'FTC', 'FTCS', 'FTEC', 'FTF', 'FTGC', 'FTHI', 'FTLB', 'FTLS', 'FTRI', 'FTSD', 'FTSL', 'FTSM', 'FTXD', 'FTXG', 'FTXH', 'FTXL', 'FTXN', 'FTXO', 'FTXR', 'FUD', 'FUE', 'FUMB', 'FUND', 'FUT', 'FUTY', 'FV', 'FVAL', 'FVC', 'FVD', 'FVL', 'FWDB', 'FXA', 'FXB', 'FXC', 'FXD', 'FXE', 'FXF', 'FXG', 'FXH', 'FXI', 'FXL', 'FXN', 'FXO', 'FXP', 'FXR', 'FXU', 'FXY', 'FXZ', 'FYC', 'FYLD', 'FYT', 'FYX', 'GAA', 'GAB', 'GAL', 'GAM', 'GAMR', 'GARS', 'GAZ', 'GBAB', 'GBDV', 'GBF', 'GBIL', 'GBUG', 'GBUY', 'GCC', 'GCE', 'GCOW', 'GCV', 'GDAT', 'GDL', 'GDMA', 'GDNA', 'GDO', 'GDV', 'GDVD', 'GDX', 'GDXJ', 'GEM', 'GENY', 'GER', 'GF', 'GFIN', 'GFY', 'GGM', 'GGN', 'GGO', 'GGT', 'GGZ', 'GHY', 'GHYB', 'GHYG', 'GIGB', 'GIGE', 'GII', 'GIM', 'GLCN', 'GLD', 'GLDI', 'GLDM', 'GLIF', 'GLIN', 'GLL', 'GLO', 'GLQ', 'GLTR', 'GLU', 'GLV', 'GMAN', 'GMF', 'GMOM', 'GMZ', 'GNAF', 'GNMA', 'GNOM', 'GNR', 'GNT', 'GOAT', 'GOAU', 'GOEX', 'GOF', 'GOVT', 'GPM', 'GQRE', 'GREK', 'GRES', 'GRF', 'GRID', 'GRN', 'GRNB', 'GRU', 'GRX', 'GSC', 'GSEE', 'GSEU', 'GSEW', 'GSG', 'GSID', 'GSIE', 'GSJY', 'GSLC', 'GSP', 'GSSC', 'GSST', 'GSUS', 'GSY', 'GTIP', 'GTO', 'GUDB', 'GUNR', 'GURU', 'GUSH', 'GUT', 'GVAL', 'GVI', 'GVIP', 'GWX', 'GXC', 'GXF', 'GXG', 'GXTG', 'GYLD', 'HACK', 'HAIL', 'HAP', 'HAUD', 'HAUZ', 'HAWX', 'HCAP', 'HCRB', 'HDAW', 'HDEF', 'HDG', 'HDGE', 'HDIV', 'HDLB', 'HDMV', 'HDV', 'HEDJ', 'HEEM', 'HEFA', 'HELX', 'HEQ', 'HERD', 'HERO', 'HEWC', 'HEWG', 'HEWI', 'HEWJ', 'HEWL', 'HEWP', 'HEWU', 'HEWW', 'HEWY', 'HEZU', 'HFRO', 'HFXE', 'HFXI', 'HFXJ', 'HGLB', 'HIBL', 'HIBS', 'HIE', 'HIO', 'HIPS', 'HIX', 'HJPX', 'HLAL', 'HMOP', 'HNDL', 'HNW', 'HOLD', 'HOMZ', 'HPF', 'HPI', 'HPS', 'HQH', 'HQL', 'HRZN', 'HSCZ', 'HSMV', 'HSPX', 'HSRT', 'HTAB', 'HTD', 'HTEC', 'HTFA', 'HTRB', 'HTUS', 'HTY', 'HUSV', 'HYB', 'HYD', 'HYDB', 'HYDW', 'HYEM', 'HYG', 'HYGH', 'HYGV', 'HYHG', 'HYI', 'HYLB', 'HYLD', 'HYLS', 'HYLV', 'HYMB', 'HYS', 'HYT', 'HYTR', 'HYUP', 'HYXE', 'HYXU', 'HYZD', 'IAE', 'IAF', 'IAGG', 'IAI', 'IAK', 'IAT', 'IAU', 'IAUF', 'IBB', 'IBCE', 'IBD', 'IBDD', 'IBDL', 'IBDM', 'IBDN', 'IBDO', 'IBDP', 'IBDQ', 'IBDR', 'IBDS', 'IBDT', 'IBDU', 'IBHA', 'IBHB', 'IBHC', 'IBHD', 'IBHE', 'IBMI', 'IBMJ', 'IBMK', 'IBML', 'IBMM', 'IBMN', 'IBMO', 'IBMP', 'IBMQ', 'IBND', 'IBTA', 'IBTB', 'IBTD', 'IBTE', 'IBTF', 'IBTG', 'IBTH', 'IBTI', 'IBTJ', 'IBUY', 'ICF', 'ICLN', 'ICOL', 'ICOW', 'ICSH', 'ICVT', 'IDE', 'IDEV', 'IDHD', 'IDHQ', 'IDIV', 'IDLB', 'IDLV', 'IDMO', 'IDNA', 'IDOG', 'IDRV', 'IDU', 'IDV', 'IDX', 'IDY', 'IECS', 'IEDI', 'IEF', 'IEFA', 'IEFN', 'IEHS', 'IEI', 'IEIH', 'IEME', 'IEMG', 'IEO', 'IETC', 'IEUR', 'IEUS', 'IEV', 'IEZ', 'IFEU', 'IFGL', 'IFN', 'IFRA', 'IFV', 'IG', 'IGA', 'IGBH', 'IGD', 'IGE', 'IGEB', 'IGF', 'IGHG', 'IGI', 'IGIB', 'IGLB', 'IGM', 'IGN', 'IGOV', 'IGR', 'IGRO', 'IGSB', 'IGV', 'IHAK', 'IHD', 'IHDG', 'IHE', 'IHF', 'IHI', 'IHIT', 'IHTA', 'IHY', 'IID', 'IIF', 'IIGD', 'IIGV', 'IIM', 'IJAN', 'IJH', 'IJJ', 'IJK', 'IJR', 'IJS', 'IJT', 'IJUL', 'ILF', 'ILTB', 'IMLP', 'IMOM', 'IMTB', 'IMTM', 'INCO', 'INDA', 'INDL', 'INDS', 'INDY', 'INFR', 'INKM', 'INSI', 'INTF', 'IOO', 'IPAC', 'IPAY', 'IPFF', 'IPKW', 'IPO', 'IPOS', 'IQDE', 'IQDF', 'IQDG', 'IQDY', 'IQI', 'IQIN', 'IQLT', 'IQM', 'IQSI', 'IQSU', 'IRBO', 'IRL', 'IRR', 'ISCF', 'ISD', 'ISDS', 'ISDX', 'ISEM', 'ISHG', 'ISMD', 'ISRA', 'ISTB', 'ISZE', 'ITA', 'ITB', 'ITEQ', 'ITM', 'ITOT', 'IUS', 'IUSB', 'IUSG', 'IUSS', 'IUSV', 'IVAL', 'IVE', 'IVES', 'IVH', 'IVLU', 'IVOG', 'IVOL', 'IVOO', 'IVOV', 'IVV', 'IVW', 'IWB', 'IWC', 'IWD', 'IWF', 'IWL', 'IWM', 'IWN', 'IWO', 'IWP', 'IWR', 'IWS', 'IWV', 'IWX', 'IWY', 'IXC', 'IXG', 'IXJ', 'IXN', 'IXP', 'IXSE', 'IXUS', 'IYC', 'IYE', 'IYF', 'IYG', 'IYH', 'IYJ', 'IYK', 'IYLD', 'IYM', 'IYR', 'IYT', 'IYW', 'IYY', 'IYZ', 'IZRL', 'JAGG', 'JCE', 'JCO', 'JCPB', 'JDD', 'JDIV', 'JDST', 'JEMD', 'JEPI', 'JEQ', 'JETS', 'JFR', 'JGH', 'JHAA', 'JHCS', 'JHEM', 'JHI', 'JHMA', 'JHMC', 'JHMD', 'JHME', 'JHMF', 'JHMH', 'JHMI', 'JHML', 'JHMM', 'JHMS', 'JHMT', 'JHMU', 'JHS', 'JHSC', 'JHY', 'JIG', 'JIGB', 'JJA', 'JJC', 'JJE', 'JJG', 'JJM', 'JJN', 'JJP', 'JJS', 'JJT', 'JJU', 'JKD', 'JKE', 'JKF', 'JKG', 'JKH', 'JKI', 'JKJ', 'JKK', 'JKL', 'JLS', 'JMBS', 'JMIN', 'JMM', 'JMOM', 'JMST', 'JMUB', 'JNK', 'JNUG', 'JO', 'JOF', 'JOYY', 'JPC', 'JPED', 'JPEM', 'JPEU', 'JPGB', 'JPGE', 'JPHF', 'JPHY', 'JPI', 'JPIN', 'JPLS', 'JPMB', 'JPME', 'JPMF', 'JPMV', 'JPN', 'JPNL', 'JPS', 'JPSE', 'JPST', 'JPT', 'JPUS', 'JPXN', 'JQC', 'JQUA', 'JRI', 'JRO', 'JRS', 'JSD', 'JSMD', 'JSML', 'JTA', 'JTD', 'JUST', 'JVAL', 'JXI', 'KALL', 'KAPR', 'KARS', 'KBA', 'KBE', 'KBWB', 'KBWD', 'KBWP', 'KBWR', 'KBWY', 'KCCB', 'KCE', 'KCNY', 'KDFI', 'KEMQ', 'KEMX', 'KF', 'KFYP', 'KGRN', 'KIE', 'KIO', 'KJAN', 'KLCD', 'KLDW', 'KMED', 'KMF', 'KNAB', 'KNG', 'KNOW', 'KOCT', 'KOIN', 'KOKU', 'KOL', 'KOLD', 'KOMP', 'KORP', 'KORU', 'KRE', 'KRMA', 'KRP', 'KSA', 'KSCD', 'KSM', 'KTF', 'KURE', 'KWEB', 'KXI', 'LABD', 'LABU', 'LACK', 'LBJ', 'LCR', 'LD', 'LDEM', 'LDP', 'LDRS', 'LDSF', 'LDUR', 'LEAD', 'LEGR', 'LEMB', 'LEND', 'LEO', 'LFEQ', 'LGH', 'LGI', 'LGLV', 'LGOV', 'LIT', 'LKOR', 'LMB', 'LMBS', 'LMLB', 'LNGR', 'LOUP', 'LOWC', 'LQD', 'LQDH', 'LQDI', 'LRGE', 'LRGF', 'LRNZ', 'LSAF', 'LSLT', 'LSST', 'LTL', 'LTPZ', 'LVHB', 'LVHD', 'LVHI', 'LVUS', 'MAAX', 'MAGA', 'MARB', 'MAV', 'MBB', 'MBSD', 'MCA', 'MCC', 'MCEF', 'MCHI', 'MCN', 'MCR', 'MCRO', 'MCX', 'MDIV', 'MDY', 'MDYG', 'MDYV', 'MEAR', 'MEN', 'MEXX', 'MFD', 'MFDX', 'MFEM', 'MFL', 'MFM', 'MFMS', 'MFT', 'MFUS', 'MFV', 'MGC', 'MGF', 'MGK', 'MGU', 'MGV', 'MHD', 'MHE', 'MHF', 'MHI', 'MHN', 'MIDF', 'MIDU', 'MIE', 'MILN', 'MINC', 'MINT', 'MIY', 'MJ', 'MJJ', 'MJO', 'MLN', 'MLPA', 'MLPB', 'MLPC', 'MLPE', 'MLPG', 'MLPI', 'MLPR', 'MLPX', 'MLPY', 'MLTI', 'MMAC', 'MMD', 'MMIN', 'MMIT', 'MMT', 'MMTM', 'MMU', 'MNA', 'MNE', 'MNP', 'MNRL', 'MOAT', 'MOM', 'MOO', 'MORT', 'MOTI', 'MOTO', 'MPA', 'MPV', 'MQT', 'MQY', 'MRGR', 'MSD', 'MSUS', 'MSVX', 'MTGP', 'MTT', 'MTUM', 'MUA', 'MUB', 'MUC', 'MUE', 'MUH', 'MUI', 'MUJ', 'MUNI', 'MUS', 'MUST', 'MUTE', 'MVF', 'MVIN', 'MVO', 'MVRL', 'MVT', 'MVV', 'MXDE', 'MXDU', 'MXE', 'MXF', 'MXI', 'MYC', 'MYD', 'MYF', 'MYI', 'MYJ', 'MYN', 'MYY', 'MZA', 'MZZ', 'NAC', 'NACP', 'NAD', 'NAIL', 'NAN', 'NANR', 'NAPR', 'NAZ', 'NBB', 'NBH', 'NBO', 'NBW', 'NCA', 'NCB', 'NCV', 'NCZ', 'NDP', 'NEA', 'NEAR', 'NEED', 'NERD', 'NETL', 'NEV', 'NFJ', 'NFLT', 'NFRA', 'NFTY', 'NGE', 'NHA', 'NHF', 'NHS', 'NIB', 'NID', 'NIE', 'NIM', 'NIQ', 'NJAN', 'NJV', 'NKG', 'NKX', 'NLR', 'NMCO', 'NMI', 'NML', 'NMS', 'NMT', 'NMY', 'NMZ', 'NNY', 'NOBL', 'NOCT', 'NOM', 'NORW', 'NPN', 'NPV', 'NQP', 'NRGD', 'NRGO', 'NRGU', 'NRGX', 'NRGZ', 'NRK', 'NRO', 'NSL', 'NTG', 'NTSX', 'NUAG', 'NUBD', 'NUDM', 'NUEM', 'NUGT', 'NUHY', 'NULC', 'NULG', 'NULV', 'NUM', 'NUMG', 'NUMV', 'NUO', 'NURE', 'NUSA', 'NUSC', 'NUSI', 'NUV', 'NUW', 'NVG', 'NXC', 'NXJ', 'NXN', 'NXP', 'NXQ', 'NXR', 'NXTG', 'NYF', 'NYV', 'NZF', 'OBOR', 'OCCI', 'OCIO', 'OCSI', 'OEF', 'OEUR', 'OGIG', 'OIA', 'OIH', 'OILK', 'OLD', 'OLEM', 'OMFL', 'OMFS', 'ONEO', 'ONEQ', 'ONEV', 'ONEY', 'ONLN', 'OPER', 'OPP', 'OSCV', 'OUNZ', 'OUSA', 'OUSM', 'OVB', 'OVF', 'OVL', 'OVM', 'OVS', 'OXLC', 'PACA', 'PAI', 'PAK', 'PALL', 'PANL', 'PAPR', 'PASS', 'PAUG', 'PAVE', 'PAWZ', 'PBD', 'PBDM', 'PBE', 'PBEE', 'PBJ', 'PBND', 'PBP', 'PBS', 'PBSM', 'PBT', 'PBTP', 'PBUS', 'PBW', 'PCEF', 'PCF', 'PCI', 'PCK', 'PCM', 'PCN', 'PCQ', 'PCY', 'PDBC', 'PDEC', 'PDI', 'PDN', 'PDP', 'PDT', 'PEJ', 'PEO', 'PEX', 'PEXL', 'PEY', 'PEZ', 'PFD', 'PFEB', 'PFF', 'PFFA', 'PFFD', 'PFFL', 'PFFR', 'PFI', 'PFIG', 'PFL', 'PFLD', 'PFM', 'PFN', 'PFO', 'PFXF', 'PGAL', 'PGF', 'PGHY', 'PGJ', 'PGM', 'PGP', 'PGX', 'PGZ', 'PHB', 'PHD', 'PHDG', 'PHK', 'PHO', 'PHT', 'PHYL', 'PHYS', 'PICB', 'PICK', 'PID', 'PIE', 'PILL', 'PIM', 'PIN', 'PIO', 'PIZ', 'PJAN', 'PJP', 'PJUL', 'PJUN', 'PKB', 'PKO', 'PKW', 'PLAT', 'PLC', 'PLCY', 'PLTM', 'PLW', 'PMAR', 'PMAY', 'PME', 'PMF', 'PML', 'PMM', 'PMO', 'PMOM', 'PMX', 'PNF', 'PNI', 'PNOV', 'PNQI', 'POCT', 'POTX', 'PPA', 'PPDM', 'PPEM', 'PPH', 'PPLC', 'PPLT', 'PPMC', 'PPR', 'PPSC', 'PPT', 'PPTY', 'PQIN', 'PQLC', 'PQSG', 'PQSV', 'PREF', 'PRF', 'PRFZ', 'PRN', 'PRNT', 'PRPL', 'PSC', 'PSCC', 'PSCD', 'PSCE', 'PSCF', 'PSCH', 'PSCI', 'PSCM', 'PSCT', 'PSCU', 'PSEP', 'PSET', 'PSF', 'PSI', 'PSJ', 'PSK', 'PSL', 'PSLV', 'PSM', 'PSMB', 'PSMC', 'PSMG', 'PSMM', 'PSP', 'PSQ', 'PSR', 'PST', 'PTBD', 'PTEU', 'PTF', 'PTH', 'PTIN', 'PTLC', 'PTMC', 'PTNQ', 'PTY', 'PUI', 'PULS', 'PUTW', 'PVAL', 'PVI', 'PWB', 'PWC', 'PWS', 'PWV', 'PWZ', 'PXE', 'PXF', 'PXH', 'PXI', 'PXJ', 'PXQ', 'PY', 'PYN', 'PYPE', 'PYZ', 'PZA', 'PZC', 'PZD', 'PZT', 'QABA', 'QAI', 'QARP', 'QAT', 'QCLN', 'QDEF', 'QDF', 'QDIV', 'QDYN', 'QED', 'QEFA', 'QEMM', 'QGRO', 'QGTA', 'QID', 'QINT', 'QLC', 'QLD', 'QLS', 'QLTA', 'QLV', 'QLVD', 'QLVE', 'QMJ', 'QMN', 'QMOM', 'QQEW', 'QQH', 'QQQ', 'QQQE', 'QQQX', 'QQXT', 'QRFT', 'QSY', 'QTEC', 'QTUM', 'QUAL', 'QUS', 'QVAL', 'QVM', 'QWLD', 'QYLD', 'RA', 'RAAX', 'RAFE', 'RALS', 'RAVI', 'RBIN', 'RBUS', 'RCD', 'RCG', 'RCS', 'RDIV', 'RDOG', 'RDVY', 'RECS', 'REET', 'REGL', 'REK', 'REM', 'REML', 'REMX', 'RESD', 'RESE', 'RESP', 'RETL', 'REVS', 'REW', 'REZ', 'RFAP', 'RFCI', 'RFDA', 'RFDI', 'RFEM', 'RFEU', 'RFFC', 'RFG', 'RFI', 'RFUN', 'RFV', 'RGI', 'RGT', 'RHS', 'RIF', 'RIGS', 'RINF', 'RING', 'RISE', 'RIV', 'RJA', 'RJI', 'RJN', 'RJZ', 'RLY', 'RMM', 'RMT', 'RNDM', 'RNDV', 'RNEM', 'RNLC', 'RNMC', 'RNP', 'RNSC', 'ROAM', 'ROBO', 'ROBT', 'RODE', 'RODI', 'RODM', 'ROKT', 'ROM', 'ROMO', 'ROOF', 'RORE', 'ROSC', 'ROSE', 'ROSEU', 'ROUS', 'RPAR', 'RPG', 'RPV', 'RQI', 'RSF', 'RSP', 'RSX', 'RSXJ', 'RTH', 'RTM', 'RUSL', 'RVNU', 'RVRS', 'RVT', 'RWCD', 'RWDC', 'RWDE', 'RWED', 'RWGV', 'RWIU', 'RWJ', 'RWK', 'RWL', 'RWLS', 'RWM', 'RWO', 'RWR', 'RWSL', 'RWUI', 'RWVG', 'RWX', 'RXD', 'RXI', 'RXL', 'RYE', 'RYF', 'RYH', 'RYJ', 'RYLD', 'RYT', 'RYU', 'RZG', 'RZV', 'SAA', 'SBB', 'SBI', 'SBIO', 'SBM', 'SBUG', 'SCC', 'SCD', 'SCHA', 'SCHB', 'SCHC', 'SCHD', 'SCHE', 'SCHF', 'SCHG', 'SCHH', 'SCHI', 'SCHJ', 'SCHK', 'SCHM', 'SCHO', 'SCHP', 'SCHQ', 'SCHR', 'SCHV', 'SCHX', 'SCHZ', 'SCID', 'SCIJ', 'SCIU', 'SCIX', 'SCJ', 'SCO', 'SCZ', 'SDAG', 'SDCI', 'SDD', 'SDEM', 'SDG', 'SDGA', 'SDIV', 'SDOG', 'SDOW', 'SDP', 'SDS', 'SDVY', 'SDY', 'SDYL', 'SECT', 'SEF', 'SEIX', 'SFHY', 'SFIG', 'SFY', 'SFYF', 'SFYX', 'SGDJ', 'SGDM', 'SGG', 'SGOL', 'SGOV', 'SH', 'SHAG', 'SHE', 'SHIP', 'SHM', 'SHV', 'SHY', 'SHYD', 'SHYG', 'SHYL', 'SIJ', 'SIL', 'SILJ', 'SIMS', 'SIVR', 'SIXA', 'SIXH', 'SIXL', 'SIXS', 'SIZE', 'SJB', 'SJNK', 'SJT', 'SKF', 'SKOR', 'SKYY', 'SLQD', 'SLT', 'SLV', 'SLVO', 'SLVP', 'SLX', 'SLY', 'SLYG', 'SLYV', 'SMB', 'SMCP', 'SMDD', 'SMDV', 'SMEZ', 'SMH', 'SMHB', 'SMIN', 'SMLF', 'SMLL', 'SMLV', 'SMM', 'SMMD', 'SMMU', 'SMMV', 'SMN', 'SMOG', 'SNLN', 'SNPE', 'SNSR', 'SNUG', 'SOCL', 'SOIL', 'SOR', 'SOVB', 'SOXL', 'SOXS', 'SOXX', 'SOYB', 'SPAB', 'SPBO', 'SPDN', 'SPDV', 'SPDW', 'SPE', 'SPEM', 'SPEU', 'SPFF', 'SPGM', 'SPGP', 'SPHB', 'SPHD', 'SPHQ', 'SPHY', 'SPIB', 'SPIP', 'SPLB', 'SPLG', 'SPLV', 'SPMB', 'SPMD', 'SPMO', 'SPMV', 'SPPP', 'SPSB', 'SPSK', 'SPSM', 'SPTI', 'SPTL', 'SPTM', 'SPTS', 'SPUS', 'SPUU', 'SPVM', 'SPVU', 'SPXB', 'SPXE', 'SPXL', 'SPXN', 'SPXS', 'SPXT', 'SPXU', 'SPXV', 'SPXX', 'SPY', 'SPYD', 'SPYG', 'SPYV', 'SPYX', 'SQEW', 'SQLV', 'SQQQ', 'SRET', 'SRLN', 'SRS', 'SRTY', 'SRV', 'SRVR', 'SSG', 'SSLY', 'SSO', 'SSPY', 'SSUS', 'STIP', 'STK', 'STLC', 'STLG', 'STLV', 'STMB', 'STOT', 'STPP', 'STPZ', 'STSB', 'SUB', 'SUSA', 'SUSB', 'SUSC', 'SUSL', 'SVXY', 'SWAN', 'SWKH', 'SWZ', 'SYE', 'SYG', 'SYLD', 'SYV', 'SZC', 'SZK', 'SZNE', 'TAAG', 'TADS', 'TAEQ', 'TAGS', 'TAIL', 'TAN', 'TAPR', 'TAWK', 'TAXF', 'TBF', 'TBLU', 'TBND', 'TBT', 'TBX', 'TCPC', 'TCTL', 'TDF', 'TDIV', 'TDTF', 'TDTT', 'TDV', 'TEAF', 'TECB', 'TECL', 'TECS', 'TEGS', 'TEI', 'TERM', 'TFI', 'TFIV', 'TFLO', 'TFLT', 'THCX', 'THD', 'THNQ', 'THQ', 'THW', 'TILT', 'TIP', 'TIPX', 'TIPZ', 'TLDH', 'TLEH', 'TLH', 'TLI', 'TLT', 'TLTD', 'TLTE', 'TMDV', 'TMF', 'TMFC', 'TMV', 'TNA', 'TOK', 'TOKE', 'TOLZ', 'TOTL', 'TPAY', 'TPHD', 'TPIF', 'TPL', 'TPLC', 'TPOR', 'TPSC', 'TPYP', 'TPZ', 'TQQQ', 'TRND', 'TRTY', 'TSI', 'TTAC', 'TTAI', 'TTP', 'TTT', 'TTTN', 'TUR', 'TUSA', 'TVIX', 'TWM', 'TWN', 'TY', 'TYBS', 'TYD', 'TYG', 'TYO', 'TZA', 'UAE', 'UAG', 'UAPR', 'UAUD', 'UAUG', 'UBG', 'UBOT', 'UBR', 'UBT', 'UCC', 'UCHF', 'UCI', 'UCIB', 'UCO', 'UCON', 'UDEC', 'UDN', 'UDOW', 'UEUR', 'UEVM', 'UFEB', 'UFO', 'UGA', 'UGAZ', 'UGBP', 'UGE', 'UGL', 'UGLD', 'UITB', 'UIVM', 'UJAN', 'UJB', 'UJPY', 'UJUL', 'UJUN', 'ULE', 'ULST', 'ULTR', 'ULVM', 'UMAR', 'UMAY', 'UMDD', 'UNG', 'UNL', 'UNOV', 'UOCT', 'UPRO', 'UPV', 'UPW', 'URA', 'URE', 'URNM', 'URTH', 'URTY', 'USA', 'USAI', 'USCI', 'USD', 'USDU', 'USDY', 'USEP', 'USEQ', 'USFR', 'USHG', 'USHY', 'USI', 'USIG', 'USL', 'USLB', 'USLV', 'USMC', 'USMF', 'USMV', 'USO', 'USOI', 'USRT', 'USSG', 'UST', 'USTB', 'USV', 'USVM', 'UTES', 'UTF', 'UTG', 'UTRN', 'UTSL', 'UUP', 'UVXY', 'UWM', 'UXI', 'UYG', 'UYM', 'VALQ', 'VALT', 'VAM', 'VAMO', 'VAW', 'VB', 'VBF', 'VBK', 'VBND', 'VBR', 'VCF', 'VCIF', 'VCIT', 'VCLT', 'VCR', 'VCSH', 'VCV', 'VDC', 'VDE', 'VEA', 'VEGA', 'VEGI', 'VEGN', 'VEU', 'VFH', 'VFL', 'VFLQ', 'VFMF', 'VFMO', 'VFMV', 'VFQY', 'VFVA', 'VGFO', 'VGI', 'VGIT', 'VGK', 'VGLT', 'VGM', 'VGSH', 'VGT', 'VHT', 'VIDI', 'VIG', 'VIGI', 'VIIX', 'VIOG', 'VIOO', 'VIOV', 'VIS', 'VIXM', 'VIXY', 'VKI', 'VKQ', 'VLT', 'VLU', 'VLUE', 'VMBS', 'VMM', 'VMO', 'VMOT', 'VNLA', 'VNM', 'VNQ', 'VNQI', 'VO', 'VOE', 'VONE', 'VONG', 'VONV', 'VOO', 'VOOG', 'VOOV', 'VOT', 'VOX', 'VPC', 'VPL', 'VPU', 'VPV', 'VQT', 'VRAI', 'VRIG', 'VRP', 'VSDA', 'VSGX', 'VSL', 'VSMV', 'VSS', 'VT', 'VTA', 'VTC', 'VTEB', 'VTHR', 'VTI', 'VTIP', 'VTN', 'VTV', 'VTWG', 'VTWO', 'VTWV', 'VUG', 'VUSE', 'VV', 'VVR', 'VWO', 'VWOB', 'VXF', 'VXUS', 'VXX', 'VYM', 'VYMI', 'WANT', 'WBIE', 'WBIF', 'WBIG', 'WBII', 'WBIL', 'WBIN', 'WBIT', 'WBIY', 'WBND', 'WCLD', 'WDIV', 'WEA', 'WEAT', 'WEBL', 'WEBS', 'WFHY', 'WFIG', 'WIA', 'WIL', 'WINC', 'WIP', 'WIW', 'WIZ', 'WLDR', 'WOMN', 'WOOD', 'WPS', 'WTMF', 'WUGI', 'WWJD', 'XAR', 'XBI', 'XBUY', 'XCEM', 'XCOM', 'XDIV', 'XES', 'XFLT', 'XHB', 'XHE', 'XHS', 'XITK', 'XLB', 'XLC', 'XLE', 'XLF', 'XLG', 'XLI', 'XLK', 'XLP', 'XLRE', 'XLSR', 'XLU', 'XLV', 'XLY', 'XME', 'XMHQ', 'XMLV', 'XMMO', 'XMPT', 'XMVM', 'XNTK', 'XOP', 'XOUT', 'XPH', 'XPP', 'XRLV', 'XRT', 'XSD', 'XSHD', 'XSHQ', 'XSLV', 'XSMO', 'XSOE', 'XSVM', 'XSW', 'XT', 'XTL', 'XTN', 'XVZ', 'XWEB', 'YANG', 'YCL', 'YCOM', 'YCS', 'YGRN', 'YINN', 'YLCO', 'YLD', 'YLDE', 'YOLO', 'YXI', 'YYY', 'ZCAN', 'ZDEU', 'ZGBR', 'ZHOK', 'ZIG', 'ZIV', 'ZJPN', 'ZMLP', 'ZOM', 'ZROZ', 'ZSL', 'ZTR']
miscellaneous_master_file='miscellaneous_mf.csv'
miscellaneous_directory='miscellaneous'
get_ticker_data(miscellaneous,miscellaneous_directory)
create_master_list(miscellaneous,miscellaneous_master_file,miscellaneous_directory)
print("-> Master file created, filename: %s\n" % miscellaneous_master_file)
purge_low_gain_volume(miscellaneous_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
non_energy_minerals = ['AA', 'AAU', 'ACH', 'AEM', 'AG', 'AGI', 'ALO', 'AP', 'AQMS', 'ARNC', 'ASM', 'ATI', 'AU', 'AUG', 'AUMN', 'AUY', 'AXU', 'BBL', 'BCC', 'BHP', 'BTG', 'BVN', 'CCJ', 'CDE', 'CENX', 'CINR', 'CLF', 'CMC', 'CMCL', 'CPAC', 'CRH', 'CRS', 'CSTE', 'CVIA', 'CX', 'EGO', 'EMX', 'EQX', 'EXK', 'EXP', 'FCX', 'FNV', 'FRTA', 'FSM', 'GAU', 'GFI', 'GGB', 'GMO', 'GOL', 'GORO', 'GPL', 'GSS', 'GSV', 'HAYN', 'HBM', 'HCR', 'HHT', 'HL', 'HMY', 'HWM', 'HYMC', 'IAG', 'IPI', 'JEL', 'JHX', 'KALU', 'KGC', 'KL', 'LEU', 'LODE', 'LOMA', 'LPX', 'MAG', 'MLM', 'MMX', 'MSB', 'MT', 'MTL', 'MTL', 'MTRN', 'MUX', 'NAK', 'NEM', 'NEXA', 'NG', 'NG', 'NUE', 'NXE', 'OR', 'OSB', 'OSN', 'PAAS', 'PKX', 'PLG', 'PLM', 'PVG', 'PZG', 'RETO', 'RGL', 'RIO', 'RS', 'RYI', 'SA', 'SAN', 'SBSW', 'SCCO', 'SCHN', 'SI', 'SILV', 'SLCA', 'SMTS', 'SN', 'SSRM', 'STL', 'SUM', 'SVM', 'SYNL', 'TECK', 'TGB', 'THM', 'TMQ', 'TMST', 'TREX', 'TRQ', 'TRX', 'TS', 'TX', 'UEC', 'UFPI', 'URG', 'USAP', 'USAS', 'USAU', 'USCR', 'USLM', 'UUUU', 'VALE', 'VEDL', 'VGZ', 'VMC', 'WOR', 'WPM', 'WRN', 'WWR', 'X', 'XPL', 'ZEUS']
non_energy_minerals_master_file='non_energy_minerals_mf.csv'
non_energy_minerals_directory='non_energy_minerals'
get_ticker_data(non_energy_minerals,non_energy_minerals_directory)
create_master_list(non_energy_minerals,non_energy_minerals_master_file,non_energy_minerals_directory)
print("-> Master file created, filename: %s\n" % non_energy_minerals_master_file)
purge_low_gain_volume(non_energy_minerals_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
producer_manufacturer = ['AAON', 'ABB', 'ACA', 'ACCO', 'ACMR', 'ADES', 'ADNT', 'AFI', 'AGCO', 'AGS', 'AIMC', 'ALG', 'ALLE', 'ALSN', 'ALV', 'AMAT', 'AME', 'AMSC', 'AMWD', 'AOS', 'APOG', 'APTV', 'ARTW', 'ASML', 'ASTE', 'ASYS', 'ATKR', 'AWI', 'AXL', 'AYI', 'AYRO', 'AZZ', 'B', 'BDC', 'BE', 'BGG', 'BHTG', 'BIMI', 'BLBD', 'BLD', 'BLDP', 'BLDR', 'BMCH', 'BMI', 'BOOM', 'BRC', 'BWA', 'BWEN', 'CAAS', 'CARR', 'CAT', 'CBAT', 'CCCL', 'CECE', 'CFX', 'CFXA', 'CIR', 'CIX', 'CMCO', 'CMI', 'CMT', 'CNHI', 'CNR', 'CPS', 'CPSH', 'CPST', 'CR', 'CREG', 'CSIQ', 'CSL', 'CVGI', 'CVR', 'CXDC', 'CYD', 'DAN', 'DDD', 'DE', 'DLPH', 'DOOR', 'DPW', 'EAF', 'EFOI', 'ELSE', 'EMR', 'ENS', 'EPAC', 'ERII', 'ETN', 'FBHS', 'FELE', 'FLOW', 'FLS', 'FORK', 'FRD', 'FSS', 'FSTR', 'GBX', 'GENC', 'GFF', 'GGG', 'GHM', 'GIFI', 'GNRC', 'GNTX', 'GRC', 'GTEC', 'GTES', 'GTLS', 'GTX', 'HCHC', 'HEBT', 'HI', 'HIHO', 'HLIO', 'HNI', 'HON', 'HSC', 'HUBB', 'HURC', 'HWCC', 'HY', 'IEP', 'IEX', 'IIIN', 'IPWR', 'IR', 'ITGR', 'ITT', 'ITW', 'JBT', 'JCI', 'JKS', 'KAI', 'KEQU', 'KIQ', 'KMT', 'KNL', 'KRNT', 'LCII', 'LEA', 'LECO', 'LFUS', 'LII', 'LIQT', 'LITE', 'LNN', 'LXFR', 'LYTS', 'MAS', 'MATW', 'MEC', 'MGA', 'MIDD', 'MLHR', 'MLI', 'MLR', 'MMM', 'MNTX', 'MOD', 'MPAA', 'MTOR', 'MTW', 'MWA', 'NAV', 'NDSN', 'NEWA', 'NL', 'NNBR', 'NNDM', 'NPO', 'NSSC', 'NWL', 'NWPX', 'NX', 'OEG', 'OESX', 'OFLX', 'OPTT', 'OSK', 'OTIS', 'PBI', 'PCAR', 'PFIE', 'PFIN', 'PGTI', 'PH', 'PKOH', 'PLOW', 'PLPC', 'PNR', 'POWL', 'PPIH', 'PPSI', 'PRLB', 'RAIL', 'RAVN', 'RBC', 'REYN', 'ROCK', 'ROK', 'ROLL', 'RXN', 'SCS', 'SEDG', 'SGBX', 'SHLO', 'SHYF', 'SIM', 'SMIT', 'SPLP', 'SPWR', 'SPXC', 'SRI', 'SSD', 'STRT', 'SUNW', 'SUP', 'SVT', 'SXI', 'SYPR', 'TAYD', 'TEN', 'TEX', 'TG', 'TGEN', 'TGLS', 'THR', 'THRM', 'TILE', 'TKR', 'TNC', 'TPIC', 'TRN', 'TRS', 'TT', 'TTC', 'TUP', 'TWI', 'TWIN', 'UAMY', 'UFAB', 'ULBI', 'VC', 'VIRC', 'VJET', 'VMI', 'VNE', 'VTSI', 'WAB', 'WBC', 'WBT', 'WIRE', 'WKHS', 'WMS', 'WNC', 'WPRT', 'WRTC', 'WSO', 'WSO.B', 'WTS', 'WWD', 'XONE', 'XYL', 'YETI', 'ZKIN']
producer_manufacturer_master_file='process_industries_mf.csv'
producer_manufacturer_directory='producer_manufacturer'
get_ticker_data(producer_manufacturer,producer_manufacturer_directory)
create_master_list(producer_manufacturer,producer_manufacturer_master_file,producer_manufacturer_directory)
print("-> Master file created, filename: %s\n" % producer_manufacturer_master_file)
purge_low_gain_volume(producer_manufacturer_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
process_industries = ['ACB', 'ADM', 'AGFS', 'AGRO', 'AIN', 'ALB', 'ALCO', 'AMCR', 'AMRS', 'AMTX', 'ANDE', 'APD', 'APHA', 'ARD', 'ASH', 'ASIX', 'ASPN', 'ATR', 'AVD', 'AVTR', 'AVY', 'AXTA', 'BAK', 'BCPC', 'BERY', 'BG', 'BLL', 'CBT', 'CC', 'CCF', 'CCK', 'CDXC', 'CDXS', 'CE', 'CF', 'CGA', 'CGC', 'CLMT', 'CLW', 'CMP', 'CRON', 'CSWI', 'CTIB', 'CTVA', 'CULP', 'DAR', 'DCI', 'DD', 'DOW', 'DQ', 'DSWL', 'ECL', 'ECL.WD', 'ELAN', 'EML', 'EMN', 'ESI', 'EVA', 'FF', 'FFHL', 'FMC', 'FOE', 'FSI', 'FUL', 'GCP', 'GEF', 'GEF.B', 'GEVO', 'GLT', 'GPK', 'GPRE', 'GRA', 'GSM', 'GURE', 'HEXO', 'HUGE', 'HUN', 'IBA', 'ICL', 'IKNX', 'INGR', 'IOSP', 'IP', 'ITP', 'KOP', 'KRA', 'KRO', 'KWR', 'LAC', 'LAKE', 'LDL', 'LIN', 'LMNR', 'LND', 'LOOP', 'LTHM', 'LXU', 'LYB', 'MBII', 'MEOH', 'MERC', 'MGPI', 'MOS', 'MTX', 'MYE', 'MYT', 'NEU', 'NGVT', 'NP', 'NTIC', 'NTR', 'ODC', 'OEC', 'OGI', 'OI', 'OLN', 'PACK', 'PEIX', 'PKG', 'PLL', 'POL', 'PPG', 'PQG', 'REGI', 'REX', 'RFP', 'RKDA', 'ROG', 'RPM', 'RYAM', 'RYCE', 'SANW', 'SCL', 'SEE', 'SEED', 'SHI', 'SHW', 'SLGN', 'SMG', 'SNDL', 'SON', 'SQM', 'SSL', 'SUZ', 'SWM', 'TANH', 'TLRY', 'TROX', 'TSE', 'TYHT', 'UAN', 'UFI', 'UFPT', 'UFS', 'VFF', 'VHI', 'VNTR', 'VRS', 'VVV', 'WDFC', 'WLK', 'WLKP', 'WRK', 'YTEN', 'ZAGG']
process_industries_master_file ='process_industries_mf.csv'
process_industries_directory='process_industries'
get_ticker_data(process_industries,process_industries_directory)
create_master_list(process_industries,process_industries_master_file,process_industries_directory)
print("-> Master file created, filename: %s\n" % process_industries_master_file)
purge_low_gain_volume(process_industries_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
retail_trade =['AAP', 'ABG', 'AEO', 'AMZN', 'AN', 'ANF', 'APRN', 'ARTL', 'ASNA', 'ATV', 'AZO', 'BABA', 'BBBY', 'BBW', 'BBY', 'BGFV', 'BGI', 'BIG', 'BJ', 'BKE', 'BNED', 'BOOT', 'BURL', 'BWMX', 'BZUN', 'CASY', 'CATO', 'CBD', 'CHS', 'CHWY', 'CJJD', 'CONN', 'COST', 'CPRI', 'CRI', 'CRMT', 'CTRN', 'CVNA', 'CVS', 'CWH', 'DBI', 'DDS', 'DG', 'DKS', 'DLTH', 'DLTR', 'DXLG', 'ELA', 'EXPR', 'EYE', 'FAMI', 'FIVE', 'FL', 'FLWS', 'FND', 'FRAN', 'GME', 'GNC', 'GO', 'GPI', 'GPS', 'GRUB', 'HD', 'HIBB', 'HOME', 'HUD', 'HVT', 'HVT.A', 'HZO', 'IAA', 'IFMK', 'IMBI', 'IMKTA', 'JD', 'JILL', 'JWN', 'KIRK', 'KMX', 'KR', 'KSS', 'KXIN', 'LAD', 'LAZY', 'LB', 'LE', 'LITB', 'LIVE', 'LL', 'LOVE', 'LOW', 'LULU', 'M', 'MED', 'MIK', 'MNRO', 'MOGU', 'MUSA', 'NGVC', 'ODP', 'OLLI', 'ONEW', 'ORLY', 'OSTK', 'PAG', 'PDD', 'PETS', 'PLCE', 'PRTS', 'PRTY', 'PSMT', 'QRTEA', 'QRTEB', 'QVCD', 'RAD', 'REAL', 'RH', 'RL', 'ROST', 'RTW', 'RVLV', 'SAH', 'SBH', 'SCVL', 'SECO', 'SFIX', 'SFM', 'SHOP', 'SIG', 'SMRT', 'SPWH', 'SYX', 'TC', 'TCS', 'TGT', 'TIF', 'TJX', 'TLRD', 'TLYS', 'TPR', 'TSCO', 'TUES', 'TWMC', 'ULTA', 'URBN', 'VIPS', 'VLGEA', 'W', 'WBA', 'WMK', 'WMT', 'WSM', 'WTRH', 'YGYI', 'YI', 'YJ', 'ZUMZ']
retail_trade_master_file='retail_trade_mf.csv'
retail_trade_directory='retail_trade'
get_ticker_data(retail_trade,retail_trade_directory)
create_master_list(retail_trade,retail_trade_master_file,retail_trade_directory)
print("-> Master file created, filename: %s\n" % retail_trade_master_file)
purge_low_gain_volume(retail_trade_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
technology_services =['ACIW', 'ACN', 'ADBE', 'ADP', 'ADS', 'ADSK', 'AEYE', 'AGYS', 'AKAM', 'ALRM', 'ALTR', 'ALYA', 'AMRH', 'AMSWA', 'ANGI', 'ANSS', 'ANY', 'APPF', 'APPN', 'APPS', 'ARCE', 'ASUR', 'ATEN', 'ATHM', 'AVID', 'AVLR', 'AYX', 'AZPN', 'BB', 'BCOR', 'BCOV', 'BIDU', 'BILI', 'BILL', 'BITA', 'BKI', 'BKYI', 'BL', 'BLIN', 'BLKB', 'BNFT', 'BOX', 'BRQS', 'BSQR', 'CACI', 'CANG', 'CARG', 'CARS', 'CASS', 'CCRC', 'CDLX', 'CDNS', 'CDW', 'CERN', 'CHKP', 'CHNG', 'CHNGU', 'CIH', 'CLDR', 'CLGX', 'CLPS', 'CMCM', 'CNXN', 'COE', 'COUP', 'CPSI', 'CREX', 'CRM', 'CRNC', 'CRTO', 'CRWD', 'CSCO', 'CSGP', 'CSGS', 'CSLT', 'CSOD', 'CSPI', 'CTEK', 'CTG', 'CTK', 'CTSH', 'CTXS', 'CVET', 'CVLT', 'CYBR', 'CYRN', 'DAO', 'DAVA', 'DBX', 'DDOG', 'DHX', 'DKNG', 'DMRC', 'DNK', 'DOCU', 'DOMO', 'DOX', 'DOYU', 'DSGX', 'DT', 'DTSS', 'DUO', 'DUOT', 'DWSN', 'DXC', 'EB', 'ECOM', 'EEFT', 'EGAN', 'EGHT', 'EGOV', 'EH', 'EIGI', 'ENV', 'EPAM', 'EPAY', 'ESTC', 'EVBG', 'EVER', 'EVOL', 'EVTC', 'EXLS', 'FB', 'FDS', 'FEYE', 'FICO', 'FIS', 'FISV', 'FIVN', 'FLNT', 'FNJN', 'FORTY', 'FSCT', 'FSLY', 'FTCH', 'FUTU', 'FVRR', 'GAN', 'GDDY', 'GDS', 'GEC', 'GIB', 'GLOB', 'GLUU', 'GOOG', 'GOOGL', 'GPN', 'GSB', 'GSUM', 'GSX', 'GTYH', 'GWRE', 'HCAT', 'HHR', 'HSTM', 'HUBS', 'HUYA', 'IAC', 'IBM', 'ICAD', 'IDEX', 'IIIV', 'INFY', 'INOD', 'INOV', 'INPX', 'INS', 'INSE', 'INTU', 'INUV', 'IO', 'IQ', 'ISDR', 'IZEA', 'JCOM', 'JFU', 'JG', 'JKHY', 'JMIA', 'JNPR', 'JOBS', 'JRJC', 'KERN', 'KRKR', 'LAIX', 'LDOS', 'LEAF', 'LINX', 'LIZI', 'LKCO', 'LLNW', 'LMPX', 'LN', 'LOGM', 'LPSN', 'LVGO', 'LYFT', 'MANH', 'MANT', 'MCHX', 'MDB', 'MDLA', 'MDRX', 'MEET', 'MELI', 'MFGP', 'MFH', 'MGIC', 'MIME', 'MITK', 'MIXT', 'MKD', 'MOBL', 'MODN', 'MOMO', 'MOXC', 'MRIN', 'MSFT', 'MSTR', 'MTBC', 'MTC', 'MTCH', 'MTLS', 'MTSL', 'MWK', 'MYSZ', 'NATI', 'NCTY', 'NEON', 'NET', 'NETE', 'NEWR', 'NH', 'NICE', 'NLOK', 'NOW', 'NSIT', 'NTCT', 'NTES', 'NTGR', 'NTNX', 'NTWK', 'NUAN', 'NVEC', 'NXGN', 'OKTA', 'OMCL', 'OOMA', 'OPRA', 'OPRX', 'ORCL', 'OSPN', 'OTEX', 'PAYC', 'PAYX', 'PBTS', 'PCTY', 'PCYG', 'PD', 'PDFS', 'PEGA', 'PERI', 'PFPT', 'PFSW', 'PHR', 'PHUN', 'PING', 'PINS', 'PLAN', 'PRGS', 'PRO', 'PRSP', 'PRTH', 'PS', 'PSN', 'PT', 'PTC', 'PYPL', 'QADA', 'QADB', 'QD', 'QIWI', 'QLYS', 'QTT', 'QTWO', 'RAMP', 'RDCM', 'RDVT', 'RENN', 'RMBL', 'RMNI', 'RNG', 'RNWK', 'RP', 'RPD', 'SABR', 'SAIC', 'SAIL', 'SAP', 'SCPL', 'SCWX', 'SDGR', 'SE', 'SFET', 'SFUN', 'SGLB', 'SHSP', 'SIFY', 'SINA', 'SLP', 'SMAR', 'SMSI', 'SNAP', 'SNCR', 'SNPS', 'SOGO', 'SOHU', 'SPLK', 'SPNS', 'SPOT', 'SPRT', 'SPSC', 'SPT', 'SQ', 'SREV', 'SSNC', 'SSNT', 'SSTK', 'STNE', 'STRM', 'SVMK', 'SWCH', 'SWI', 'SY', 'SYKE', 'SYNA', 'SYNC', 'TAOP', 'TCX', 'TEAM', 'TENB', 'TKAT', 'TLND', 'TME', 'TNAV', 'TRHC', 'TRU', 'TRUE', 'TRVG', 'TTGT', 'TW', 'TWLO', 'TWOU', 'TWTR', 'TYL', 'TZOO', 'UBER', 'UEPS', 'UIS', 'UPLD', 'UPWK', 'UXIN', 'VALU', 'VEEV', 'VERI', 'VHC', 'VMW', 'VNET', 'VRNS', 'VRNT', 'VRSN', 'VRTU', 'WB', 'WBAI', 'WDAY', 'WEI', 'WEX', 'WIFI', 'WIMI', 'WIT', 'WIX', 'WK', 'WKEY', 'WORK', 'WUBA', 'WYY', 'XAIR', 'XELA', 'XNET', 'XP', 'XRF', 'YELP', 'YEXT', 'YNDX', 'YRD', 'YVR', 'YY', 'Z', 'ZDGE', 'ZEN', 'ZG', 'ZI', 'ZIXI', 'ZM', 'ZNGA', 'ZS', 'ZUO']
technology_services_master_file='technology_services_mf.csv'
technology_services_directory='technology_services'
get_ticker_data(technology_services,technology_services_directory)
create_master_list(technology_services,technology_services_master_file,technology_services_directory)
print("-> Master file created, filename: %s\n" % technology_services_master_file)
purge_low_gain_volume(technology_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
transportation = ['AAL', 'AAWW', 'AIRT', 'ALGT', 'ALK', 'ARCB', 'ASC', 'ASR', 'ATSG', 'AZUL', 'BEST', 'CAAP', 'CEA', 'CHRW', 'CMRE', 'CNI', 'CP', 'CPA', 'CPLP', 'CSX', 'CTRM', 'CVTI', 'CYRX', 'DAC', 'DAL', 'DHT', 'DLNG', 'DSKE', 'DSSI', 'DSX', 'ECHO', 'EDRY', 'EGLE', 'ERA', 'ESEA', 'EURN', 'EXPD', 'FDX', 'FRO', 'FWRD', 'GASS', 'GFL', 'GFLU', 'GLBS', 'GLNG', 'GLOG', 'GLOP', 'GMLP', 'GNK', 'GOGL', 'GOL', 'GPP', 'GRIN', 'GSH', 'HA', 'HMLP', 'HTLD', 'HUBG', 'INSW', 'JBHT', 'JBLU', 'KEX', 'KNOP', 'KNX', 'KSU', 'LPG', 'LSTR', 'LTM', 'LUV', 'MATX', 'MESA', 'MRTN', 'NAT', 'NM',   'NMCI', 'NMM', 'NNA', 'NSC', 'NVGS', 'ODFL', 'OMAB', 'OSG', 'PAC', 'PATI', 'PRSC', 'PSHG', 'PSV', 'PTSI', 'PXS', 'RLGT', 'RYAAY', 'SAIA', 'SALT', 'SAVE', 'SB', 'SBLK', 'SFL', 'SINO', 'SKYW', 'SMHI', 'SNDR', 'STCN', 'STNG', 'TFII', 'TGP', 'TK', 'TNK', 'TNP', 'TOPS', 'TRMD', 'UAL', 'ULH', 'UNP', 'UPS', 'USAK', 'USX', 'VLRS', 'VRRM', 'WERN', 'XPO', 'YRCW', 'ZNH', 'ZTO']
transportation_master_file='transportation_mf.csv'
transportation_directory='transportation'
get_ticker_data(transportation,transportation_directory)
create_master_list(transportation,transportation_master_file,transportation_directory)
print("-> Master file created, filename: %s\n" % transportation_master_file)
purge_low_gain_volume(transportation_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
utilities = ['AEE', 'AEP', 'AES', 'AGR', 'ALE', 'AQN', 'ARTNA', 'AT', 'ATO', 'AVA', 'AWK', 'AWR', 'AY', 'AZRE', 'BEP', 'BIP', 'BIPC', 'BKH', 'BWXT', 'CDZI', 'CEPU', 'CIG', 'CIG.C', 'CLNE', 'CMS', 'CMSA', 'CMSC', 'CNP', 'CPK', 'CWCO', 'CWEN', 'CWEN.A', 'CWT', 'DCUE', 'DRUA', 'DTE', 'DTJ', 'DTP', 'DTW', 'DTY', 'DUK', 'DUKB', 'DUKH', 'EAI', 'EBR', 'EBR.B', 'ED', 'EDN', 'EE', 'EIX', 'ELC', 'ELLO', 'ELP', 'ELU', 'EMP', 'ENIA', 'ENIC', 'ENJ', 'ENO', 'ES', 'ETR', 'EVRG', 'EVSI', 'EXC', 'EZT', 'FE', 'FTS', 'GNE', 'GPJA', 'GWRS', 'HE', 'HNP', 'IDA', 'JE', 'JHB', 'KEN', 'KEP', 'LNT', 'MDU', 'MGEE', 'MSEX', 'NEE', 'NEP', 'NES', 'NGG', 'NI', 'NJR', 'NRG', 'NWE', 'NWN', 'OGE', 'OGS', 'ORA', 'OTTR', 'PAM', 'PCG', 'PCYO', 'PEG', 'PNM', 'PNW', 'POR', 'PPL', 'RGCO', 'RUN', 'SBS', 'SGU', 'SJI', 'SJIU', 'SJW', 'SKYS', 'SMLP', 'SO', 'SOJA', 'SOJB', 'SOJC', 'SOLN', 'SPH', 'SPKE', 'SR', 'SRE', 'SWX', 'TAC', 'TERP', 'UGI', 'UTL', 'VSLR', 'VST', 'VVPR', 'WEC', 'WTRG', 'WTRU', 'XEL', 'YORW']
utilities_master_file = 'utilities_mf.csv'
utilities_directory='utilities'
get_ticker_data(utilities,utilities_directory)
create_master_list(utilities,utilities_master_file,utilities_directory)
print("-> Master file created, filename: %s\n" % utilities_master_file)
purge_low_gain_volume(utilities_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
commercial_services =['AACG', 'ABM', 'ACTG', 'ADPT', 'ADT', 'AESE', 'ALJJ', 'ALT', 'AMN', 'ANTE', 'APDN', 'ARC', 'ASGN', 'ATOM', 'ATTO', 'AUTO', 'BAH', 'BAND', 'BBSI', 'BBU', 'BCEL', 'BCO', 'BEDU', 'BFAM', 'BGSF', 'BOMN', 'BOXL', 'BR', 'BV', 'CAE', 'CATM', 'CBZ', 'CCC', 'CCO', 'CCRN', 'CDAY', 'CDK', 'CELC', 'CELP', 'CHGG', 'CIDM', 'CLEU', 'CLXT', 'CMPR', 'CNDT', 'CNET', 'CNTG', 'CPRT', 'CRAI', 'CRL', 'CSTM', 'DFIN', 'DL', 'DLHC', 'DLX', 'DSS', 'EAST', 'EBF', 'EDNT', 'EEX', 'EFX', 'ETSY', 'EVH', 'EVOP', 'FC', 'FCN', 'FEDU', 'FLT', 'FORR', 'FRG', 'G', 'GAIA', 'GPX', 'GRPN', 'GSL', 'GVP', 'HCKT', 'HCSG', 'HHS', 'HMSY', 'HQI', 'HSII', 'HSON', 'HURN', 'ICFI', 'ICLK', 'III', 'INWK', 'IPG', 'ISIG', 'IT', 'JOB', 'KAR', 'KC', 'KE', 'KELYA', 'KELYB', 'KFRC', 'KFY', 'KROS', 'LAUR', 'LINC', 'LIVX', 'LMFA', 'LOPE', 'LRMR', 'LTBR', 'MAN', 'MARA', 'MCO', 'MDCA', 'MEDP', 'MEDS', 'METX', 'MGI', 'MHH', 'MKGI', 'MMS', 'MNDO', 'MTA', 'NCMI', 'NEW', 'NEWT', 'NLSN', 'NSP', 'NTIP', 'NVEE', 'NXTD', 'OMC', 'OMEX', 'ONE', 'OPGN', 'PAE', 'PAGS', 'PAYS', 'PCOM', 'PFMT', 'PGNY', 'PIXY', 'PPD', 'PRAA', 'PRAH', 'PRFT', 'PRGX', 'QNST', 'QUAD', 'QUOT', 'RBA', 'RCM', 'REDU', 'REKR', 'RELX', 'REZI', 'RGP', 'RHI', 'RPAY', 'RRD', 'RSSS', 'RUBI', 'RYB', 'SCOR', 'SGRP', 'SIC', 'SJ', 'SLRX', 'SP', 'SPGI', 'SRAX', 'SRT', 'STAF', 'STG', 'STMP', 'TBI', 'TEDU', 'TISI', 'TNET', 'TRI', 'TRXC', 'TSRI', 'TTD', 'TTEC', 'TUFN', 'USIO', 'UTI', 'VEC', 'VERB', 'VOLT', 'VVI', 'VVNT', 'WAFU', 'WLDN', 'WMG', 'WNS', 'WPP', 'XCUR', 'ZCMD']
commercial_services_master_file = 'commercial_services_mf.csv'
commercial_services_directory='commercial_services'
get_ticker_data(commercial_services,commercial_services_directory)
create_master_list(commercial_services,commercial_services_master_file,commercial_services_directory)
print("-> Master file created, filename: %s\n" % commercial_services_master_file)
purge_low_gain_volume(commercial_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')


combine_masters_into_one('sector','master_sector_file.xlsx')





#industries
#blockchain
blockchain_industry =['ACN', 'AMD', 'BITCF', 'BLOC', 'BTCS', 'BTL', 'BTL', 'BTSC', 'CODE', 'DCC', 'DGGXF', 'DPW', 'FTFT', 'GBTC', 'GCAP', 'GROW', 'HIVE', 'IBM', 'INTC', 'INTV', 'JPM', 'LFIN', 'MARA', 'MARK', 'MGTI', 'MSFT', 'NDAQ', 'NETE', 'NVDA', 'OMGT', 'OSTK', 'PFE', 'PRELF', 'QIWI', 'RIOT', 'SAP', 'SIEB', 'SING', 'SQ', 'SRAX', 'SSC', 'TSM', 'UBIA', 'UEPS', 'XBLK', 'XNET', 'ZNGA']
blockchain_master_file = "blockchainmasterList.csv"
blockchain_directory = 'blockchain'
print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
get_industry_ticker_data(blockchain_industry,blockchain_directory)
create_industry_master_list(blockchain_industry,blockchain_master_file,blockchain_directory)
print("-> Master file created, filename: %s\n" % blockchain_master_file)
purge_industry_low_gain_volume(blockchain_master_file)
print('-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

#cannabis
cannabis_industry = ['ABBV', 'ACB', 'ARNA', 'BLOZF', 'CANN', 'CARA', 'CGC', 'CRBP', 'CRON',
               'CVSI', 'DIGP', 'GRNH', 'GRWC', 'GWPH', 'IIPR', 'KAYS', 'KSHB', 'MCIG',
               'MJNA', 'MO', 'MSRT', 'POTN', 'SMG', 'TGODF', 'TLRY', 'TRTC', 'TURV', 'YOLO',
               'ZYNE']
cannabis_master_file = "cannabismasterList.csv"
cannabis_directory = 'cannabis'
print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
get_industry_ticker_data(cannabis_industry,cannabis_directory)
create_industry_master_list(cannabis_industry,cannabis_master_file,cannabis_directory)
print("-> Master file created, filename: %s\n" % cannabis_master_file)
purge_industry_low_gain_volume(cannabis_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')



#ticker source for below - https://www.tradingview.com/markets/stocks-usa/sectorandindustry-sector/


print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
wireless_telecommunications=['AMOV', 'AMX', 'ATEX', 'BCOM', 'CEL', 'CHL', 'GOGO', 'GSAT', 'IRDM', 'MBT', 'ORBC', 'PTNR', 'RCI', 'SKM', 'SPOK', 'TDS', 'TIGO', 'TKC', 'TMUS', 'TSU', 'USM', 'UZB', 'VEON', 'VOD']
wireless_telecommunications_master_file='wireless_telecomm.csv'
wireless_communications_directory = 'wireless_communications'
get_industry_ticker_data(wireless_telecommunications,wireless_communications_directory)
create_industry_master_list(wireless_telecommunications,wireless_telecommunications_master_file,wireless_communications_directory)
print("-> Master file created, filename: %s\n" % wireless_telecommunications_master_file)
purge_industry_low_gain_volume(wireless_telecommunications_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
advertising_marketing_services = ['ANTE', 'AUTO', 'BOMN', 'CCO', 'CNET', 'EAST', 'GRPN', 'HHS', 'ICLK', 'INWK', 'IPG', 'ISIG', 'MDCA', 'MKGI', 'NCMI', 'NLSN', 'OMC', 'QNST', 'QUOT', 'RUBI', 'SRAX', 'TTD', 'WPP']
advertising_marketing_services_master_file='ad_marketing_services_mf.csv'
advertising_marketing_services_directory = 'advertising_marketing_services'
get_industry_ticker_data(advertising_marketing_services ,advertising_marketing_services_directory)
create_industry_master_list(advertising_marketing_services ,advertising_marketing_services_master_file,advertising_marketing_services_directory)
print("-> Master file created, filename: %s\n" % advertising_marketing_services_master_file)
purge_industry_low_gain_volume(advertising_marketing_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
aerospace_defense = ['AAXN', 'AIR', 'AIRI', 'AJR', 'ASTC', 'ATRO', 'AVAV', 'BA', 'CUB', 'CVU', 'CW', 'CODA', 'DCODA', 'DCOD', 'GD', 'RADA', 'TDG', 'TDY', 'ERJ', 'ESE', 'ESLT', 'FLIR', 'GE', 'HEI', 'HEI.A', 'HII', 'HXL', 'ISSC', 'KAMN', 'KTOS', 'KVHI', 'LHX', 'LMT', 'MAGS', 'MAXR', 'MOG.A', 'MOG.B', 'MSI', 'NOC', 'PKE', 'RTX', 'SIF', 'SPCE', 'SPR', 'SSTI', 'SWBI', 'TATT', 'TGI', 'TXT', 'UAVS']
aerospace_defense_master_file='aerospace_defense_mf.csv'
aerospace_defense_directory = 'aerospace_defense'
get_industry_ticker_data(aerospace_defense ,aerospace_defense_directory)
create_industry_master_list(aerospace_defense ,aerospace_defense_master_file,aerospace_defense_directory)
print("-> Master file created, filename: %s\n" % aerospace_defense_master_file)
purge_industry_low_gain_volume(aerospace_defense_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
agricultural_commodities_milling = ['ALCO', 'APHA', 'BG', 'CGC', 'CRON', 'ELAN', 'HEXO', 'HUGE', 'IBA', 'INGR', 'LMNR', 'MGPI', 'OGI', 'RYCE', 'SANW', 'TLRY', 'TYHT', 'VFF', 'SNDL', 'LND', 'DARD', 'ACB', 'ADM', 'AGRO', 'ANDE', 'SEED']
agricultural_commodities_milling_master_file='ag_commodities_milling_mf.csv'
agricultural_commodities_milling_directory = 'agricultural_commodities_milling'
get_industry_ticker_data(agricultural_commodities_milling ,agricultural_commodities_milling_directory)
create_industry_master_list(agricultural_commodities_milling ,agricultural_commodities_milling_master_file,agricultural_commodities_milling_directory)
print("-> Master file created, filename: %s\n" % agricultural_commodities_milling_master_file)
purge_industry_low_gain_volume(agricultural_commodities_milling_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
air_freight_couriers = ['AAWW', 'AIRT', 'BEST', 'CHRW', 'CYRX', 'EXPD', 'FDX', 'FWRD', 'HUBG', 'RLGT', 'STCN', 'UPS', 'ZTO']
air_freight_couriers_master_file='air_freight_couriers_mf.csv'
air_freight_couriers_directory = 'air_freight_couriers'
get_industry_ticker_data(air_freight_couriers ,air_freight_couriers_directory)
create_industry_master_list(air_freight_couriers ,air_freight_couriers_master_file,air_freight_couriers_directory)
print("-> Master file created, filename: %s\n" % air_freight_couriers_master_file)
purge_industry_low_gain_volume(air_freight_couriers_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
airlines = ['AAL', 'ALGT', 'ALK', 'ATSG', 'AZUL', 'CEA', 'CPA', 'DAL', 'GOL', 'HA', 'JBLU', 'LTM', 'LUV', 'MESA', 'RYAAY', 'SAVE', 'SKYW', 'UAL', 'VLRS', 'ZNH']
airlines_master_file='airlines_mf.csv'
airlines_directory = 'airlines'
get_industry_ticker_data(airlines ,airlines_directory)
create_industry_master_list(airlines ,airlines_master_file,airlines_directory)
print("-> Master file created, filename: %s\n" % airlines_master_file)
purge_industry_low_gain_volume(airlines_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
alternative_power_generation = ['AY', 'AZRE', 'BIP', 'ELLO', 'JHB', 'NEP', 'ORA', 'SKYS', 'TERP']
alternative_power_generation_master_file='alt_power_generation_mf.csv'
alternative_power_generation_directory = 'alternative_power_generation'
get_industry_ticker_data(alternative_power_generation ,alternative_power_generation_directory)
create_industry_master_list(alternative_power_generation ,alternative_power_generation_master_file,alternative_power_generation_directory)
print("-> Master file created, filename: %s\n" % alternative_power_generation_master_file)
purge_industry_low_gain_volume(alternative_power_generation_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
aluminum = ['AA', 'ACH', 'ARNC', 'CENX', 'HWM', 'KALU']
aluminum_master_file='aluminum_mf.csv'
aluminum_directory = 'aluminum'
get_industry_ticker_data(aluminum ,aluminum_directory)
create_industry_master_list(aluminum ,aluminum_master_file,aluminum_directory)
print("-> Master file created, filename: %s\n" % aluminum_master_file)
purge_industry_low_gain_volume(aluminum_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
apparel_footwear = ['APEX', 'CAL', 'COLM', 'CROX', 'DECK', 'DLA', 'EVK', 'GES', 'GIII', 'GIL', 'GOOS', 'HBI', 'ICON', 'JRSH', 'KBSF', 'KTB', 'LEVI', 'NAKD', 'NKE', 'OXM', 'PVH', 'RCKY', 'RUHN', 'SGC', 'SHOO', 'SKX', 'SQBG', 'TBLT', 'UA', 'UAA', 'VFC', 'VNCE', 'VRA', 'WWW', 'XELB']
apparel_footwear_master_file='apparel_footwear_mf.csv'
apparel_footwear_directory = 'apparel_footwear'
get_industry_ticker_data(apparel_footwear ,apparel_footwear_directory)
create_industry_master_list(apparel_footwear ,apparel_footwear_master_file,apparel_footwear_directory)
print("-> Master file created, filename: %s\n" % apparel_footwear_master_file)
purge_industry_low_gain_volume(apparel_footwear_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
apparel_footwear_retail =['AEO', 'ANF', 'ASNA', 'BKE', 'BOOT', 'BURL', 'CATO', 'CHS', 'CPRI', 'CRI', 'CTRN', 'DBI', 'DXLG', 'EXPR', 'FL', 'FRAN', 'GPS', 'JILL', 'JWN', 'LB', 'LULU', 'MOGU', 'PLCE', 'RL', 'ROST', 'RTW', 'SCVL', 'SECO', 'SFIX', 'SMRT', 'TJX', 'TLRD', 'TLYS', 'TPR', 'URBN', 'ZUMZ']
apparel_footwear_retail_master_file='apparel_footwear_retail_mf.csv'
apparel_footwear_retail_directory = 'apparel_footwear_retail'
get_industry_ticker_data(apparel_footwear_retail ,apparel_footwear_retail_directory)
create_industry_master_list(apparel_footwear_retail ,apparel_footwear_retail_master_file,apparel_footwear_retail_directory)
print("-> Master file created, filename: %s\n" % apparel_footwear_retail_master_file)
purge_industry_low_gain_volume(apparel_footwear_retail_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
auto_after = ['CTB', 'DORM', 'GT', 'HZN', 'LKQ', 'SMP', 'XPEL']
auto_after_master_file='auto_after_mf.csv'
auto_after_directory = 'auto_after'
get_industry_ticker_data(auto_after ,auto_after_directory)
create_industry_master_list(auto_after ,auto_after_master_file,auto_after_directory)
print("-> Master file created, filename: %s\n" % auto_after_master_file)
purge_industry_low_gain_volume(auto_after_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
auto_parts_oem = ['ADNT', 'ALV', 'APTV', 'AXL', 'AYRO', 'BWA', 'CAAS', 'CPS', 'CXDC', 'DAN', 'DLPH', 'GNTX', 'GTX', 'KIQ', 'LEA', 'LIQT', 'MGA', 'MOD', 'MPAA', 'MTOR', 'SRI', 'STRT', 'SUP', 'SYPR', 'TEN', 'THRM', 'UFAB', 'VC', 'VNE', 'WBC', 'WKHS']
auto_parts_oem_master_file='auto_parts_oem_mf.csv'
auto_parts_oem_directory = 'auto_parts_oem'
get_industry_ticker_data(auto_parts_oem ,auto_parts_oem_directory)
create_industry_master_list(auto_parts_oem ,auto_parts_oem_master_file,auto_parts_oem_directory)
print("-> Master file created, filename: %s\n" % auto_parts_oem_master_file)
purge_industry_low_gain_volume(auto_parts_oem_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
beverages_alcoholic=['ABEV', 'BF.A', 'BF.B', 'BREW', 'BUD', 'CCU', 'DEO', 'NBEV', 'SAM', 'STZ', 'STZ.B', 'TAP', 'TAP.A', 'WVVI']
beverages_alcoholic_master_file='beverages_alcoholic_mf.csv'
beverages_alcoholic_directory = 'beverages_alcoholic'
get_industry_ticker_data(beverages_alcoholic ,beverages_alcoholic_directory)
create_industry_master_list(beverages_alcoholic ,beverages_alcoholic_master_file,beverages_alcoholic_directory)
print("-> Master file created, filename: %s\n" % beverages_alcoholic_master_file)
purge_industry_low_gain_volume(beverages_alcoholic_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
beverages_non_alcoholic=['AKO.A', 'AKO.B', 'CCEP', 'CELH', 'COKE', 'FIZZ', 'FMX', 'FTFT', 'KDP', 'KO', 'KOF', 'MNST', 'PEP', 'PRMW', 'REED', 'WTER']
beverages_non_alcoholic_master_file='beverages_non_alcoholic_mf.csv'
beverages_non_alcoholic_directory = 'beverages_non_alcoholic'
get_industry_ticker_data(beverages_non_alcoholic ,beverages_non_alcoholic_directory)
create_industry_master_list(beverages_non_alcoholic ,beverages_non_alcoholic_master_file,beverages_non_alcoholic_directory)
print("-> Master file created, filename: %s\n" % beverages_non_alcoholic_master_file)
purge_industry_low_gain_volume(beverages_non_alcoholic_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
biotechnology=['A', 'ABIO', 'ABUS', 'ACAD', 'ACHV', 'ACOR', 'ADAP', 'ADMA', 'ADRO', 'ADVM', 'ADXS', 'AEZS', 'AFMD', 'AGEN', 'AGIO', 'AGTC', 'AIKI', 'AIM', 'AIMT', 'AKBA', 'ALDX', 'ALEC', 'ALIM', 'ALKS', 'ALLO', 'ALNY', 'ALPN', 'ALXN', 'AMGN', 'AMTI', 'ANAB', 'ANCN', 'ANIX', 'ANPC', 'APOP', 'APTO', 'APVO', 'AQB', 'ARAV', 'ARDX', 'ARMP', 'ARNA', 'ARVN', 'ARWR', 'ASLN', 'ASND', 'ATHE', 'ATHX', 'ATNM', 'ATRA', 'AUPH', 'AUTL', 'AVCO', 'AVEO', 'AVRO', 'AVXL', 'AXGT', 'AXLA', 'AZRX', 'BCDA', 'BCLI', 'BCRX', 'BCYC', 'BDTX', 'BEAM', 'BFRA', 'BGNE', 'BIIB', 'BLCM', 'BLU', 'BLUE', 'BMRN', 'BNGO', 'BNTC', 'BNTX', 'BPMC', 'BPTH', 'BTAI', 'CABA', 'CALA', 'CANF', 'CAPR', 'CARA', 'CASI', 'CATB', 'CBAY', 'CBIO', 'CBLI', 'CBMG', 'CBPO', 'CCXI', 'CDMO', 'CDTX', 'CFRX', 'CGEN', 'CHMA', 'CHRS', 'CLLS', 'CLRB', 'CLVS', 'CMRX', 'CNCE', 'CODX', 'CRBP', 'CRIS', 'CRSP', 'CRTX', 'CSBR', 'CTMX', 'CUE', 'CVM', 'CWBR', 'CYAD', 'CYCC', 'CYTK', 'DBVT', 'DCPH', 'DFFN', 'DMAC', 'DNLI', 'DRNA', 'DTIL', 'DVAX', 'DYAI', 'EARS', 'EBS', 'EDAP', 'EDIT', 'EIDX', 'EIGR', 'ELOX', 'ENTA', 'ENZ', 'EPZM', 'EQ', 'ERYP', 'ESPR', 'ETTX', 'EVFM', 'EVGN', 'EXAS', 'EXEL', 'FATE', 'FBIO', 'FENC', 'FGEN', 'FIXX', 'FLXN', 'FOLD', 'FPRX', 'FREQ', 'FWP', 'GALT', 'GBT', 'GENE', 'GERN', 'GILD', 'GLPG', 'GLYC', 'GNCA', 'GNFT', 'HALO', 'HARP', 'HBIO', 'HRTX', 'HTBX', 'HTGM', 'IBIO', 'IDRA', 'IFRX', 'IGC', 'ILMN', 'IMGN', 'IMMP', 'IMMU', 'IMV', 'INCY', 'INFI', 'INMB', 'INSM', 'IONS', 'IOVA', 'IPHA', 'ISEE', 'ITCI', 'ITRM', 'JNCE', 'KDMN', 'KMDA', 'KOD', 'KPTI', 'KURA', 'KZIA', 'KZR', 'LCTX', 'LGND', 'LIFE', 'LJPC', 'LMNL', 'LMNX', 'LPTX', 'LUMO', 'LXRX', 'MACK', 'MBOT', 'MCRB', 'MESO', 'MGEN', 'MGNX', 'MGTA', 'MGTX', 'MITO', 'MNKD', 'MNOV', 'MNPR', 'MNTA', 'MOR', 'MORF', 'MREO', 'MRNA', 'MRSN', 'MRTX', 'MRUS', 'MTEM', 'MTNB', 'MTP', 'MYGN', 'MYOV', 'NBIX', 'NBRV', 'NBY', 'NCNA', 'NDRA', 'NEPT', 'NERV', 'NGM', 'NK', 'NLTX', 'NMTR', 'NNVC', 'NRBO', 'NSTG', 'NTLA', 'NVAX', 'NVIV', 'NXTC', 'OCGN', 'OGEN', 'OMER', 'ONCY', 'ONTX', 'ONVO', 'OPNT', 'ORGS', 'OTIC', 'OTLK', 'PACB', 'PASG', 'PBYI', 'PDLI', 'PFNX', 'PGEN', 'PGNX', 'PHIO', 'PIRS', 'PLRX', 'PRPO', 'PRQR', 'PRTA', 'PRTK', 'PRVL', 'PSNL', 'PSTI', 'PSTV', 'PTCT', 'PTGX', 'PTI', 'PTN', 'PULM', 'QURE', 'RAPT', 'RARE', 'RCEL', 'RCKT', 'RCUS', 'RDUS', 'REGN', 'REPH', 'REPL', 'REXN', 'RGNX', 'RIGL', 'RTRX', 'RUBY', 'RYTM', 'SBBP', 'SEEL', 'SELB', 'SESN', 'SGEN', 'SGMO', 'SLDB', 'SLS', 'SMMT', 'SNCA', 'SNDX', 'SNES', 'SNGX', 'SNSS', 'SONN', 'SPPI', 'SRNE', 'SRPT', 'SRRA', 'SRRK', 'STML', 'STOK', 'SURF', 'SVRA', 'SYBX', 'SYN', 'SYRS', 'TBIO', 'TCON', 'TECH', 'TGTX', 'TLC', 'TLSA', 'TOCA', 'TORC', 'TRIL', 'TRPX', 'TRVN', 'TTOO', 'TWST', 'TXG', 'UBX', 'UMRX', 'URGN', 'VBIV', 'VBLT', 'VCNX', 'VCYT', 'VIE', 'VIR', 'VIVO', 'VKTX', 'VNDA', 'VRML', 'VRTX', 'VSTM', 'VTGN', 'VTVT', 'VXRT', 'VYGR', 'WINT', 'WVE', 'XBIO', 'XBIT', 'XENE', 'XFOR', 'XLRN', 'XNCR', 'XOMA', 'XTLB', 'XXII', 'ZIOP', 'ZLAB', 'ZNTL']
biotechnology_master_file='biotechnology_mf.csv'
biotechnology_directory = 'biotechnology'
get_industry_ticker_data(biotechnology ,biotechnology_directory)
create_industry_master_list(biotechnology ,biotechnology_master_file,biotechnology_directory)
print("-> Master file created, filename: %s\n" % biotechnology_master_file)
purge_industry_low_gain_volume(biotechnology_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
broadcasting=['BBGI', 'CETV', 'CMLS', 'ETM', 'EVC', 'FENG', 'FOX', 'FOXA', 'GTN', 'GTN.A', 'HMTV', 'IHRT', 'MDIA', 'NTN', 'NXST', 'SALM', 'SBGI', 'SGA', 'SIRI', 'SSP', 'TGNA', 'TSQ', 'TV', 'UONE', 'UONEK', 'VIAC', 'VIACA']
broadcasting_master_file='broadcasting_mf.csv'
broadcasting_directory = 'broadcasting'
get_industry_ticker_data(broadcasting ,broadcasting_directory)
create_industry_master_list(broadcasting ,broadcasting_master_file,broadcasting_directory)
print("-> Master file created, filename: %s\n" % broadcasting_master_file)
purge_industry_low_gain_volume(broadcasting_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
building_products=['AFI', 'ALLE', 'AMWD', 'AOS', 'APOG', 'AWI', 'BLD', 'BLDR', 'BMCH', 'CCCL', 'CNR', 'DOOR', 'FBHS', 'GFF', 'LII', 'LYTS', 'MAS', 'NSSC', 'NX', 'OTIS', 'PFIE', 'PGTI', 'PLPC', 'SSD', 'TGLS', 'TILE', 'WSO', 'WSO.B']
building_products_master_file='building_products_mf.csv'
building_products_directory = 'building_products'
get_industry_ticker_data(building_products ,building_products_directory)
create_industry_master_list(building_products ,building_products_master_file,building_products_directory)
print("-> Master file created, filename: %s\n" % building_products_master_file)
purge_industry_low_gain_volume(building_products_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
cable_satellite_tv=['AMCX', 'ATUS', 'CABO', 'CHTR', 'CMCSA', 'DISCA', 'DISCB', 'DISCK', 'DISH', 'LBRDA', 'LBRDK', 'LBTYA', 'LBTYB', 'LBTYK', 'LILA', 'LILAK', 'NFLX', 'SJR']
cable_satellite_tv_master_file='cable_satellite_tV_mf.csv'
cable_satellite_tv_directory = 'cable_satellite_tv'
get_industry_ticker_data(cable_satellite_tv ,cable_satellite_tv_directory)
create_industry_master_list(cable_satellite_tv ,cable_satellite_tv_master_file,cable_satellite_tv_directory)
print("-> Master file created, filename: %s\n" % cable_satellite_tv_master_file)
purge_industry_low_gain_volume(cable_satellite_tv_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
casinos_gaming=['BYD', 'CHDN', 'CNTY', 'CPHC', 'CZR', 'ERI', 'EVRI', 'FLL', 'GDEN', 'GRVY', 'IGT', 'LVS', 'MGM', 'MLCO', 'MSC', 'NWGI', 'PENN', 'SGMS', 'TRWH', 'WYNN']
casinos_gaming_master_file='casinos_gaming_mf.csv'
casinos_gaming_directory = 'casinos_gaming'
get_industry_ticker_data(casinos_gaming ,casinos_gaming_directory)
create_industry_master_list(casinos_gaming ,casinos_gaming_master_file,casinos_gaming_directory)
print("-> Master file created, filename: %s\n" % casinos_gaming_master_file)
purge_industry_low_gain_volume(casinos_gaming_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
catalog_specialty_distribution=['ARTL','ATV','BZUN','DADA','YJ']
catalog_specialty_distribution_master_file='catalog_special_distro_mf.csv'
catalog_specialty_distribution_directory = 'catalog_specialty_distribution'
get_industry_ticker_data(catalog_specialty_distribution ,catalog_specialty_distribution_directory)
create_industry_master_list(catalog_specialty_distribution ,catalog_specialty_distribution_master_file,catalog_specialty_distribution_directory)
print("-> Master file created, filename: %s\n" % catalog_specialty_distribution_master_file)
purge_industry_low_gain_volume(catalog_specialty_distribution_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
chemicals_agricultural=['AGFS', 'AVD', 'CF', 'CGA', 'CTVA', 'FMC', 'ICL', 'LAC', 'LXU', 'MBII', 'MOS', 'NTR', 'ODC', 'PLL', 'RKDA', 'SMG', 'SQM', 'UAN', 'YTEN']
chemicals_agricultural_master_file='chemicals_agricultural_mf.csv'
chemicals_agricultural_directory = 'chemicals_agricultural'
get_industry_ticker_data(chemicals_agricultural ,chemicals_agricultural_directory)
create_industry_master_list(chemicals_agricultural ,chemicals_agricultural_master_file,chemicals_agricultural_directory)
print("-> Master file created, filename: %s\n" % chemicals_agricultural_master_file)
purge_industry_low_gain_volume(chemicals_agricultural_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
chemicals_major_diversified=['CE', 'DD', 'EMN', 'ESI', 'FSI', 'HUN', 'IKNX', 'NGVT', 'VVV', 'WLK']
chemicals_major_diversified_master_file='chem_major_diversified_mf.csv'
chemicals_major_diversified_directory = 'chemicals_major_diversified'
get_industry_ticker_data(chemicals_major_diversified ,chemicals_major_diversified_directory)
create_industry_master_list(chemicals_major_diversified ,chemicals_major_diversified_master_file,chemicals_major_diversified_directory)
print("-> Master file created, filename: %s\n" % chemicals_major_diversified_master_file)
purge_industry_low_gain_volume(chemicals_major_diversified_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
chemicals_specialty=['ALB', 'AMRS', 'AMTX', 'APD', 'ASH', 'ASIX', 'AVTR', 'BAK', 'BCPC', 'CDXC', 'CDXS', 'CHX', 'CMP', 'DOW', 'DQ', 'DSWL', 'ECL', 'ECL.WD', 'FF', 'FFHL', 'GEVO', 'GPRE', 'GRA', 'GURE', 'IOSP', 'KOP', 'KRO', 'KWR', 'LIN', 'LOOP', 'LTHM', 'LYB', 'MEOH', 'MTX', 'MYT', 'NEU', 'OEC', 'PEIX', 'POL', 'PQG', 'REGI', 'REX', 'ROG', 'RYAM', 'SCL', 'SHI', 'SSL', 'TANH', 'TROX', 'VHI', 'VNTR', 'WLKP']
chemicals_specialty_master_file='chemicals_specialty_mf.csv'
chemicals_specialty_directory = 'chemicals_specialty'
get_industry_ticker_data(chemicals_specialty ,chemicals_specialty_directory)
create_industry_master_list(chemicals_specialty ,chemicals_specialty_master_file,chemicals_specialty_directory)
print("-> Master file created, filename: %s\n" % chemicals_specialty_master_file)
purge_industry_low_gain_volume(chemicals_specialty_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
coal=['ARCH', 'AREC', 'ARLP', 'BTU', 'CCR', 'CEIX', 'CTRA', 'HCC', 'HNRG', 'METC', 'NC', 'NRP', 'SXC']
coal_master_file='coal_mf.csv'
coal_directory = 'coal'
get_industry_ticker_data(coal ,coal_directory)
create_industry_master_list(coal ,coal_master_file,coal_directory)
print("-> Master file created, filename: %s\n" % coal_master_file)
purge_industry_low_gain_volume(coal_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
commercial_printing_forms=['ARC', 'CMPR', 'DLX', 'DSS', 'EBF', 'LIVX', 'MTA', 'QUAD', 'RRD', 'WMG']
commercial_printing_forms_master_file='commercial_print_form_mf.csv'
commercial_printing_forms_directory = 'commercial_printing_forms'
get_industry_ticker_data(commercial_printing_forms ,commercial_printing_forms_directory)
create_industry_master_list(commercial_printing_forms ,commercial_printing_forms_master_file,commercial_printing_forms_directory)
print("-> Master file created, filename: %s\n" % commercial_printing_forms_master_file)
purge_industry_low_gain_volume(commercial_printing_forms_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
computer_communications=['AAOI', 'ACIA', 'ALLT', 'ANET', 'FFIV', 'FTNT', 'ITI', 'OBLG', 'PANW', 'RDWR', 'SILC', 'SMCI', 'UUU', 'VCRA']
computer_communications_master_file='computer_communications_mf.csv'
computer_communications_directory = 'computer_communications'
get_industry_ticker_data(computer_communications ,computer_communications_directory)
create_industry_master_list(computer_communications ,computer_communications_master_file,computer_communications_directory)
print("-> Master file created, filename: %s\n" % computer_communications_master_file)
purge_industry_low_gain_volume(computer_communications_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
computer_peripherals=['ALOT', 'DGII', 'EXTR', 'IMMR', 'LOGI', 'NTAP', 'PSTG', 'QMCO', 'QUMU', 'SCKT', 'SSYS', 'STX', 'TACT', 'WDC', 'XRX', 'ZBRA']
computer_peripherals_master_file='computer_peripherals_mf.csv'
computer_peripherals_directory = 'computer_peripherals'
get_industry_ticker_data(computer_peripherals ,computer_peripherals_directory)
create_industry_master_list(computer_peripherals ,computer_peripherals_master_file,computer_peripherals_directory)
print("-> Master file created, filename: %s\n" % computer_peripherals_master_file)
purge_industry_low_gain_volume(computer_peripherals_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
computer_processing_hardware=['CTS','DELL','HPE','HPQ','NCR','OSS','TDC']
computer_processing_hardware_master_file='computer_processing_hw_mf.csv'
computer_processing_hardware_directory = 'computer_processing_hardware'
get_industry_ticker_data(computer_processing_hardware ,computer_processing_hardware_directory)
create_industry_master_list(computer_processing_hardware ,computer_processing_hardware_master_file,computer_processing_hardware_directory)
print("-> Master file created, filename: %s\n" % computer_processing_hardware_master_file)
purge_industry_low_gain_volume(computer_processing_hardware_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
construction_materials=['CPAC', 'CRH', 'CVIA', 'CX', 'EXP', 'FRTA', 'HHT', 'JHX', 'LOMA', 'MLM', 'SND', 'SUM', 'USCR', 'USLM', 'VMC']
construction_materials_master_file='construction_materials_mf.csv'
construction_materials_directory = 'construction_materials'
get_industry_ticker_data(construction_materials ,construction_materials_directory)
create_industry_master_list(construction_materials ,construction_materials_master_file,construction_materials_directory)
print("-> Master file created, filename: %s\n" % construction_materials_master_file)
purge_industry_low_gain_volume(construction_materials_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
consumer_sundries=['CENT','CENTA','DOGZ']
consumer_sundries_master_file='consumer_sundries_mf.csv'
consumer_sundries_directory = 'consumer_sundries'
get_industry_ticker_data(consumer_sundries ,consumer_sundries_directory)
create_industry_master_list(consumer_sundries ,consumer_sundries_master_file,consumer_sundries_directory)
print("-> Master file created, filename: %s\n" % consumer_sundries_master_file)
purge_industry_low_gain_volume(consumer_sundries_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
containers_packaging=['AMCR', 'ARD', 'ATR', 'AVY', 'BERY', 'BLL', 'CCK', 'GEF', 'GEF.B', 'GPK', 'IP', 'MYE', 'OI', 'PACK', 'PKG', 'SEE', 'SLGN', 'SON', 'UFPT', 'WRK']
containers_packaging_master_file='containers_packaging_mf.csv'
containers_packaging_directory = 'containers_packaging'
get_industry_ticker_data(containers_packaging ,containers_packaging_directory)
create_industry_master_list(containers_packaging ,containers_packaging_master_file,containers_packaging_directory)
print("-> Master file created, filename: %s\n" % containers_packaging_master_file)
purge_industry_low_gain_volume(containers_packaging_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
contract_drilling=['BORR', 'HP', 'ICD', 'NBR', 'NE', 'PACD', 'PDS', 'PTEN', 'QES', 'RIG', 'SDRL', 'VAL']
contract_drilling_master_file='contract_drilling_mf.csv'
contract_drilling_directory = 'contract_drilling'
get_industry_ticker_data(contract_drilling ,contract_drilling_directory)
create_industry_master_list(contract_drilling ,contract_drilling_master_file,contract_drilling_directory)
print("-> Master file created, filename: %s\n" % contract_drilling_master_file)
purge_industry_low_gain_volume(contract_drilling_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
data_processing_services=['ADP', 'ADS', 'AMRH', 'BCOR', 'CASS', 'CCRC', 'CLGX', 'CSGS', 'CSLT', 'DBX', 'DWSN', 'DXC', 'EEFT', 'EVTC', 'EXLS', 'FDS', 'FIS', 'FISV', 'FLNT', 'FVRR', 'GPN', 'INOD', 'INOV', 'ISDR', 'KERN', 'MOXC', 'MRIN', 'NH', 'OPRX', 'PAYX', 'PCYG', 'PFPT', 'PYPL', 'QIWI', 'RAMP', 'TRHC', 'TYL', 'WEX', 'WYY', 'YEXT', 'ZEN', 'ZIXI']
data_processing_services_master_file='data_processing_service_mf.csv'
data_processing_services_directory = 'data_processing_services'
get_industry_ticker_data(data_processing_services ,data_processing_services_directory)
create_industry_master_list(data_processing_services ,data_processing_services_master_file,data_processing_services_directory)
print("-> Master file created, filename: %s\n" % data_processing_services_master_file)
purge_industry_low_gain_volume(data_processing_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
department_stores=['DDS','KSS','M','OLLI']
department_stores_master_file='department_stores_mf.csv'
department_stores_directory = 'department_stores'
get_industry_ticker_data(department_stores ,department_stores_directory)
create_industry_master_list(department_stores ,department_stores_master_file,department_stores_directory)
print("-> Master file created, filename: %s\n" % department_stores_master_file)
purge_industry_low_gain_volume(department_stores_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
discount_stores=['BIG','DG','DLTR','FIVE','PSMT']
discount_stores_master_file='discount_stores_mf.csv'
discount_stores_directory = 'discount_stores'
get_industry_ticker_data(discount_stores ,discount_stores_directory)
create_industry_master_list(discount_stores ,discount_stores_master_file,discount_stores_directory)
print("-> Master file created, filename: %s\n" % discount_stores_master_file)
purge_industry_low_gain_volume(discount_stores_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
drugstore_chains=['CJJD','CVS','PETS','RAD','WBA','YI']
drugstore_chains_master_file='drugstore_chains_mf.csv'
drugstore_chains_directory = 'drugstore_chains'
get_industry_ticker_data(drugstore_chains ,drugstore_chains_directory)
create_industry_master_list(drugstore_chains ,drugstore_chains_master_file,drugstore_chains_directory)
print("-> Master file created, filename: %s\n" % drugstore_chains_master_file)
purge_industry_low_gain_volume(drugstore_chains_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
electric_utilities=['AEE', 'AEP', 'AES', 'AGR', 'ALE', 'AQN', 'AT', 'AVA', 'BEP', 'BKH', 'BWXT', 'CEPU', 'CIG', 'CIG.C', 'CMS', 'CMSA', 'CMSC', 'CNP', 'CWEN', 'CWEN.A', 'D', 'DCUE', 'DRUA', 'DTE', 'DTJ', 'DTP', 'DTW', 'DTY', 'DUK', 'DUKB', 'DUKH', 'EAI', 'EBR', 'EBR.B', 'ED', 'EDN', 'EE', 'EIX', 'ELC', 'ELP', 'ELU', 'EMP', 'ENIA', 'ENIC', 'ENJ', 'ENO', 'ES', 'ETR', 'EVRG', 'EVSI', 'EXC', 'EZT', 'FE', 'FTS', 'GNE', 'GPJA', 'HE', 'HNP', 'IDA', 'JE', 'KEN', 'KEP', 'LNT', 'MGEE', 'NEE', 'NGG', 'NRG', 'NWE', 'OGE', 'OTTR', 'PAM', 'PCG', 'PEG', 'PNM', 'PNW', 'POR', 'PPL', 'RUN', 'SO', 'SOJA', 'SOJB', 'SOJC', 'SOLN', 'SPKE', 'TAC', 'UTL', 'VSLR', 'VST', 'VVPR', 'WEC', 'XEL']
electric_utilities_master_file='electric_Utilities_mf.csv'
electric_utilities_directory = 'electric_utilities'
get_industry_ticker_data(electric_utilities ,electric_utilities_directory)
create_industry_master_list(electric_utilities ,electric_utilities_master_file,electric_utilities_directory)
print("-> Master file created, filename: %s\n" % electric_utilities_master_file)
purge_industry_low_gain_volume(electric_utilities_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
electrical_products=['ABB', 'AME', 'AMSC', 'ATKR', 'AYI', 'AZZ', 'BDC', 'BE', 'BLDP', 'CBAT', 'CPST', 'CSIQ', 'DPW', 'EAF', 'EFOI', 'EMR', 'ENS', 'ERII', 'ETN', 'GNRC', 'HUBB', 'IPWR', 'ITGR', 'JKS', 'LFUS', 'LITE', 'OEG', 'OESX', 'OPTT', 'POWL', 'PPSI', 'ROK', 'SEDG', 'SPLP', 'SPWR', 'SUNW', 'THR', 'TPIC', 'ULBI']
electrical_products_master_file='electrical_products_mf.csv'
electrical_products_directory = 'electrical_products'
get_industry_ticker_data(electrical_products ,electrical_products_directory)
create_industry_master_list(electrical_products ,electrical_products_master_file,electrical_products_directory)
print("-> Master file created, filename: %s\n" % electrical_products_master_file)
purge_industry_low_gain_volume(electrical_products_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
electronic_components=['APH', 'BELFA', 'BELFB', 'BHE', 'CAMT', 'CLS', 'CREE', 'ELTK', 'ESP', 'FLEX', 'FSLR', 'GLW', 'IEC', 'JBL', 'KEM', 'KN', 'KTCC', 'LEDS', 'LPTH', 'MEI', 'NPTN', 'NSYS', 'NVT', 'OLED', 'PLUG', 'PLXS', 'PRCP', 'RESN', 'RFIL', 'SANM', 'SENS', 'SGMA', 'SMTX', 'TEL', 'TTMI', 'VICR', 'VPG', 'VSH']
electronic_components_master_file='electronic_components_mf.csv'
electronic_components_directory = 'electronic_components'
get_industry_ticker_data(electronic_components ,electronic_components_directory)
create_industry_master_list(electronic_components ,electronic_components_master_file,electronic_components_directory)
print("-> Master file created, filename: %s\n" % electronic_components_master_file)
purge_industry_low_gain_volume(electronic_components_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
electronic_equipment_instruments=['AMOT', 'BW', 'CAJ', 'CETX', 'CGNX', 'CLIR', 'CLWT', 'COHU', 'CYBE', 'DAIO', 'DBD', 'EMAN', 'FARO', 'FCEL', 'FEIM', 'FTV', 'HOLI', 'IDN', 'IIVI', 'ISNS', 'ITRI', 'KEYS', 'KODK', 'LGL', 'LUNA', 'MKSI', 'MRCY', 'MTSC', 'NOVT', 'NTP', 'OBAS', 'PAR', 'POLA', 'ROP', 'SEAC', 'SPCB', 'ST', 'TRMB', 'TRNS', 'VIAV', 'VUZI']
electronic_equipment_instruments_master_file='electronic_equip_instru_mf.csv'
electronic_equipment_instruments_directory = 'electronic_equipment_instruments'
get_industry_ticker_data(electronic_equipment_instruments ,electronic_equipment_instruments_directory)
create_industry_master_list(electronic_equipment_instruments ,electronic_equipment_instruments_master_file,electronic_equipment_instruments_directory)
print("-> Master file created, filename: %s\n" % electronic_equipment_instruments_master_file)
purge_industry_low_gain_volume(electronic_equipment_instruments_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
electronic_production_equipment=['ACLS', 'AEHR', 'AEIS', 'AXTI', 'BRKS', 'CVV', 'DAKT', 'EMKR', 'FORM', 'IMTE', 'INTT', 'IVAC', 'KLAC', 'LPL', 'LRCX', 'MICT', 'MVIS', 'NVMI', 'ONTO', 'PI', 'PLAB', 'REFR', 'RELL', 'SGOC', 'SONO', 'TER', 'VECO', 'VERT.U', 'VRT']
electronic_production_equipment_master_file='electronic_prod_equip_mf.csv'
electronic_production_equipment_directory = 'electronic_production_equipment'
get_industry_ticker_data(electronic_production_equipment ,electronic_production_equipment_directory)
create_industry_master_list(electronic_production_equipment ,electronic_production_equipment_master_file,electronic_production_equipment_directory)
print("-> Master file created, filename: %s\n" % electronic_production_equipment_master_file)
purge_industry_low_gain_volume(electronic_production_equipment_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
electronics_appliance_stores=['BBY','CONN','TWMC']
electronics_appliance_stores_master_file='electronics_appl_stores_mf.csv'
electronics_appliance_stores_directory = 'electronics_appliance_stores'
get_industry_ticker_data(electronics_appliance_stores ,electronics_appliance_stores_directory)
create_industry_master_list(electronics_appliance_stores ,electronics_appliance_stores_master_file,electronics_appliance_stores_directory)
print("-> Master file created, filename: %s\n" % electronics_appliance_stores_master_file)
purge_industry_low_gain_volume(electronics_appliance_stores_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
electronics_appliances=['ARLO', 'BNSO', 'GNSS', 'GPRO', 'HBB', 'HEAR', 'HELE', 'IRBT', 'KOSS', 'MSN', 'NPK', 'ROKU', 'SNE', 'SPB', 'UEIC', 'VIOT', 'WHR']
electronics_appliances_master_file='electronics_appliances_mf.csv'
electronics_appliances_directory = 'electronics_appliances'
get_industry_ticker_data(electronics_appliances ,electronics_appliances_directory)
create_industry_master_list(electronics_appliances ,electronics_appliances_master_file,electronics_appliances_directory)
print("-> Master file created, filename: %s\n" % electronics_appliances_master_file)
purge_industry_low_gain_volume(electronics_appliances_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
electronics_distributors=['ARW', 'AVT', 'LTRX', 'PLUS', 'RCON', 'SCSC', 'SNX', 'TAIT', 'TECD', 'WSTG']
electronics_distributors_master_file='electronics_distro_mf.csv'
electronics_distributors_directory = 'electronics_distributors'
get_industry_ticker_data(electronics_distributors ,electronics_distributors_directory)
create_industry_master_list(electronics_distributors ,electronics_distributors_master_file,electronics_distributors_directory)
print("-> Master file created, filename: %s\n" % electronics_distributors_master_file)
purge_industry_low_gain_volume(electronics_distributors_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
engineering_construction=['ACM', 'AEGN', 'AGX', 'AMRC', 'DY', 'EME', 'EXPO', 'FIX', 'FLR', 'FTEK', 'GLDD', 'GRAM', 'GV', 'GVA', 'HIL', 'IEA', 'IESC', 'J', 'KBR', 'MG', 'MTRX', 'MTZ', 'MYRG', 'NFE', 'ORN', 'PECK', 'PRIM', 'PWR', 'RCMT', 'ROAD', 'STN', 'STRL', 'TPC', 'VSEC']
engineering_construction_master_file='engineering_construct_mf.csv'
engineering_construction_directory = 'engineering_construction'
get_industry_ticker_data(engineering_construction ,engineering_construction_directory)
create_industry_master_list(engineering_construction ,engineering_construction_master_file,engineering_construction_directory)
print("-> Master file created, filename: %s\n" % engineering_construction_master_file)
purge_industry_low_gain_volume(engineering_construction_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
environmental_services=['ADSW', 'AQUA', 'AWX', 'CHRA', 'CLH', 'CVA', 'CWST', 'ECOL', 'HCCI', 'HDSN', 'JAN', 'PESI', 'QRHC', 'RSG', 'SMED', 'SRCL', 'TTEK', 'VTNR', 'WCN', 'WM']
environmental_services_master_file='environmental_services_mf.csv'
environmental_services_directory = 'environmental_services'
get_industry_ticker_data(environmental_services ,environmental_services_directory)
create_industry_master_list(environmental_services ,environmental_services_master_file,environmental_services_directory)
print("-> Master file created, filename: %s\n" % environmental_services_master_file)
purge_industry_low_gain_volume(environmental_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
finance_rental_leasing=['AAN', 'ACY', 'AER', 'AGM', 'AGM.A', 'AHCO', 'AIHS', 'AL', 'APG', 'ASFI', 'ASPS', 'ATAX', 'ATLC', 'BMRG', 'BMRG.U', 'CACC', 'CAI', 'CAR', 'CNF', 'COOP', 'CPSS', 'CURO', 'DNJR', 'DXF', 'ECPG', 'ELVT', 'ENVA', 'EZPW', 'FCFS', 'FINV', 'FLY', 'FPAY', 'GATX', 'GDOT', 'GLG', 'GMTA', 'GSKY', 'GYC', 'HEES', 'HRI', 'HTZ', 'HX', 'HYRE', 'IMH', 'IX', 'JFIN', 'LADR', 'LC', 'LX', 'MA', 'MGRC', 'MINI', 'MOGO', 'MRLN', 'NAVI', 'NICK', 'NNI', 'NSCO', 'OCN', 'OMF', 'ONDK', 'OPRT', 'PFSI', 'PYS', 'QFIN', 'R', 'RCII', 'RM', 'SC', 'SLM', 'SYF', 'TGH', 'TREE', 'TRTN', 'UHAL', 'URI', 'V', 'VEL', 'WD', 'WINA', 'WLFC', 'WRLD', 'WSC', 'WU', 'XYF']
finance_rental_leasing_master_file='finance_rental_leasing_mf.csv'
finance_rental_leasing_directory = 'finance_rental_leasing'
get_industry_ticker_data(finance_rental_leasing ,finance_rental_leasing_directory)
create_industry_master_list(finance_rental_leasing ,finance_rental_leasing_master_file,finance_rental_leasing_directory)
print("-> Master file created, filename: %s\n" % finance_rental_leasing_master_file)
purge_industry_low_gain_volume(finance_rental_leasing_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
financial_conglomerates=['ACAM', 'ACAMU', 'ACTT', 'ACTTU', 'AGBA', 'AGMH', 'ALAC', 'ALACU', 'ALUS', 'ALUS.U', 'AMCI', 'AMCIU', 'AMHC', 'AMHCU', 'AMK', 'ANDA', 'ANDAU', 'APXT', 'APXTU', 'ARYA', 'ARYAU', 'ARYBU', 'ATCX', 'AVCT', 'AXP', 'BBCP', 'BFIN', 'BIOX', 'BRBS', 'BROG', 'C', 'CCAC', 'CCAC.U', 'CCH', 'CCH.U', 'CCX', 'CCX.U', 'CCXX', 'CCXX.U', 'CFFA', 'CFFAU', 'CGROU', 'CHAQ.U', 'CHPM', 'CHPMU', 'CIIC', 'CIICU', 'CIZN', 'CNNE', 'COHN', 'CPAA', 'CPAAU', 'CRSA', 'CRSAU', 'CZWI', 'DFNS', 'DFNS.U', 'DFPH', 'DFPHU', 'DMYT', 'DMYT.U', 'DPHC', 'DPHCU', 'EFC', 'ENOB', 'EQH', 'ESSC', 'ESSCU', 'EXPC', 'EXPCU', 'EXPI', 'FBC', 'FBIZ', 'FEAC', 'FEAC.U', 'FMAO', 'FMCI', 'FMCIU', 'FPAC', 'FPAC.U', 'FSBW', 'FSRV', 'FSRVU', 'FTAC', 'FTACU', 'FVAC.U', 'GDYN', 'GFED', 'GHIV', 'GHIVU', 'GIK.U', 'GIX', 'GIX.U', 'GLEO', 'GLEO.U', 'GMHI', 'GMHIU', 'GNRS', 'GNRSU', 'GPAQ', 'GPAQU', 'GRAF', 'GRAF.U', 'GRNV', 'GSBC', 'GSMG', 'GXGX', 'GXGXU', 'HBCP', 'HCAC', 'HCACU', 'HCCH', 'HCCO', 'HCCOU', 'HUSN', 'HYAC', 'HYACU', 'IGIC', 'IMVT', 'IMVTU', 'IMXI',  'ING', 'INSU', 'INSUU', 'IPOB.U', 'IPOC.U', 'IPV', 'IPV.U', 'JFK', 'JIH', 'JIH.U', 'JWS.U', 'LACQ', 'LATN', 'LATNU', 'LCA', 'LCAHU', 'LFAC', 'LGC', 'LGVW.U', 'LHC', 'LIVKU', 'LOAC', 'LOAK', 'LOAK.U', 'LSAC', 'LSACU', 'MCACU', 'MCMJ', 'MFAC', 'MFAC.U', 'MNCL', 'NBAC', 'NBACU', 'NEBU', 'NEBUU', 'NFH', 'NFIN', 'NFINU', 'NOVSU', 'NPA', 'NPAUU', 'OAC', 'OAC.U', 'OFED', 'OPES', 'ORGO', 'ORSN', 'ORSNU', 'PAAC', 'PAACU', 'PCPL.U', 'PIC', 'PIC.U', 'PMBC', 'PRS', 'PRU', 'PTAC', 'PTACU', 'PVBC', 'RIOT', 'RMG', 'RMI', 'ROCHU', 'RPLA', 'RPLA.U', 'SAMA', 'SAMAU', 'SAQN', 'SBE', 'SBFG', 'SCPE', 'SCVX', 'SCVX.U', 'SFE', 'SFTW', 'SFTW.U', 'SHG', 'SHLL', 'SHLL.U', 'SMMC', 'SMMCU', 'SOAC.U', 'SPAQ', 'SPAQ.U', 'SRAC', 'SRL', 'SSPK', 'SSPKU', 'TBNK', 'TDAC', 'TDACU', 'THBR', 'THBRU', 'THCA', 'THCAU', 'THCB', 'TOTA', 'TRNE', 'TRNE.U', 'TZAC', 'USWS', 'WINS', 'WPF.U', 'WSBF', 'WTBA', 'ZGYH', 'ZGYHU']
financial_conglomerates_master_file='financial_conglomerates_mf.csv'
financial_conglomerates_directory = 'financial_conglomerates'
get_industry_ticker_data(financial_conglomerates ,financial_conglomerates_directory)
create_industry_master_list(financial_conglomerates ,financial_conglomerates_master_file,financial_conglomerates_directory)
print("-> Master file created, filename: %s\n" % financial_conglomerates_master_file)
purge_industry_low_gain_volume(financial_conglomerates_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
financial_publishing_services=['MCO','SPGI','TRI']
financial_publishing_services_master_file='fin_publishing_services_mf.csv'
financial_publishing_services_directory = 'financial_publishing_services'
get_industry_ticker_data(financial_publishing_services ,financial_publishing_services_directory)
create_industry_master_list(financial_publishing_services ,financial_publishing_services_master_file,financial_publishing_services_directory)
print("-> Master file created, filename: %s\n" % financial_publishing_services_master_file)
purge_industry_low_gain_volume(financial_publishing_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
food_major_diversified=['BGS', 'BYND', 'CAG', 'CPB', 'DTEA', 'GIS', 'K', 'KHC', 'LNDC', 'MDLZ', 'RELV', 'SMPL', 'THS']
food_major_diversified_master_file='food_major_diversified_mf.csv'
food_major_diversified_directory = 'food_major_diversified'
get_industry_ticker_data(food_major_diversified ,food_major_diversified_directory)
create_industry_master_list(food_major_diversified ,food_major_diversified_master_file,food_major_diversified_directory)
print("-> Master file created, filename: %s\n" % food_major_diversified_master_file)
purge_industry_low_gain_volume(food_major_diversified_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
food_meat_fish_dairy=['BRFS','HRL','LWAY','PLIN','PPC','SAFM','SEC','TSN']
food_meat_fish_dairy_master_file='food_meat_fish_dairy_mf.csv'
food_meat_fish_dairy_directory = 'food_meat_fish_dairy'
get_industry_ticker_data(food_meat_fish_dairy ,food_meat_fish_dairy_directory)
create_industry_master_list(food_meat_fish_dairy ,food_meat_fish_dairy_master_file,food_meat_fish_dairy_directory)
print("-> Master file created, filename: %s\n" % food_meat_fish_dairy_master_file)
purge_industry_low_gain_volume(food_meat_fish_dairy_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
food_specialty_candy=['BRID', 'CALM', 'CVGW', 'FARM', 'FDP', 'FLO', 'FRPT', 'HAIN', 'HSY', 'JBSS', 'JJSF', 'JVA', 'LANC', 'LK', 'LW', 'MKC', 'MKC.V', 'NATR', 'NOMD', 'PETZ', 'PLAG', 'POST', 'RIBT', 'RMCF', 'SENEA', 'SJM', 'STKL', 'TR', 'TWNK']
food_specialty_candy_master_file='food_specialty_candy_mf.csv'
food_specialty_candy_directory = 'food_specialty_candy'
get_industry_ticker_data(food_specialty_candy ,food_specialty_candy_directory)
create_industry_master_list(food_specialty_candy ,food_specialty_candy_master_file,food_specialty_candy_directory)
print("-> Master file created, filename: %s\n" % food_specialty_candy_master_file)
purge_industry_low_gain_volume(food_specialty_candy_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
food_distributors=['CHEF', 'HFFG', 'PFGC', 'SPTN', 'SYY', 'UNFI', 'USFD', 'WILC']
food_distributors_master_file='food_distributors_mf.csv'
food_distributors_directory = 'food_distributors'
get_industry_ticker_data(food_distributors ,food_distributors_directory)
create_industry_master_list(food_distributors ,food_distributors_master_file,food_distributors_directory)
print("-> Master file created, filename: %s\n" % food_distributors_master_file)
purge_industry_low_gain_volume(food_distributors_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
food_retail=['CBD', 'FAMI', 'GO', 'IFMK', 'IMKTA', 'KR', 'NGVC', 'SFM', 'VLGEA', 'WMK', 'WMT']
food_retail_master_file='food_retail_mf.csv'
food_retail_directory = 'food_retail'
get_industry_ticker_data(food_retail ,food_retail_directory)
create_industry_master_list(food_retail ,food_retail_master_file,food_retail_directory)
print("-> Master file created, filename: %s\n" % food_retail_master_file)
purge_industry_low_gain_volume(food_retail_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
forest_products=['BCC','JELD','LPX','OSB','TREX','UFPI']
forest_products_master_file='forest_products_mf.csv'
forest_products_directory = 'forest_products'
get_industry_ticker_data(forest_products ,forest_products_directory)
create_industry_master_list(forest_products ,forest_products_master_file,forest_products_directory)
print("-> Master file created, filename: %s\n" % forest_products_master_file)
purge_industry_low_gain_volume(forest_products_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
gas_distributors=['ATO', 'BIPC', 'CLNE', 'CPK', 'MDU', 'NI', 'NJR', 'NWN', 'OGS', 'RGCO', 'SGU', 'SJI', 'SJIU', 'SMLP', 'SPH', 'SR', 'SRE', 'SWX', 'UGI']
gas_distributors_master_file='gas_distributors_mf.csv'
gas_distributors_directory = 'gas_distributors'
get_industry_ticker_data(gas_distributors ,gas_distributors_directory)
create_industry_master_list(gas_distributors ,gas_distributors_master_file,gas_distributors_directory)
print("-> Master file created, filename: %s\n" % gas_distributors_master_file)
purge_industry_low_gain_volume(gas_distributors_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
home_furnishings=['BSET', 'CRWS', 'CSPR', 'DXYN', 'ETH', 'FLXS', 'HOFT', 'KBAL', 'LCUT', 'LEG', 'LZB', 'MHK', 'NTZ', 'NVFY', 'OC', 'SNBR', 'TPX']
home_furnishings_master_file='home_furnishings_mf.csv'
home_furnishings_directory = 'home_furnishings'
get_industry_ticker_data(home_furnishings ,home_furnishings_directory)
create_industry_master_list(home_furnishings ,home_furnishings_master_file,home_furnishings_directory)
print("-> Master file created, filename: %s\n" % home_furnishings_master_file)
purge_industry_low_gain_volume(home_furnishings_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
home_improvement_chains=['FND','HD','LOW']
home_improvement_chains_master_file='home_improvement_chains_mf.csv'
home_improvement_chains_directory = 'home_improvement_chains'
get_industry_ticker_data(home_improvement_chains ,home_improvement_chains_directory)
create_industry_master_list(home_improvement_chains ,home_improvement_chains_master_file,home_improvement_chains_directory)
print("-> Master file created, filename: %s\n" % home_improvement_chains_master_file)
purge_industry_low_gain_volume(home_improvement_chains_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
homebuilding=['BZH', 'CCS', 'CHCI', 'CVCO', 'DHI', 'DRTT', 'FTDR', 'GRBK', 'HGSH', 'HOV', 'KBH', 'LEGH', 'LEN', 'LEN.B', 'LGIH', 'MDC', 'MHO', 'MTH', 'NVR', 'NWHM', 'PATK', 'PHM', 'SKY', 'TMHC', 'TOL', 'TPH']
homebuilding_master_file='homebuilding_mf.csv'
homebuilding_directory = 'homebuilding'
get_industry_ticker_data(homebuilding ,homebuilding_directory)
create_industry_master_list(homebuilding ,homebuilding_master_file,homebuilding_directory)
print("-> Master file created, filename: %s\n" % homebuilding_master_file)
purge_industry_low_gain_volume(homebuilding_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
hospital_nursing_management=['BKD', 'CYH', 'EHC', 'FVE', 'GEN', 'HCA', 'JYNT', 'NHC', 'PINC', 'SEM', 'SGRY', 'SSY', 'TDOC', 'THC', 'UHS', 'VMD']
hospital_nursing_management_master_file='hospital_nursing_mgmt_mf.csv'
hospital_nursing_management_directory = 'hospital_nursing_management'
get_industry_ticker_data(hospital_nursing_management ,hospital_nursing_management_directory)
create_industry_master_list(hospital_nursing_management ,hospital_nursing_management_master_file,hospital_nursing_management_directory)
print("-> Master file created, filename: %s\n" % hospital_nursing_management_master_file)
purge_industry_low_gain_volume(hospital_nursing_management_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
hotels_resorts_cruise_lines=['CCL', 'CHH', 'CUK', 'CVEO', 'GHG', 'H', 'HLT', 'HTHT', 'IHG', 'MAR', 'MCRI', 'MTN', 'NCLH', 'PLYA', 'RCL', 'RLH', 'RRR', 'STAY', 'TH', 'WH', 'WYND']
hotels_resorts_cruise_lines_master_file='hotels_resorts_cruise_mf.csv'
hotels_resorts_cruise_lines_directory = 'hotels_resorts_cruise_lines'
get_industry_ticker_data(hotels_resorts_cruise_lines ,hotels_resorts_cruise_lines_directory)
create_industry_master_list(hotels_resorts_cruise_lines ,hotels_resorts_cruise_lines_master_file,hotels_resorts_cruise_lines_directory)
print("-> Master file created, filename: %s\n" % hotels_resorts_cruise_lines_master_file)
purge_industry_low_gain_volume(hotels_resorts_cruise_lines_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
household_personal_care=['CHD', 'CL', 'CLX', 'COTY', 'EL', 'ELF', 'ENR', 'EPC', 'IFF', 'IFFT', 'IPAR', 'KMB', 'NHTC', 'NTCO', 'NUS', 'OBCI', 'PG', 'REV', 'UG', 'UL', 'UN', 'USNA']
household_personal_care_master_file='household_personal_care_mf.csv'
household_personal_care_directory = 'household_personal_care'
get_industry_ticker_data(household_personal_care ,household_personal_care_directory)
create_industry_master_list(household_personal_care ,household_personal_care_master_file,household_personal_care_directory)
print("-> Master file created, filename: %s\n" % household_personal_care_master_file)
purge_industry_low_gain_volume(household_personal_care_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
industrial_conglomerates=['AGS', 'HI', 'HON', 'IEP', 'MMM', 'NL', 'NWL', 'SPXC', 'TT', 'WRTC', 'ZKIN']
industrial_conglomerates_master_file='industrial_conglomerate_mf.csv'
industrial_conglomerates_directory = 'industrial_conglomerates'
get_industry_ticker_data(industrial_conglomerates ,industrial_conglomerates_directory)
create_industry_master_list(industrial_conglomerates ,industrial_conglomerates_master_file,industrial_conglomerates_directory)
print("-> Master file created, filename: %s\n" % industrial_conglomerates_master_file)
purge_industry_low_gain_volume(industrial_conglomerates_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
industrial_machinery=['AAON', 'ACMR', 'ADES', 'AIMC', 'AMAT', 'ASML', 'ASYS', 'B', 'BHTG', 'BIMI', 'BMI', 'BWEN', 'CARR', 'CECE', 'CFX', 'CFXA', 'CIR', 'CMT', 'CREG', 'DDD', 'ELSE', 'EPAC', 'FELE', 'FLOW', 'FLS', 'GENC', 'GGG', 'GHM', 'GRC', 'GTEC', 'GTES', 'GTLS', 'HEBT', 'HLIO', 'HURC', 'IEX', 'IR', 'ITT', 'ITW', 'JBT', 'KAI', 'KMT', 'KRNT', 'LECO', 'LXFR', 'MEC', 'MIDD', 'MWA', 'NDSN', 'NEWA', 'NNDM', 'NPO', 'OFLX', 'PFIN', 'PH', 'PPIH', 'PRLB', 'RBC', 'RXN', 'SMIT', 'TAYD', 'TGEN', 'TNC', 'VJET', 'VTSI', 'WBT', 'WTS', 'WWD', 'XONE', 'XYL']
industrial_machinery_master_file='industrial_machinery_mf.csv'
industrial_machinery_directory = 'industrial_machinery'
get_industry_ticker_data(industrial_machinery ,industrial_machinery_directory)
create_industry_master_list(industrial_machinery ,industrial_machinery_master_file,industrial_machinery_directory)
print("-> Master file created, filename: %s\n" % industrial_machinery_master_file)
purge_industry_low_gain_volume(industrial_machinery_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
industrial_specialties=['ASPN', 'AXTA', 'CBT', 'CC', 'CCF', 'CLMT', 'CSWI', 'CTIB', 'DCI', 'EML', 'EVA', 'FOE', 'FUL', 'GCP', 'GSM', 'KRA', 'LAKE', 'LDL', 'NTIC', 'OLN', 'PPG', 'RPM', 'SHW', 'TSE', 'WDFC', 'ZAGG']
industrial_specialties_master_file='industrial_specialties_mf.csv'
industrial_specialties_directory = 'industrial_specialties'
get_industry_ticker_data(industrial_specialties ,industrial_specialties_directory)
create_industry_master_list(industrial_specialties ,industrial_specialties_master_file,industrial_specialties_directory)
print("-> Master file created, filename: %s\n" % industrial_specialties_master_file)
purge_industry_low_gain_volume(industrial_specialties_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
information_technology_services=['ACN', 'ALRM', 'ALYA', 'ANY', 'APPF', 'ATEN', 'BKYI', 'BLIN', 'BOX', 'BRQS', 'CACI', 'CDW', 'CERN', 'CLPS', 'CNXN', 'CPSI', 'CSCO', 'CSPI', 'CTEK', 'CTG', 'CTSH', 'DAVA', 'DMRC', 'DOX', 'DSGX', 'DTSS', 'DUOT', 'EIGI', 'EPAM', 'EVOL', 'FEYE', 'FIVN', 'FORTY', 'FSCT', 'FSLY', 'GDS', 'GIB', 'GLOB', 'GWRE', 'HUBS', 'IBM', 'IDEX', 'INFY', 'INPX', 'INUV', 'JKHY', 'JNPR', 'LDOS', 'LINX', 'MDRX', 'MFH', 'MITK', 'MIXT', 'MOBL', 'MODN', 'MOMO', 'MTBC', 'MTC', 'MTSL', 'NEON', 'NET', 'NEWR', 'NICE', 'NOW', 'NSIT', 'NTCT', 'NTGR', 'NTWK', 'NXGN', 'OMCL', 'OOMA', 'OSPN', 'PDFS', 'PEGA', 'PRO', 'PS', 'PSN', 'QADA', 'QADB', 'QTWO', 'RMBL', 'RP', 'RPD', 'SAIC', 'SFET', 'SPLK', 'SQ', 'SREV', 'SSNC', 'SSNT', 'STRM', 'SVMK', 'SWCH', 'SYKE', 'SYNA', 'SYNC', 'TAOP', 'TNAV', 'TRU', 'TZOO', 'UIS', 'UPLD', 'VMW', 'VRNS', 'VRNT', 'VRTU', 'WDAY', 'WIFI', 'WIT', 'WIX', 'WK', 'WKEY', 'XELA', 'Z', 'ZG']
information_technology_services_master_file='info_tech_services_mf.csv'
information_technology_services_directory = 'information_technology_services'
get_industry_ticker_data(information_technology_services ,information_technology_services_directory)
create_industry_master_list(information_technology_services ,information_technology_services_master_file,information_technology_services_directory)
print("-> Master file created, filename: %s\n" % information_technology_services_master_file)
purge_industry_low_gain_volume(information_technology_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
insurance_brokers_services=['AJG', 'AON', 'BFYT', 'BRO', 'CRD.A', 'CRD.B', 'CRVL', 'EBIX', 'ESGR', 'GSHD', 'HUIZ', 'KINS', 'MJCO', 'MMC', 'VRSK', 'WLTW']
insurance_brokers_services_master_file='insur_brokers_services_mf.csv'
insurance_brokers_services_directory = 'insurance_brokers_services'
get_industry_ticker_data(insurance_brokers_services ,insurance_brokers_services_directory)
create_industry_master_list(insurance_brokers_services ,insurance_brokers_services_master_file,insurance_brokers_services_directory)
print("-> Master file created, filename: %s\n" % insurance_brokers_services_master_file)
purge_industry_low_gain_volume(insurance_brokers_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
integrated_oil=['AMPY', 'AXAS', 'BP', 'CNX', 'COG', 'CVX', 'E', 'EC', 'EQNR', 'ESTE', 'FLMN', 'GDP', 'HESM', 'IMO', 'INDO', 'NFG', 'PBR', 'PBR.A', 'PRT', 'PTR', 'PVL', 'RDS.A', 'RDS.B', 'SNDE', 'SNMP', 'SNP', 'SU', 'TALO', 'TELL', 'TOT', 'VNOM', 'XOM', 'YPF']
integrated_oil_master_file='integrated_oil_mf.csv'
integrated_oil_directory = 'integrated_oil'
get_industry_ticker_data(integrated_oil ,integrated_oil_directory)
create_industry_master_list(integrated_oil ,integrated_oil_master_file,integrated_oil_directory)
print("-> Master file created, filename: %s\n" % integrated_oil_master_file)
purge_industry_low_gain_volume(integrated_oil_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
internet_retail=['AMZN', 'APRN', 'BABA', 'CHWY', 'DLTH', 'FLWS', 'IMBI', 'JD', 'LE', 'LITB', 'LIVE', 'OSTK', 'PDD', 'PRTS', 'QRTEA', 'QRTEB', 'QVCD', 'REAL', 'RVLV', 'SHOP', 'SYX', 'TC', 'VIPS', 'W', 'WTRH', 'YGYI']
internet_retail_master_file='internet_retail_mf.csv'
internet_retail_directory = 'internet_retail'
get_industry_ticker_data(internet_retail ,internet_retail_directory)
create_industry_master_list(internet_retail ,internet_retail_master_file,internet_retail_directory)
print("-> Master file created, filename: %s\n" % internet_retail_master_file)
purge_industry_low_gain_volume(internet_retail_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
internet_software_services=['AEYE', 'AKAM', 'ANGI', 'APPS', 'ARCE', 'ATHM', 'BCOV', 'BIDU', 'BILI', 'BITA', 'CARG', 'CARS', 'CDLX', 'CHKP', 'CIH', 'COE', 'CRTO', 'CSGP', 'CYRN', 'DHX', 'DKNG', 'DNK', 'DUO', 'EGAN', 'EGOV', 'EVER', 'FB', 'FTCH', 'FUTU', 'GDDY', 'GLUU', 'GOOG', 'GOOGL', 'GSX', 'HHR', 'IAC', 'IQ', 'JCOM', 'JFU', 'JG', 'JMIA', 'JOBS', 'JRJC', 'KRKR', 'LEAF', 'LLNW', 'LMPX', 'LOGM', 'LPSN', 'MCHX', 'MEET', 'MELI', 'MKD', 'MSTR', 'MTCH', 'MYSZ', 'NETE', 'NTES', 'PERI', 'PFSW', 'PINS', 'QD', 'QTT', 'RENN', 'SE', 'SFUN', 'SHSP', 'SIFY', 'SINA', 'SNAP', 'SNCR', 'SOGO', 'SOHU', 'SPOT', 'SSTK', 'SY', 'TCX', 'TME', 'TRUE', 'TRVG', 'TTGT', 'TW', 'TWTR', 'UXIN', 'VALU', 'VHC', 'VNET', 'VRSN', 'WB', 'WBAI', 'WEI', 'WUBA', 'XNET', 'XP', 'XRF', 'YELP', 'YNDX', 'YRD', 'YY', 'ZI', 'ZNGA']
internet_software_services_master_file='internet_sw_services_mf.csv'
internet_software_services_directory = 'internet_software_services'
get_industry_ticker_data(internet_software_services ,internet_software_services_directory)
create_industry_master_list(internet_software_services ,internet_software_services_master_file,internet_software_services_directory)
print("-> Master file created, filename: %s\n" % internet_software_services_master_file)
purge_industry_low_gain_volume(internet_software_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
investment_banks_brokers=['AMP', 'AMRK', 'AMTD', 'BGCP', 'CBOE', 'CME', 'COWN', 'ETFC', 'EVR', 'FRHC', 'GCAP', 'GHL', 'GS', 'HKIB', 'HLI', 'IBKR', 'ICE', 'INTL', 'JEF', 'JMP', 'JMPNL', 'LAZ', 'LPLA', 'LYL', 'MC', 'MKTX', 'MS', 'NDAQ', 'NHLD', 'NMR', 'NOAH', 'OPY', 'PACQ', 'PIPR', 'PJT', 'RILY', 'RILYI', 'RJF', 'SCHW', 'SF', 'SIEB', 'TIGR', 'VIRT', 'WHG', 'YIN']
investment_banks_brokers_master_file='investment_bank_brokers_mf.csv'
investment_banks_brokers_directory = 'investment_banks_brokers'
get_industry_ticker_data(investment_banks_brokers ,investment_banks_brokers_directory)
create_industry_master_list(investment_banks_brokers ,investment_banks_brokers_master_file,investment_banks_brokers_directory)
print("-> Master file created, filename: %s\n" % investment_banks_brokers_master_file)
purge_industry_low_gain_volume(investment_banks_brokers_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
investment_managers=['AAMC', 'AB', 'AC', 'ACRE', 'AINC', 'AINV', 'AJX', 'AJXA', 'AMG', 'APAM', 'APO', 'ARCC', 'ARES', 'ATCO', 'ATIF', 'BAM', 'BBDC', 'BEN', 'BKCC', 'BLK', 'BSIG', 'BX', 'CCAP', 'CG', 'CGBD', 'CNS', 'CODI', 'CORR', 'CPTA', 'CSWC', 'DHIL', 'EQS', 'EV', 'FDUS', 'FHI', 'FOCS', 'FSK', 'FTAI', 'GAIN', 'GBDC', 'GBL', 'GECC', 'GLAD', 'GLADD', 'GRNQ', 'GROW', 'GSBD', 'HCXY', 'HCXZ', 'HLNE', 'HNNA', 'HQY', 'HTGC', 'ICMB', 'IFS', 'INFO', 'IVZ', 'JHG', 'JP', 'JT', 'KKR', 'KYN', 'LM', 'LMHA', 'LMHB', 'MAIN', 'MCI', 'MDLQ', 'MDLX', 'MDLY', 'MFIN', 'MIC', 'MN', 'MORN', 'MRCC', 'MRCCL', 'MSCI', 'MVC', 'MVCD', 'MYFW', 'NMFC', 'OCFT', 'OCSL', 'OFS', 'OFSSZ', 'ORCC', 'OXSQ', 'PBB', 'PBY', 'PFG', 'PFLT', 'PHCF', 'PNNT', 'PSEC', 'PTMN', 'PUYI', 'PZN', 'RAND', 'RFM', 'SAF', 'SAMG', 'SAR', 'SCA', 'SCM', 'SCU', 'SEIC', 'SLRC', 'SSSS', 'SUNS', 'SVVC', 'TCRD', 'TCRZ', 'TPVG', 'TPVY', 'TROW', 'TSLX', 'TURN', 'VCTR', 'VRTS', 'WDR', 'WETF', 'WHF', 'WHFBZ']
investment_managers_master_file='investment_managers_mf.csv'
investment_managers_directory = 'investment_managers'
get_industry_ticker_data(investment_managers ,investment_managers_directory)
create_industry_master_list(investment_managers ,investment_managers_master_file,investment_managers_directory)
print("-> Master file created, filename: %s\n" % investment_managers_master_file)
purge_industry_low_gain_volume(investment_managers_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
investment_trusts_mutual_funds=['AAAU','AADR', 'AAXJ', 'ABEQ', 'ACES', 'ACIO', 'ACP', 'ACSG', 'ACSI', 'ACT', 'ACV', 'ACWF', 'ACWI', 'ACWV', 'ACWX', 'ADME', 'ADRE', 'ADX', 'AEF', 'AESR', 'AFB', 'AFIF', 'AFK', 'AFLG', 'AFMC', 'AFSM', 'AFT', 'AFTY', 'AGD', 'AGG', 'AGGP', 'AGGY', 'AGQ', 'AGT', 'AGZ', 'AGZD', 'AIA', 'AIEQ', 'AIF', 'AIIQ', 'AIO', 'AIQ', 'AIRR', 'ALFA', 'ALTS', 'ALTY', 'AMCA', 'AMJ', 'AMLP', 'AMOM', 'AMU', 'AMUB', 'AMZA', 'ANGL', 'AOA', 'AOD', 'AOK', 'AOM', 'AOR', 'ARB', 'ARCM', 'ARDC', 'ARGT', 'ARKF', 'ARKG', 'ARKK', 'ARKQ', 'ARKW', 'ARMR', 'ASA', 'ASEA', 'ASET', 'ASG', 'ASHR', 'ASHS', 'ASHX', 'ATMP', 'AUSF', 'AVDE', 'AVDV', 'AVEM', 'AVK', 'AVUS', 'AVUV', 'AWAY', 'AWF', 'AWP', 'AWTM', 'AZAA', 'AZBA', 'BAB', 'BAF', 'BAL', 'BANX', 'BAPR', 'BAR', 'BATT', 'BAUG', 'BBAX', 'BBC', 'BBCA', 'BBEU', 'BBF', 'BBH', 'BBIN', 'BBJP', 'BBK', 'BBMC', 'BBN', 'BBP', 'BBRE', 'BBSA', 'BBUS', 'BCD', 'BCI', 'BCM', 'BCSF', 'BCV', 'BCX', 'BDCS', 'BDCX', 'BDCY', 'BDCZ', 'BDEC', 'BDJ', 'BDRY', 'BETZ', 'BFEB', 'BFIT', 'BFK', 'BFO', 'BFOR', 'BFY', 'BFZ', 'BGB', 'BGH', 'BGIO', 'BGR', 'BGRN', 'BGT', 'BGX', 'BGY', 'BHK', 'BHV', 'BIB', 'BIBL', 'BICK', 'BIF', 'BIL', 'BIS', 'BIT', 'BIV', 'BIZD', 'BJAN', 'BJK', 'BJUL', 'BJUN', 'BKAG', 'BKEM', 'BKF', 'BKHY', 'BKIE', 'BKK', 'BKLC', 'BKLN', 'BKMC', 'BKN', 'BKSB', 'BKSE', 'BKT', 'BLCN', 'BLE', 'BLES', 'BLHY', 'BLOK', 'BLV', 'BLW', 'BMAR', 'BMAY', 'BME', 'BMEZ', 'BMLP', 'BND', 'BNDC', 'BNDW', 'BNDX', 'BNKD', 'BNKO', 'BNKU', 'BNKZ', 'BNO', 'BNOV', 'BNY', 'BOCT', 'BOE', 'BOIL', 'BOND', 'BOSS', 'BOTZ', 'BOUT', 'BPT', 'BQH', 'BRF', 'BRZU', 'BSAE', 'BSBE', 'BSCE', 'BSCK', 'BSCL', 'BSCM', 'BSCN', 'BSCO', 'BSCP', 'BSCQ', 'BSCR', 'BSCS', 'BSCT', 'BSD', 'BSDED', 'BSE', 'BSEP', 'BSJK', 'BSJL', 'BSJM', 'BSJN', 'BSJO', 'BSJP', 'BSJQ', 'BSJR', 'BSL', 'BSML', 'BSMM', 'BSMN', 'BSMO', 'BSMP', 'BSMQ', 'BSMR', 'BSMS', 'BSMT', 'BST', 'BSTZ', 'BSV', 'BTA', 'BTAL', 'BTEC', 'BTO', 'BTT', 'BTYS', 'BTZ', 'BUG', 'BUI', 'BUL', 'BUY', 'BUYZ', 'BVAL', 'BWG', 'BWX', 'BWZ', 'BXMX', 'BYLD', 'BYM', 'BZM', 'BZQ', 'CACG', 'CAF', 'CALF', 'CANE', 'CAPE', 'CARZ', 'CATH', 'CBH', 'CBON', 'CCD', 'CCOR', 'CDC', 'CDL', 'CEE', 'CEF', 'CEFD', 'CEFS', 'CEM', 'CEMB', 'CEN', 'CET', 'CEV', 'CEW', 'CEY', 'CEZ', 'CFA', 'CFCV', 'CFO', 'CGO', 'CGW', 'CHAD', 'CHAU', 'CHEP', 'CHGX', 'CHI', 'CHIC', 'CHIE', 'CHIH', 'CHII', 'CHIK', 'CHIL', 'CHIM', 'CHIQ', 'CHIR', 'CHIS', 'CHIU', 'CHIX', 'CHN', 'CHNA', 'CHW', 'CHY', 'CIBR', 'CID', 'CIF', 'CII', 'CIK', 'CIL', 'CIZ', 'CKX', 'CLIX', 'CLM', 'CLOU', 'CLRG', 'CLTL', 'CMBS', 'CMDY', 'CMF', 'CMU', 'CN', 'CNBS', 'CNCR', 'CNRG', 'CNXT', 'CNYA', 'COM', 'COMB', 'COMT', 'COPX', 'CORN', 'CORP', 'COW', 'COWZ', 'CPER', 'CPI', 'CPZ', 'CQQQ', 'CRAK', 'CRBN', 'CRF', 'CROC', 'CROP', 'CSA', 'CSB', 'CSD', 'CSF', 'CSM', 'CSML', 'CSQ', 'CTR', 'CUBA', 'CURE', 'CUT', 'CVY', 'CWB', 'CWEB', 'CWI', 'CWS', 'CXE', 'CXH', 'CXSE', 'CYB', 'CZA', 'DALI', 'DALT', 'DAUD', 'DAUG', 'DAX', 'DBA', 'DBAW', 'DBB', 'DBC', 'DBE', 'DBEF', 'DBEH', 'DBEM', 'DBEU', 'DBEZ', 'DBGR', 'DBJP', 'DBL', 'DBLV', 'DBMF', 'DBO', 'DBP', 'DBS', 'DBV', 'DCF', 'DCHF', 'DDF', 'DDG', 'DDIV', 'DDLS', 'DDM', 'DDWM', 'DEED', 'DEEF', 'DEF', 'DEFA', 'DEM', 'DES', 'DEUR', 'DEUS', 'DEW', 'DEX', 'DFE', 'DFEB', 'DFEN', 'DFJ', 'DFND', 'DFNL', 'DFP', 'DFVL', 'DFVS', 'DGAZ', 'DGBP', 'DGL', 'DGLD', 'DGP', 'DGRE', 'DGRO', 'DGRS', 'DGRW', 'DGS', 'DGT', 'DGZ', 'DHF', 'DHS', 'DHY', 'DIA', 'DIAL', 'DIAX', 'DIG', 'DIM', 'DINT', 'DIV', 'DIVA', 'DIVB', 'DIVC', 'DIVO', 'DIVY', 'DJCB', 'DJCI', 'DJD', 'DJP', 'DJPY', 'DLN', 'DLS', 'DLY', 'DMAY', 'DMB', 'DMDV', 'DMF', 'DMO', 'DMRE', 'DMRI', 'DMRL', 'DMRM', 'DMRS', 'DNI', 'DNL', 'DNOV', 'DNP', 'DOG', 'DOL', 'DON', 'DOO', 'DPG', 'DPST', 'DRIP', 'DRIV', 'DRN', 'DRSK', 'DRV', 'DRW', 'DSE', 'DSI', 'DSL', 'DSLV', 'DSM', 'DSTL', 'DSU', 'DTD', 'DTEC', 'DTF', 'DTH', 'DTN', 'DTUL', 'DTUS', 'DTYL', 'DUC', 'DUG', 'DURA', 'DUSA', 'DUSL', 'DUST', 'DVLU', 'DVOL', 'DVP', 'DVY', 'DVYA', 'DVYE', 'DWAS', 'DWAT', 'DWAW', 'DWCR', 'DWEQ', 'DWFI', 'DWLD', 'DWM', 'DWMC', 'DWMF', 'DWPP', 'DWSH', 'DWUS', 'DWX', 'DXD', 'DXGE', 'DXJ', 'DXJS', 'DYNF', 'DZK', 'DZZ', 'EAD', 'EAGG', 'EASG', 'EASI', 'EBIZ', 'EBND', 'ECC', 'ECCX', 'ECCY', 'ECF', 'ECH', 'ECLN', 'ECNS', 'ECON', 'ECOW', 'ECOZ', 'EDC', 'EDD', 'EDEN', 'EDF', 'EDI', 'EDIV', 'EDOG', 'EDOW', 'EDV', 'EDZ', 'EEA', 'EEH', 'EELV', 'EEM', 'EEMA', 'EEMD', 'EEMO', 'EEMS', 'EEMV', 'EEMX', 'EES', 'EET', 'EEV', 'EFA', 'EFAD', 'EFAS', 'EFAV', 'EFAX', 'EFF', 'EFG', 'EFL', 'EFNL', 'EFO', 'EFR', 'EFT', 'EFU', 'EFV', 'EFZ', 'EGF', 'EGIF', 'EGPT', 'EHI', 'EHT', 'EIC', 'EIDO', 'EIM', 'EINC', 'EIRL', 'EIS', 'EJAN', 'EJUL', 'EKAR', 'ELD', 'EMAG', 'EMB', 'EMBD', 'EMBH', 'EMCB', 'EMD', 'EMDV', 'EMF', 'EMFM', 'EMGF', 'EMHY', 'EMIF', 'EMLC', 'EMLP', 'EMMF', 'EMNT', 'EMO', 'EMQQ', 'EMSG', 'EMSH', 'EMTL', 'EMTY', 'EMXC', 'ENFR', 'ENOR', 'ENTR', 'ENX', 'ENZL', 'EOD', 'EOI', 'EOS', 'EOT', 'EPHE', 'EPI', 'EPOL', 'EPP', 'EPRF', 'EPS', 'EPU', 'EPV', 'EQAL', 'EQL', 'EQRR', 'EQWL', 'ERC', 'ERH', 'ERM', 'ERSX', 'ERUS', 'ERX', 'ERY', 'ESCR', 'ESEB', 'ESG', 'ESGD', 'ESGE', 'ESGG', 'ESGN', 'ESGS', 'ESGU', 'ESGV', 'ESHY', 'ESML', 'ESNG', 'ESPO', 'ETB', 'ETG', 'ETHO', 'ETJ', 'ETO', 'ETV', 'ETW', 'ETX', 'ETY', 'EUDG', 'EUDV', 'EUFN', 'EUFX', 'EUM', 'EUMV', 'EUO', 'EURL', 'EURZ', 'EUSA', 'EUSC', 'EVF', 'EVG', 'EVM', 'EVN', 'EVSTC', 'EVT', 'EVV', 'EVX', 'EVY', 'EWA', 'EWC', 'EWCO', 'EWD', 'EWG', 'EWGS', 'EWH', 'EWI', 'EWJ', 'EWJE', 'EWJV', 'EWK', 'EWL', 'EWM', 'EWMC', 'EWN', 'EWO', 'EWP', 'EWQ', 'EWRE', 'EWS', 'EWSC', 'EWT', 'EWU', 'EWUS', 'EWV', 'EWW', 'EWX', 'EWY', 'EWZ', 'EWZS', 'EXD', 'EXG', 'EXI', 'EYLD', 'EZA', 'EZJ', 'EZM', 'EZU', 'FAAR', 'FAB', 'FAD', 'FALN', 'FAM', 'FAN', 'FAS', 'FAUG', 'FAUS', 'FAX', 'FAZ', 'FBCG', 'FBCV', 'FBGX', 'FBND', 'FBT', 'FBZ', 'FCA', 'FCAL', 'FCAN', 'FCEF', 'FCG', 'FCO', 'FCOM', 'FCOR', 'FCPI', 'FCT', 'FCTR', 'FCVT', 'FDD', 'FDEM', 'FDEU', 'FDEV', 'FDG', 'FDHY', 'FDIS', 'FDIV', 'FDL', 'FDLO', 'FDM', 'FDMO', 'FDN', 'FDNI', 'FDRR', 'FDT', 'FDTS', 'FDVV', 'FEI', 'FEM', 'FEMB', 'FEMS', 'FEN', 'FENY', 'FEO', 'FEP', 'FEUL', 'FEUZ', 'FEX', 'FEZ', 'FFA', 'FFC', 'FFEB', 'FFEU', 'FFHG', 'FFIU', 'FFR', 'FFSG', 'FFTG', 'FFTI', 'FFTY', 'FGB', 'FGD', 'FGM', 'FHK', 'FHLC', 'FIBR', 'FID', 'FIDI', 'FIDU', 'FIEE', 'FIF', 'FIHD', 'FILL', 'FINX', 'FISR', 'FITE', 'FIV', 'FIVA', 'FIVG', 'FIW', 'FIXD', 'FIYY', 'FJNK', 'FJP', 'FKO', 'FKU', 'FLAT', 'FLAU', 'FLAX', 'FLBL', 'FLBR', 'FLC', 'FLCA', 'FLCB', 'FLCH', 'FLCO', 'FLDR', 'FLEE', 'FLEH', 'FLEU', 'FLFR', 'FLGB', 'FLGE', 'FLGR', 'FLHK', 'FLHY', 'FLIA', 'FLIN', 'FLIY', 'FLJH', 'FLJP', 'FLKR', 'FLLA', 'FLLV', 'FLM', 'FLMB', 'FLMI', 'FLMX', 'FLN', 'FLOT', 'FLQD', 'FLQE', 'FLQG', 'FLQH', 'FLQL', 'FLQM', 'FLQS', 'FLRN', 'FLRT', 'FLRU', 'FLSA', 'FLSP', 'FLSW', 'FLTB', 'FLTR', 'FLTW', 'FLV', 'FLYT', 'FLZA', 'FM', 'FMAT', 'FMAY', 'FMB', 'FMF', 'FMHI', 'FMIL', 'FMK', 'FMN', 'FMO', 'FMY', 'FNCL', 'FNDA', 'FNDB', 'FNDC', 'FNDE', 'FNDF', 'FNDX', 'FNGD', 'FNGO', 'FNGS', 'FNGU', 'FNGZ', 'FNI', 'FNK', 'FNOV', 'FNX', 'FNY', 'FOF', 'FOVL', 'FPA', 'FPE', 'FPEI', 'FPF', 'FPL', 'FPX', 'FPXE', 'FPXI', 'FQAL', 'FRA', 'FRAK', 'FRDM', 'FREL', 'FRI', 'FRLG', 'FSD', 'FSLF', 'FSMB', 'FSMD', 'FSTA', 'FSZ', 'FT', 'FTA', 'FTAG', 'FTC', 'FTCS', 'FTEC', 'FTF', 'FTGC', 'FTHI', 'FTLB', 'FTLS', 'FTRI', 'FTSD', 'FTSL', 'FTSM', 'FTXD', 'FTXG', 'FTXH', 'FTXL', 'FTXN', 'FTXO', 'FTXR', 'FUD', 'FUE', 'FUMB', 'FUND', 'FUT', 'FUTY', 'FV', 'FVAL', 'FVC', 'FVD', 'FVL', 'FWDB', 'FXA', 'FXB', 'FXC', 'FXD', 'FXE', 'FXF', 'FXG', 'FXH', 'FXI', 'FXL', 'FXN', 'FXO', 'FXP', 'FXR', 'FXU', 'FXY', 'FXZ', 'FYC', 'FYLD', 'FYT', 'FYX', 'GAA', 'GAB', 'GAL', 'GAM', 'GAMR', 'GARS', 'GAZ', 'GBAB', 'GBDV', 'GBF', 'GBIL', 'GBUG', 'GBUY', 'GCC', 'GCE', 'GCOW', 'GCV', 'GDAT', 'GDL', 'GDMA', 'GDNA', 'GDO', 'GDV', 'GDVD', 'GDX', 'GDXJ', 'GEM', 'GENY', 'GER', 'GF', 'GFIN', 'GFY', 'GGM', 'GGN', 'GGO', 'GGT', 'GGZ', 'GHY', 'GHYB', 'GHYG', 'GIGB', 'GIGE', 'GII', 'GIM', 'GLCN', 'GLD', 'GLDI', 'GLDM', 'GLIF', 'GLIN', 'GLL', 'GLO', 'GLQ', 'GLTR', 'GLU', 'GLV', 'GMAN', 'GMF', 'GMOM', 'GMZ', 'GNAF', 'GNMA', 'GNOM', 'GNR', 'GNT', 'GOAT', 'GOAU', 'GOEX', 'GOF', 'GOVT', 'GPM', 'GQRE', 'GREK', 'GRES', 'GRF', 'GRID', 'GRN', 'GRNB', 'GRU', 'GRX', 'GSC', 'GSEE', 'GSEU', 'GSEW', 'GSG', 'GSID', 'GSIE', 'GSJY', 'GSLC', 'GSP', 'GSSC', 'GSST', 'GSUS', 'GSY', 'GTIP', 'GTO', 'GUDB', 'GUNR', 'GURU', 'GUSH', 'GUT', 'GVAL', 'GVI', 'GVIP', 'GWX', 'GXC', 'GXF', 'GXG', 'GXTG', 'GYLD', 'HACK', 'HAIL', 'HAP', 'HAUD', 'HAUZ', 'HAWX', 'HCAP', 'HCRB', 'HDAW', 'HDEF', 'HDG', 'HDGE', 'HDIV', 'HDLB', 'HDMV', 'HDV', 'HEDJ', 'HEEM', 'HEFA', 'HELX', 'HEQ', 'HERD', 'HERO', 'HEWC', 'HEWG', 'HEWI', 'HEWJ', 'HEWL', 'HEWP', 'HEWU', 'HEWW', 'HEWY', 'HEZU', 'HFRO', 'HFXE', 'HFXI', 'HFXJ', 'HGLB', 'HIBL', 'HIBS', 'HIE', 'HIO', 'HIPS', 'HIX', 'HJPX', 'HLAL', 'HMOP', 'HNDL', 'HNW', 'HOLD', 'HOMZ', 'HPF', 'HPI', 'HPS', 'HQH', 'HQL', 'HRZN', 'HSCZ', 'HSMV', 'HSPX', 'HSRT', 'HTAB', 'HTD', 'HTEC', 'HTFA', 'HTRB', 'HTUS', 'HTY', 'HUSV', 'HYB', 'HYD', 'HYDB', 'HYDW', 'HYEM', 'HYG', 'HYGH', 'HYGV', 'HYHG', 'HYI', 'HYLB', 'HYLD', 'HYLS', 'HYLV', 'HYMB', 'HYS', 'HYT', 'HYTR', 'HYUP', 'HYXE', 'HYXU', 'HYZD', 'IAE', 'IAF', 'IAGG', 'IAI', 'IAK', 'IAT', 'IAU', 'IAUF', 'IBB', 'IBCE', 'IBD', 'IBDD', 'IBDL', 'IBDM', 'IBDN', 'IBDO', 'IBDP', 'IBDQ', 'IBDR', 'IBDS', 'IBDT', 'IBDU', 'IBHA', 'IBHB', 'IBHC', 'IBHD', 'IBHE', 'IBMI', 'IBMJ', 'IBMK', 'IBML', 'IBMM', 'IBMN', 'IBMO', 'IBMP', 'IBMQ', 'IBND', 'IBTA', 'IBTB', 'IBTD', 'IBTE', 'IBTF', 'IBTG', 'IBTH', 'IBTI', 'IBTJ', 'IBUY', 'ICF', 'ICLN', 'ICOL', 'ICOW', 'ICSH', 'ICVT', 'IDE', 'IDEV', 'IDHD', 'IDHQ', 'IDIV', 'IDLB', 'IDLV', 'IDMO', 'IDNA', 'IDOG', 'IDRV', 'IDU', 'IDV', 'IDX', 'IDY', 'IECS', 'IEDI', 'IEF', 'IEFA', 'IEFN', 'IEHS', 'IEI', 'IEIH', 'IEME', 'IEMG', 'IEO', 'IETC', 'IEUR', 'IEUS', 'IEV', 'IEZ', 'IFEU', 'IFGL', 'IFN', 'IFRA', 'IFV', 'IG', 'IGA', 'IGBH', 'IGD', 'IGE', 'IGEB', 'IGF', 'IGHG', 'IGI', 'IGIB', 'IGLB', 'IGM', 'IGN', 'IGOV', 'IGR', 'IGRO', 'IGSB', 'IGV', 'IHAK', 'IHD', 'IHDG', 'IHE', 'IHF', 'IHI', 'IHIT', 'IHTA', 'IHY', 'IID', 'IIF', 'IIGD', 'IIGV', 'IIM', 'IJAN', 'IJH', 'IJJ', 'IJK', 'IJR', 'IJS', 'IJT', 'IJUL', 'ILF', 'ILTB', 'IMLP', 'IMOM', 'IMTB', 'IMTM', 'INCO', 'INDA', 'INDL', 'INDS', 'INDY', 'INFR', 'INKM', 'INSI', 'INTF', 'IOO', 'IPAC', 'IPAY', 'IPFF', 'IPKW', 'IPO', 'IPOS', 'IQDE', 'IQDF', 'IQDG', 'IQDY', 'IQI', 'IQIN', 'IQLT', 'IQM', 'IQSI', 'IQSU', 'IRBO', 'IRL', 'IRR', 'ISCF', 'ISD', 'ISDS', 'ISDX', 'ISEM', 'ISHG', 'ISMD', 'ISRA', 'ISTB', 'ISZE', 'ITA', 'ITB', 'ITEQ', 'ITM', 'ITOT', 'IUS', 'IUSB', 'IUSG', 'IUSS', 'IUSV', 'IVAL', 'IVE', 'IVES', 'IVH', 'IVLU', 'IVOG', 'IVOL', 'IVOO', 'IVOV', 'IVV', 'IVW', 'IWB', 'IWC', 'IWD', 'IWF', 'IWL', 'IWM', 'IWN', 'IWO', 'IWP', 'IWR', 'IWS', 'IWV', 'IWX', 'IWY', 'IXC', 'IXG', 'IXJ', 'IXN', 'IXP', 'IXSE', 'IXUS', 'IYC', 'IYE', 'IYF', 'IYG', 'IYH', 'IYJ', 'IYK', 'IYLD', 'IYM', 'IYR', 'IYT', 'IYW', 'IYY', 'IYZ', 'IZRL', 'JAGG', 'JCE', 'JCO', 'JCPB', 'JDD', 'JDIV', 'JDST', 'JEMD', 'JEPI', 'JEQ', 'JETS', 'JFR', 'JGH', 'JHAA', 'JHCS', 'JHEM', 'JHI', 'JHMA', 'JHMC', 'JHMD', 'JHME', 'JHMF', 'JHMH', 'JHMI', 'JHML', 'JHMM', 'JHMS', 'JHMT', 'JHMU', 'JHS', 'JHSC', 'JHY', 'JIG', 'JIGB', 'JJA', 'JJC', 'JJE', 'JJG', 'JJM', 'JJN', 'JJP', 'JJS', 'JJT', 'JJU', 'JKD', 'JKE', 'JKF', 'JKG', 'JKH', 'JKI', 'JKJ', 'JKK', 'JKL', 'JLS', 'JMBS', 'JMIN', 'JMM', 'JMOM', 'JMST', 'JMUB', 'JNK', 'JNUG', 'JO', 'JOF', 'JOYY', 'JPC', 'JPED', 'JPEM', 'JPEU', 'JPGB', 'JPGE', 'JPHF', 'JPHY', 'JPI', 'JPIN', 'JPLS', 'JPMB', 'JPME', 'JPMF', 'JPMV', 'JPN', 'JPNL', 'JPS', 'JPSE', 'JPST', 'JPT', 'JPUS', 'JPXN', 'JQC', 'JQUA', 'JRI', 'JRO', 'JRS', 'JSD', 'JSMD', 'JSML', 'JTA', 'JTD', 'JUST', 'JVAL', 'JXI', 'KALL', 'KAPR', 'KARS', 'KBA', 'KBE', 'KBWB', 'KBWD', 'KBWP', 'KBWR', 'KBWY', 'KCCB', 'KCE', 'KCNY', 'KDFI', 'KEMQ', 'KEMX', 'KF', 'KFYP', 'KGRN', 'KIE', 'KIO', 'KJAN', 'KLCD', 'KLDW', 'KMED', 'KMF', 'KNAB', 'KNG', 'KNOW', 'KOCT', 'KOIN', 'KOKU', 'KOL', 'KOLD', 'KOMP', 'KORP', 'KORU', 'KRE', 'KRMA', 'KRP', 'KSA', 'KSCD', 'KSM', 'KTF', 'KURE', 'KWEB', 'KXI', 'LABD', 'LABU', 'LACK', 'LBJ', 'LCR', 'LD', 'LDEM', 'LDP', 'LDRS', 'LDSF', 'LDUR', 'LEAD', 'LEGR', 'LEMB', 'LEND', 'LEO', 'LFEQ', 'LGH', 'LGI', 'LGLV', 'LGOV', 'LIT', 'LKOR', 'LMBS', 'LMLB', 'LNGR', 'LOUP', 'LOWC', 'LQD', 'LQDH', 'LQDI', 'LRGE', 'LRGF', 'LRNZ', 'LSAF', 'LSLT', 'LSST', 'LTL', 'LTPZ', 'LVHB', 'LVHD', 'LVHI', 'LVUS', 'MAAX', 'MAGA', 'MARB', 'MAV', 'MBB', 'MBSD', 'MCA', 'MCC', 'MCEF', 'MCHI', 'MCN', 'MCR', 'MCRO', 'MCX', 'MDIV', 'MDY', 'MDYG', 'MDYV', 'MEAR', 'MEN', 'MEXX', 'MFD', 'MFDX', 'MFEM', 'MFL', 'MFM', 'MFMS', 'MFT', 'MFUS', 'MFV', 'MGC', 'MGF', 'MGK', 'MGU', 'MGV', 'MHD', 'MHE', 'MHF', 'MHI', 'MHN', 'MIDF', 'MIDU', 'MIE', 'MILN', 'MINC', 'MINT', 'MIY', 'MJ', 'MJJ', 'MJO', 'MLN', 'MLPA', 'MLPB', 'MLPC', 'MLPE', 'MLPG', 'MLPI', 'MLPR', 'MLPX', 'MLPY', 'MLTI', 'MMAC', 'MMD', 'MMIN', 'MMIT', 'MMT', 'MMTM', 'MMU', 'MNA', 'MNE', 'MNP', 'MNRL', 'MOAT', 'MOM', 'MOO', 'MORT', 'MOTI', 'MOTO', 'MPA', 'MPV', 'MQT', 'MQY', 'MRGR', 'MSD', 'MSUS', 'MSVX', 'MTGP', 'MTT', 'MTUM', 'MUA', 'MUB', 'MUC', 'MUE', 'MUH', 'MUI', 'MUJ', 'MUNI', 'MUS', 'MUST', 'MUTE', 'MVF', 'MVIN', 'MVO', 'MVRL', 'MVT', 'MVV', 'MXDE', 'MXDU', 'MXE', 'MXF', 'MXI', 'MYC', 'MYD', 'MYF', 'MYI', 'MYJ', 'MYN', 'MYY', 'MZA', 'MZZ', 'NAC', 'NACP', 'NAD', 'NAIL', 'NAN', 'NANR', 'NAPR', 'NAZ', 'NBB', 'NBH', 'NBO', 'NBW', 'NCA', 'NCB', 'NCV', 'NCZ', 'NDP', 'NEA', 'NEAR', 'NEED', 'NERD', 'NETL', 'NEV', 'NFJ', 'NFLT', 'NFRA', 'NFTY', 'NGE', 'NHA', 'NHF', 'NHS', 'NIB', 'NID', 'NIE', 'NIM', 'NIQ', 'NJAN', 'NJV', 'NKG', 'NKX', 'NLR', 'NMCO', 'NMI', 'NML', 'NMS', 'NMT', 'NMY', 'NMZ', 'NNY', 'NOBL', 'NOCT', 'NOM', 'NORW', 'NPN', 'NPV', 'NQP', 'NRGD', 'NRGO', 'NRGU', 'NRGX', 'NRGZ', 'NRK', 'NRO', 'NSL', 'NTG', 'NTSX', 'NUAG', 'NUBD', 'NUDM', 'NUEM', 'NUGT', 'NUHY', 'NULC', 'NULG', 'NULV', 'NUM', 'NUMG', 'NUMV', 'NUO', 'NURE', 'NUSA', 'NUSC', 'NUSI', 'NUV', 'NUW', 'NVG', 'NXC', 'NXJ', 'NXN', 'NXP', 'NXQ', 'NXR', 'NXTG', 'NYF', 'NYV', 'NZF', 'OBOR', 'OCCI', 'OCIO', 'OCSI', 'OEF', 'OEUR', 'OGIG', 'OIA', 'OIH', 'OILK', 'OLD', 'OLEM', 'OMFL', 'OMFS', 'ONEO', 'ONEQ', 'ONEV', 'ONEY', 'ONLN', 'OPER', 'OPP', 'OSCV', 'OUNZ', 'OUSA', 'OUSM', 'OVB', 'OVF', 'OVL', 'OVM', 'OVS', 'OXLC', 'PACA', 'PAI', 'PAK', 'PALL', 'PAPR', 'PASS', 'PAUG', 'PAVE', 'PAWZ', 'PBD', 'PBDM', 'PBE', 'PBEE', 'PBJ', 'PBND', 'PBP', 'PBS', 'PBSM', 'PBT', 'PBTP', 'PBUS', 'PBW', 'PCEF', 'PCF', 'PCI', 'PCK', 'PCM', 'PCN', 'PCQ', 'PCY', 'PDBC', 'PDEC', 'PDI', 'PDN', 'PDP', 'PDT', 'PEJ', 'PEO', 'PEX', 'PEXL', 'PEY', 'PEZ', 'PFD', 'PFEB', 'PFF', 'PFFA', 'PFFD', 'PFFL', 'PFFR', 'PFI', 'PFIG', 'PFL', 'PFLD', 'PFM', 'PFN', 'PFO', 'PFXF', 'PGAL', 'PGF', 'PGHY', 'PGJ', 'PGM', 'PGP', 'PGX', 'PGZ', 'PHB', 'PHD', 'PHDG', 'PHK', 'PHO', 'PHT', 'PHYL', 'PHYS', 'PICB', 'PICK', 'PID', 'PIE', 'PILL', 'PIM', 'PIN', 'PIO', 'PIZ', 'PJAN', 'PJP', 'PJUL', 'PJUN', 'PKB', 'PKO', 'PKW', 'PLAT', 'PLC', 'PLCY', 'PLTM', 'PLW', 'PMAR', 'PMAY', 'PMF', 'PML', 'PMM', 'PMO', 'PMOM', 'PMX', 'PNF', 'PNI', 'PNOV', 'PNQI', 'POCT', 'POTX', 'PPA', 'PPDM', 'PPEM', 'PPH', 'PPLC', 'PPLT', 'PPMC', 'PPR', 'PPSC', 'PPT', 'PPTY', 'PQIN', 'PQLC', 'PQSG', 'PQSV', 'PREF', 'PRF', 'PRFZ', 'PRN', 'PRNT', 'PSC', 'PSCC', 'PSCD', 'PSCE', 'PSCF', 'PSCH', 'PSCI', 'PSCM', 'PSCT', 'PSCU', 'PSEP', 'PSET', 'PSF', 'PSI', 'PSJ', 'PSK', 'PSL', 'PSLV', 'PSM', 'PSMB', 'PSMC', 'PSMG', 'PSMM', 'PSP', 'PSQ', 'PSR', 'PST', 'PTBD', 'PTEU', 'PTF', 'PTH', 'PTIN', 'PTLC', 'PTMC', 'PTNQ', 'PTY', 'PUI', 'PULS', 'PUTW', 'PVAL', 'PVI', 'PWB', 'PWC', 'PWS', 'PWV', 'PWZ', 'PXE', 'PXF', 'PXH', 'PXI', 'PXJ', 'PXQ', 'PY', 'PYN', 'PYPE', 'PYZ', 'PZA', 'PZC', 'PZD', 'PZT', 'QABA', 'QAI', 'QARP', 'QAT', 'QCLN', 'QDEF', 'QDF', 'QDIV', 'QDYN', 'QED', 'QEFA', 'QEMM', 'QGRO', 'QGTA', 'QID', 'QINT', 'QLC', 'QLD', 'QLS', 'QLTA', 'QLV', 'QLVD', 'QLVE', 'QMJ', 'QMN', 'QMOM', 'QQEW', 'QQH', 'QQQ', 'QQQE', 'QQQX', 'QQXT', 'QRFT', 'QSY', 'QTEC', 'QTUM', 'QUAL', 'QUS', 'QVAL', 'QVM', 'QWLD', 'QYLD', 'RA', 'RAAX', 'RAFE', 'RALS', 'RAVI', 'RBIN', 'RBUS', 'RCD', 'RCG', 'RCS', 'RDIV', 'RDOG', 'RDVY', 'RECS', 'REET', 'REGL', 'REK', 'REM', 'REML', 'REMX', 'RESD', 'RESE', 'RESP', 'RETL', 'REVS', 'REW', 'REZ', 'RFAP', 'RFCI', 'RFDA', 'RFDI', 'RFEM', 'RFEU', 'RFFC', 'RFG', 'RFI', 'RFUN', 'RFV', 'RGI', 'RGT', 'RHS', 'RIF', 'RIGS', 'RINF', 'RING', 'RISE', 'RIV', 'RJA', 'RJI', 'RJN', 'RJZ', 'RLY', 'RMM', 'RMT', 'RNDM', 'RNDV', 'RNEM', 'RNLC', 'RNMC', 'RNP', 'RNSC', 'ROAM', 'ROBO', 'ROBT', 'RODE', 'RODI', 'RODM', 'ROKT', 'ROM', 'ROMO', 'ROOF', 'RORE', 'ROSC', 'ROUS', 'RPAR', 'RPG', 'RPV', 'RQI', 'RSF', 'RSP', 'RSX', 'RSXJ', 'RTH', 'RTM', 'RUSL', 'RVNU', 'RVRS', 'RVT', 'RWCD', 'RWDC', 'RWDE', 'RWED', 'RWGV', 'RWIU', 'RWJ', 'RWK', 'RWL', 'RWLS', 'RWM', 'RWO', 'RWR', 'RWSL', 'RWUI', 'RWVG', 'RWX', 'RXD', 'RXI', 'RXL', 'RYE', 'RYF', 'RYH', 'RYJ', 'RYLD', 'RYT', 'RYU', 'RZG', 'RZV', 'SAA', 'SBB', 'SBI', 'SBIO', 'SBM', 'SBUG', 'SCC', 'SCD', 'SCHA', 'SCHB', 'SCHC', 'SCHD', 'SCHE', 'SCHF', 'SCHG', 'SCHH', 'SCHI', 'SCHJ', 'SCHK', 'SCHM', 'SCHO', 'SCHP', 'SCHQ', 'SCHR', 'SCHV', 'SCHX', 'SCHZ', 'SCID', 'SCIJ', 'SCIU', 'SCIX', 'SCJ', 'SCO', 'SCZ', 'SDAG', 'SDCI', 'SDD', 'SDEM', 'SDG', 'SDGA', 'SDIV', 'SDOG', 'SDOW', 'SDP', 'SDS', 'SDVY', 'SDY', 'SDYL', 'SECT', 'SEF', 'SEIX', 'SFHY', 'SFIG', 'SFY', 'SFYF', 'SFYX', 'SGDJ', 'SGDM', 'SGG', 'SGOL', 'SGOV', 'SH', 'SHAG', 'SHE', 'SHM', 'SHV', 'SHY', 'SHYD', 'SHYG', 'SHYL', 'SIJ', 'SIL', 'SILJ', 'SIMS', 'SIVR', 'SIXA', 'SIXH', 'SIXL', 'SIXS', 'SIZE', 'SJB', 'SJNK', 'SJT', 'SKF', 'SKOR', 'SKYY', 'SLQD', 'SLT', 'SLV', 'SLVO', 'SLVP', 'SLX', 'SLY', 'SLYG', 'SLYV', 'SMB', 'SMCP', 'SMDD', 'SMDV', 'SMEZ', 'SMH', 'SMHB', 'SMIN', 'SMLF', 'SMLL', 'SMLV', 'SMM', 'SMMD', 'SMMU', 'SMMV', 'SMN', 'SMOG', 'SNLN', 'SNPE', 'SNSR', 'SNUG', 'SOCL', 'SOIL', 'SOR', 'SOVB', 'SOXL', 'SOXS', 'SOXX', 'SOYB', 'SPAB', 'SPBO', 'SPDN', 'SPDV', 'SPDW', 'SPE', 'SPEM', 'SPEU', 'SPFF', 'SPGM', 'SPGP', 'SPHB', 'SPHD', 'SPHQ', 'SPHY', 'SPIB', 'SPIP', 'SPLB', 'SPLG', 'SPLV', 'SPMB', 'SPMD', 'SPMO', 'SPMV', 'SPPP', 'SPSB', 'SPSK', 'SPSM', 'SPTI', 'SPTL', 'SPTM', 'SPTS', 'SPUS', 'SPUU', 'SPVM', 'SPVU', 'SPXB', 'SPXE', 'SPXL', 'SPXN', 'SPXS', 'SPXT', 'SPXU', 'SPXV', 'SPXX', 'SPY', 'SPYD', 'SPYG', 'SPYV', 'SPYX', 'SQEW', 'SQLV', 'SQQQ', 'SRET', 'SRLN', 'SRS', 'SRTY', 'SRV', 'SRVR', 'SSG', 'SSLY', 'SSO', 'SSPY', 'SSUS', 'STIP', 'STK', 'STLC', 'STLG', 'STLV', 'STMB', 'STOT', 'STPP', 'STPZ', 'STSB', 'SUB', 'SUSA', 'SUSB', 'SUSC', 'SUSL', 'SVXY', 'SWAN', 'SWZ', 'SYE', 'SYG', 'SYLD', 'SYV', 'SZC', 'SZK', 'SZNE', 'TAAG', 'TADS', 'TAEQ', 'TAGS', 'TAIL', 'TAN', 'TAPR', 'TAWK', 'TAXF', 'TBF', 'TBLU', 'TBND', 'TBT', 'TBX', 'TCPC', 'TCTL', 'TDF', 'TDIV', 'TDTF', 'TDTT', 'TDV', 'TEAF', 'TECB', 'TECL', 'TECS', 'TEGS', 'TEI', 'TERM', 'TFI', 'TFIV', 'TFLO', 'TFLT', 'THCX', 'THD', 'THNQ', 'THQ', 'THW', 'TILT', 'TIP', 'TIPX', 'TIPZ', 'TLDH', 'TLEH', 'TLH', 'TLI', 'TLT', 'TLTD', 'TLTE', 'TMDV', 'TMF', 'TMFC', 'TMV', 'TNA', 'TOK', 'TOKE', 'TOLZ', 'TOTL', 'TPAY', 'TPHD', 'TPIF', 'TPL', 'TPLC', 'TPOR', 'TPSC', 'TPYP', 'TPZ', 'TQQQ', 'TRND', 'TRTY', 'TSI', 'TTAC', 'TTAI', 'TTP', 'TTT', 'TTTN', 'TUR', 'TUSA', 'TVIX', 'TWM', 'TWN', 'TY', 'TYBS', 'TYD', 'TYG', 'TYO', 'TZA', 'UAE', 'UAG', 'UAPR', 'UAUD', 'UAUG', 'UBG', 'UBOT', 'UBR', 'UBT', 'UCC', 'UCHF', 'UCI', 'UCIB', 'UCO', 'UCON', 'UDEC', 'UDN', 'UDOW', 'UEUR', 'UEVM', 'UFEB', 'UFO', 'UGA', 'UGAZ', 'UGBP', 'UGE', 'UGL', 'UGLD', 'UITB', 'UIVM', 'UJAN', 'UJB', 'UJPY', 'UJUL', 'UJUN', 'ULE', 'ULST', 'ULTR', 'ULVM', 'UMAR', 'UMAY', 'UMDD', 'UNG', 'UNL', 'UNOV', 'UOCT', 'UPRO', 'UPV', 'UPW', 'URA', 'URE', 'URNM', 'URTH', 'URTY', 'USA', 'USAI', 'USCI', 'USD', 'USDU', 'USDY', 'USEP', 'USEQ', 'USFR', 'USHG', 'USHY', 'USI', 'USIG', 'USL', 'USLB', 'USLV', 'USMC', 'USMF', 'USMV', 'USO', 'USOI', 'USRT', 'USSG', 'UST', 'USTB', 'USV', 'USVM', 'UTES', 'UTF', 'UTG', 'UTRN', 'UTSL', 'UUP', 'UVXY', 'UWM', 'UXI', 'UYG', 'UYM', 'VALQ', 'VALT', 'VAM', 'VAMO', 'VAW', 'VB', 'VBF', 'VBK', 'VBND', 'VBR', 'VCF', 'VCIF', 'VCIT', 'VCLT', 'VCR', 'VCSH', 'VCV', 'VDC', 'VDE', 'VEA', 'VEGA', 'VEGI', 'VEGN', 'VETS', 'VEU', 'VFH', 'VFL', 'VFLQ', 'VFMF', 'VFMO', 'VFMV', 'VFQY', 'VFVA', 'VGFO', 'VGI', 'VGIT', 'VGK', 'VGLT', 'VGM', 'VGSH', 'VGT', 'VHT', 'VIDI', 'VIG', 'VIGI', 'VIIX', 'VIOG', 'VIOO', 'VIOV', 'VIS', 'VIXM', 'VIXY', 'VKI', 'VKQ', 'VLT', 'VLU', 'VLUE', 'VMBS', 'VMM', 'VMO', 'VMOT', 'VNLA', 'VNM', 'VNQ', 'VNQI', 'VO', 'VOE', 'VONE', 'VONG', 'VONV', 'VOO', 'VOOG', 'VOOV', 'VOT', 'VOX', 'VPC', 'VPL', 'VPU', 'VPV', 'VQT', 'VRAI', 'VRIG', 'VRP', 'VSDA', 'VSGX', 'VSL', 'VSMV', 'VSS', 'VT', 'VTA', 'VTC', 'VTEB', 'VTHR', 'VTI', 'VTIP', 'VTN', 'VTV', 'VTWG', 'VTWO', 'VTWV', 'VUG', 'VUSE', 'VV', 'VVR', 'VWO', 'VWOB', 'VXF', 'VXUS', 'VXX', 'VYM', 'VYMI', 'WANT', 'WBIE', 'WBIF', 'WBIG', 'WBII', 'WBIL', 'WBIN', 'WBIT', 'WBIY', 'WBND', 'WCLD', 'WDIV', 'WEA', 'WEAT', 'WEBL', 'WEBS', 'WFHY', 'WFIG', 'WIA', 'WIL', 'WINC', 'WIP', 'WIW', 'WIZ', 'WLDR', 'WOMN', 'WOOD', 'WPS', 'WTMF', 'WUGI', 'WWJD', 'XAR', 'XBI', 'XBUY', 'XCEM', 'XCOM', 'XDIV', 'XES', 'XFLT', 'XHB', 'XHE', 'XHS', 'XITK', 'XLB', 'XLC', 'XLE', 'XLF', 'XLG', 'XLI', 'XLK', 'XLP', 'XLRE', 'XLSR', 'XLU', 'XLV', 'XLY', 'XME', 'XMHQ', 'XMLV', 'XMMO', 'XMPT', 'XMVM', 'XNTK', 'XOP', 'XOUT', 'XPH', 'XPP', 'XRLV', 'XRT', 'XSD', 'XSHD', 'XSHQ', 'XSLV', 'XSMO', 'XSOE', 'XSVM', 'XSW', 'XT', 'XTL', 'XTN', 'XVZ', 'XWEB', 'YANG', 'YCL', 'YCOM', 'YCS', 'YGRN', 'YINN', 'YLCO', 'YLD', 'YLDE', 'YOLO', 'YXI', 'YYY', 'ZCAN', 'ZDEU', 'ZGBR', 'ZHOK', 'ZIG', 'ZIV', 'ZJPN', 'ZMLP', 'ZROZ', 'ZSL', 'ZTR']
investment_trusts_mutual_funds_master_file='itmf.csv'
investment_trusts_mutual_funds_directory = 'investment_trusts_mutual_funds'
get_industry_ticker_data(investment_trusts_mutual_funds ,investment_trusts_mutual_funds_directory)
create_industry_master_list(investment_trusts_mutual_funds ,investment_trusts_mutual_funds_master_file,investment_trusts_mutual_funds_directory)
print("-> Master file created, filename: %s\n" % investment_trusts_mutual_funds_master_file)
purge_industry_low_gain_volume(investment_trusts_mutual_funds_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
life_health_insurance=['AEL', 'AFL', 'AMSF', 'ANAT', 'BHF', 'BHFAL', 'CIA', 'EHTH', 'EIG', 'FFG', 'GL', 'GNW', 'GWGH', 'IHC', 'LFC', 'LNC', 'MET', 'MFC', 'NWLI', 'PRI', 'RGA', 'RZB', 'UNM', 'UNMA', 'VERY', 'VOYA']
life_health_insurance_master_file='life_health_insurance_mf.csv'
life_health_insurance_directory = 'life_health_insurance'
get_industry_ticker_data(life_health_insurance ,life_health_insurance_directory)
create_industry_master_list(life_health_insurance ,life_health_insurance_master_file,life_health_insurance_directory)
print("-> Master file created, filename: %s\n" % life_health_insurance_master_file)
purge_industry_low_gain_volume(life_health_insurance_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
major_banks=['ALRS', 'AMNB', 'AMTB', 'AMTBB', 'AROW', 'AVAL', 'BAC', 'BBD', 'BBDO', 'BBVA', 'BCS', 'BK', 'BMO', 'BNS', 'CADE', 'CASH', 'CBNK', 'CBTX', 'CHCO', 'CIT', 'CM', 'CMA', 'COF', 'CS', 'DB', 'ESQ', 'FBK', 'FBMS', 'FNLC', 'HOPE', 'HSBC', 'ITUB', 'JPM', 'KEY', 'LARK', 'LYG', 'MCB', 'MFG', 'MUFG', 'NKSH', 'OFG', 'OPOF', 'PBCT', 'PEBO', 'PNC', 'PRK', 'RBS', 'RF', 'RMBI', 'RY', 'SAN', 'SFNC', 'SMFG', 'SUPV', 'TCF', 'T', 'UBS', 'USB', 'WBK', 'WF', 'WFC', 'ZIONL']
major_banks_master_file='major_banks_mf.csv'
major_banks_directory = 'major_banks'
get_industry_ticker_data(major_banks ,major_banks_directory)
create_industry_master_list(major_banks ,major_banks_master_file,major_banks_directory)
print("-> Master file created, filename: %s\n" % major_banks_master_file)
purge_industry_low_gain_volume(major_banks_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
major_telecommunications=['ALSK', 'ATNI', 'AWRE', 'BCE', 'CHA', 'CHT', 'CHU', 'CNSL', 'KT', 'OIBR.C', 'ORAN', 'SHEN', 'T', 'TEO', 'TEUM', 'TLK', 'TU', 'VIV', 'VZ']
major_telecommunications_master_file='major_telecomm_mf.csv'
major_telecommunications_directory = 'major_telecommunications'
get_industry_ticker_data(major_telecommunications ,major_telecommunications_directory)
create_industry_master_list(major_telecommunications ,major_telecommunications_master_file,major_telecommunications_directory)
print("-> Master file created, filename: %s\n" % major_telecommunications_master_file)
purge_industry_low_gain_volume(major_telecommunications_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
managed_health_care=['ANTM','CI','CNC','GTS','HUM','MGLN','MOH','UNH']
managed_health_care_master_file='managed_health_care_mf.csv'
managed_health_care_directory = 'managed_health_care'
get_industry_ticker_data(managed_health_care ,managed_health_care_directory)
create_industry_master_list(managed_health_care ,managed_health_care_master_file,managed_health_care_directory)
print("-> Master file created, filename: %s\n" % managed_health_care_master_file)
purge_industry_low_gain_volume(managed_health_care_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
marine_shipping=['ASC', 'CMRE', 'CPLP', 'CTRM', 'DAC', 'DHT', 'DLNG', 'DSSI', 'DSX', 'EDRY', 'EGLE', 'ESEA', 'EURN', 'FRO', 'GASS', 'GLBS', 'GLNG', 'GLOG', 'GLOP', 'GMLP', 'GNK', 'GOGL', 'GRIN', 'INSW', 'KEX', 'KNOP', 'LPG', 'NAT', 'NM', 'NMCI', 'NMM', 'NNA', 'NVGS', 'OSG', 'PSHG', 'PSV', 'PXS', 'SALT', 'SB', 'SBLK', 'SFL', 'SINO', 'STNG', 'TGP', 'TK', 'TNK', 'TNP', 'TOPS', 'TRMD']
marine_shipping_master_file='marine_shipping_mf.csv'
marine_shipping_directory = 'marine_shipping'
get_industry_ticker_data(marine_shipping ,marine_shipping_directory)
create_industry_master_list(marine_shipping ,marine_shipping_master_file,marine_shipping_directory)
print("-> Master file created, filename: %s\n" % marine_shipping_master_file)
purge_industry_low_gain_volume(marine_shipping_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
medical_distributors=['ABC','CAH','HLF','HSIC','MCK','MOHO','MYO','OMI','PDCO']
medical_distributors_master_file='medical_distributors_mf.csv'
medical_distributors_directory = 'medical_distributors'
get_industry_ticker_data(medical_distributors ,medical_distributors_directory)
create_industry_master_list(medical_distributors ,medical_distributors_master_file,medical_distributors_directory)
print("-> Master file created, filename: %s\n" % medical_distributors_master_file)
purge_industry_low_gain_volume(medical_distributors_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
medical_nursing_services=['ACHC', 'ADUS', 'AIH', 'AMED', 'AMS', 'APTX', 'ARA', 'BIOC', 'CATS', 'CCM', 'CHE', 'CO', 'CRDF', 'CRHM', 'CSTL', 'CSU', 'DVA', 'ENSG', 'FLGT', 'FMS', 'GH', 'GHSI', 'HNGR', 'IMAC', 'LHCG', 'MD', 'NEO', 'NTRA', 'NVTA', 'ONCS', 'ONEM', 'OPCH', 'OPTN', 'PDEX', 'PNTG', 'RDNT', 'SDC', 'TVTY', 'USPH', 'XGN']
medical_nursing_services_master_file='med_nursing_services_mf.csv'
medical_nursing_services_directory = 'medical_nursing_services'
get_industry_ticker_data(medical_nursing_services ,medical_nursing_services_directory)
create_industry_master_list(medical_nursing_services ,medical_nursing_services_master_file,medical_nursing_services_directory)
print("-> Master file created, filename: %s\n" % medical_nursing_services_master_file)
purge_industry_low_gain_volume(medical_nursing_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
medical_specialties=['ABMD', 'ABT', 'AEMD', 'AHPI', 'AKER', 'ALC', 'ALGN', 'ANGO', 'APEN', 'APT', 'APYX', 'ARAY', 'ATEC', 'ATOS', 'ATRC', 'ATRI', 'ATRS', 'AVDL', 'AVGR', 'AVNS', 'AXDX', 'AXGN', 'AXNX', 'BAX', 'BDX', 'BEAT', 'BIO', 'BIOL', 'BLFS', 'BLPH', 'BMRA', 'BRKR', 'BSGM', 'BSX', 'BWAY', 'CDNA', 'CEMI', 'CERS', 'CFMS', 'CGIX', 'CHEK', 'CHFS', 'CLDX', 'CLGN', 'CLPT', 'CLSN', 'CMD', 'CNMD', 'COO', 'CRMD', 'CRY', 'CSII', 'CTSO', 'CUTR', 'CYAN', 'DCTH', 'DHR', 'DMTK', 'DRAD', 'DRIO', 'DXCM', 'DXR', 'DYNT', 'ECOR', 'EKSO', 'ELGX', 'ELMD', 'ENLV', 'ESTA', 'EW', 'EYES', 'FLDM', 'FONR', 'GKOS', 'GMAB', 'GMED', 'GNMK', 'HAE', 'HJLI', 'HOLX', 'HRC', 'HSDT', 'HSKA', 'IART', 'ICCC', 'ICUI', 'IDXX', 'IIN', 'INFU', 'INGN', 'INMD', 'INO', 'INSP', 'IRIX', 'IRMD', 'IRTC', 'ISRG', 'ITMR', 'IVC', 'KIDS', 'KRMD', 'LIVN', 'LLIT', 'LMAT', 'LNTH', 'MASI', 'MDGS', 'MDT', 'MLAB', 'MLSS', 'MMSI', 'MOTS', 'MSON', 'MTD', 'MYOS', 'NAOV', 'NARI', 'NEOG', 'NEPH', 'NMRD', 'NSPR', 'NTRP', 'NTUS', 'NURO', 'NUVA', 'NVCN', 'NVCR', 'NVRO', 'NVST', 'NYMX', 'OBLN', 'OCX', 'OFIX', 'OSUR', 'OXFD', 'PAVM', 'PEN', 'PHG', 'PKI', 'PLSE', 'POAI', 'PODD', 'PROF', 'QDEL', 'QGEN', 'QTNT', 'RMD', 'RMED', 'RMTI', 'RTIX', 'RVP', 'RWLK', 'SIBN', 'SIEN', 'SILK', 'SINT', 'SLNO', 'SNN', 'SOLY', 'SPNE', 'SRDX', 'SRTS', 'SSKN', 'STAA', 'STE', 'STIM', 'STXS', 'SWAV', 'SYK', 'TCMD', 'TELA', 'TENX', 'TFX', 'THMO', 'TMDI', 'TMDX', 'TMO', 'TNDM', 'TRIB', 'UTMD', 'VAPO', 'VAR', 'VCEL', 'VERO', 'VIVE', 'VNRX', 'VRAY', 'VREX', 'WAT', 'WMGI', 'WST', 'XRAY', 'XTNT', 'ZBH', 'ZSAN', 'ZYXI']
medical_specialties_master_file='medical_specialties_mf.csv'
medical_specialties_directory = 'medical_specialties'
get_industry_ticker_data(medical_specialties ,medical_specialties_directory)
create_industry_master_list(medical_specialties ,medical_specialties_master_file,medical_specialties_directory)
print("-> Master file created, filename: %s\n" % medical_specialties_master_file)
purge_industry_low_gain_volume(medical_specialties_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
metal_fabrication=['BOOM', 'CIX', 'CVR', 'FRD', 'GIFI', 'HCHC', 'HSC', 'HWCC', 'IIIN', 'MATW', 'MLI', 'NNBR', 'NWPX', 'PKOH', 'ROCK', 'ROLL', 'SGBX', 'SHLO', 'SIM', 'SVT', 'TKR', 'UAMY', 'VMI', 'WIRE']
metal_fabrication_master_file='metal_fabrication_mf.csv'
metal_fabrication_directory = 'metal_fabrication'
get_industry_ticker_data(metal_fabrication ,metal_fabrication_directory)
create_industry_master_list(metal_fabrication ,metal_fabrication_master_file,metal_fabrication_directory)
print("-> Master file created, filename: %s\n" % metal_fabrication_master_file)
purge_industry_low_gain_volume(metal_fabrication_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
miscellaneous=['CCNC','CLSK','LMB','PANL','PME','PRPL','ROSE','ROSEU','SHIP','SWKH','ZOM']
miscellaneous_master_file='miscellaneous_mf.csv'
miscellaneous_directory = 'miscellaneous'
get_industry_ticker_data(miscellaneous ,miscellaneous_directory)
create_industry_master_list(miscellaneous ,miscellaneous_master_file,miscellaneous_directory)
print("-> Master file created, filename: %s\n" % miscellaneous_master_file)
purge_industry_low_gain_volume(miscellaneous_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
miscellaneous_commercial_services=['AACG', 'ABM', 'ACTG', 'ADPT', 'ADT', 'AESE', 'ALJJ', 'ALT', 'APDN', 'ATOM', 'ATTO', 'BAH', 'BAND', 'BBU', 'BCEL', 'BCO', 'BEDU', 'BFAM', 'BOXL', 'BR', 'BV', 'CAE', 'CATM', 'CBZ', 'CCC', 'CDAY', 'CDK', 'CELC', 'CELP', 'CHGG', 'CIDM', 'CLEU', 'CLXT', 'CNDT', 'CNTG', 'CPRT', 'CRAI', 'CRL', 'CSTM', 'DFIN', 'DL', 'EDNT', 'EEX', 'EFX', 'ETSY', 'EVH', 'EVOP', 'FC', 'FCN', 'FEDU', 'FLT', 'FORR', 'FOUR', 'FRG', 'G', 'GAIA', 'GPX', 'GSL', 'HCKT', 'HCSG', 'HMSY', 'HURN', 'ICFI', 'III', 'IT', 'KAR', 'KC', 'KE', 'KROS', 'LAUR', 'LINC', 'LMFA', 'LOPE', 'LRMR', 'LTBR', 'MARA', 'MEDP', 'MEDS', 'METX', 'MGI', 'MMS', 'MNDO', 'NEW', 'NEWT', 'NTIP', 'NVEE', 'NXTD', 'OMEX', 'ONE', 'OPGN', 'PAE', 'PAGS', 'PAYS', 'PCOM', 'PFMT', 'PGNY', 'PPD', 'PRAA', 'PRAH', 'PRGX', 'RBA', 'RCM', 'REDU', 'REKR', 'RELX', 'REZI', 'RPAY', 'RSSS', 'RYB', 'SCOR', 'SGRP', 'SIC', 'SJ', 'SLRX', 'SP', 'SRT', 'STG', 'STMP', 'TEDU', 'TISI', 'TNET', 'TRXC', 'TTEC', 'TUFN', 'USIO', 'UTI', 'VEC', 'VERB', 'VVI', 'VVNT', 'WAFU', 'WLDN', 'WNS', 'XCUR', 'ZCMD']
miscellaneous_commercial_services_master_file='misc_commercial_services_mf.csv'
miscellaneous_commercial_services_directory = 'miscellaneous_commercial_services'
get_industry_ticker_data(miscellaneous_commercial_services ,miscellaneous_commercial_services_directory)
create_industry_master_list(miscellaneous_commercial_services ,miscellaneous_commercial_services_master_file,miscellaneous_commercial_services_directory)
print("-> Master file created, filename: %s\n" % miscellaneous_commercial_services_master_file)
purge_industry_low_gain_volume(miscellaneous_commercial_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
miscellaneous_manufacturing=['BRC', 'CPSH', 'CR', 'CSL', 'FORK', 'HIHO', 'LCII', 'PNR', 'RAVN', 'REYN', 'SXI', 'TG', 'TRS', 'TUP', 'WMS', 'YETI']
miscellaneous_manufacturing_master_file='misc_manufacturing_mf.csv'
miscellaneous_manufacturing_directory = 'miscellaneous_manufacturing'
get_industry_ticker_data(miscellaneous_manufacturing ,miscellaneous_manufacturing_directory)
create_industry_master_list(miscellaneous_manufacturing ,miscellaneous_manufacturing_master_file,miscellaneous_manufacturing_directory)
print("-> Master file created, filename: %s\n" % miscellaneous_manufacturing_master_file)
purge_industry_low_gain_volume(miscellaneous_manufacturing_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
motor_vehicles=['F', 'FCAU', 'FUV', 'GM', 'HMC', 'HOG', 'KNDI', 'NIO', 'NIU', 'NKLA', 'RACE', 'REVG', 'SOLO', 'TM', 'TSLA', 'TTM']
motor_vehicles_master_file='motor_Vehicles_mf.csv'
motor_vehicles_directory = 'motor_vehicles'
get_industry_ticker_data(motor_vehicles ,motor_vehicles_directory)
create_industry_master_list(motor_vehicles ,motor_vehicles_master_file,motor_vehicles_directory)
print("-> Master file created, filename: %s\n" % motor_vehicles_master_file)
purge_industry_low_gain_volume(motor_vehicles_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
movies_entertainment=['ACEL', 'AMC', 'BATRA', 'BATRK', 'BTN', 'BWL.A', 'CNK', 'CSSE', 'DIS', 'DLB', 'DLPN', 'DS', 'DVD', 'EROS', 'FUN', 'FWONA', 'FWONK', 'GMBL', 'GNUS', 'IMAX', 'LGF.A', 'LGF.B', 'LSXMA', 'LSXMB', 'LSXMK', 'LYV', 'MANU', 'MCS', 'MSGE', 'MSGN', 'MSGS', 'PLAY', 'RICK', 'SEAS', 'SIX', 'WSG', 'WWE']
movies_entertainment_master_file='movies_entertainment_mf.csv'
movies_entertainment_directory = 'movies_entertainment'
get_industry_ticker_data(movies_entertainment ,movies_entertainment_directory)
create_industry_master_list(movies_entertainment ,movies_entertainment_master_file,movies_entertainment_directory)
print("-> Master file created, filename: %s\n" % movies_entertainment_master_file)
purge_industry_low_gain_volume(movies_entertainment_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
multi_line_insurance=['AAME', 'ACGL', 'AEB', 'AEG', 'AIG', 'AIZ', 'ARGD', 'ARGO', 'AXS', 'BRK.A', 'BRK.B', 'BRP', 'CNA', 'CNO', 'DGICA', 'DGICB', 'FANH', 'HIG', 'HMN', 'KMPR', 'KNSL', 'NODK', 'PUK', 'SG', 'SLF', 'SLQT', 'SNFCA', 'TRV', 'WTM', 'WTRE']
multi_line_insurance_master_file='multi_line_insurance_mf.csv'
multi_line_insurance_directory = 'multi_line_insurance'
get_industry_ticker_data(multi_line_insurance ,multi_line_insurance_directory)
create_industry_master_list(multi_line_insurance ,multi_line_insurance_master_file,multi_line_insurance_directory)
print("-> Master file created, filename: %s\n" % multi_line_insurance_master_file)
purge_industry_low_gain_volume(multi_line_insurance_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
office_equipment_supplies=['ACCO','HNI','JCI','KEQU','KNL','MLHR','PBI','SCS','VIRC']
office_equipment_supplies_master_file='office_equip_supplies_mf.csv'
office_equipment_supplies_directory = 'office_equipment_supplies'
get_industry_ticker_data(office_equipment_supplies ,office_equipment_supplies_directory)
create_industry_master_list(office_equipment_supplies ,office_equipment_supplies_master_file,office_equipment_supplies_directory)
print("-> Master file created, filename: %s\n" % office_equipment_supplies_master_file)
purge_industry_low_gain_volume(office_equipment_supplies_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
oil_gas_pipelines=['ALTM', 'AM', 'BKEP', 'BPMP', 'CEQP', 'CNXM', 'CQP', 'DCP', 'DKL', 'ENB', 'ENBA', 'ENBL', 'ENLC', 'EPD', 'EQM', 'ET', 'ETRN', 'GEL', 'HEP', 'KMI', 'LNG', 'MMLP', 'MMP', 'MPLX', 'NBLX', 'NGL', 'OKE', 'OMP', 'PAA', 'PAGP', 'PBA', 'PSXP', 'RTLR', 'SHLX', 'SRLP', 'TCP', 'TRP', 'USDP', 'WMB']
oil_gas_pipelines_master_file='oil_gas_pipelines_mf.csv'
oil_gas_pipelines_directory = 'oil_gas_pipelines'
get_industry_ticker_data(oil_gas_pipelines ,oil_gas_pipelines_directory)
create_industry_master_list(oil_gas_pipelines ,oil_gas_pipelines_master_file,oil_gas_pipelines_directory)
print("-> Master file created, filename: %s\n" % oil_gas_pipelines_master_file)
purge_industry_low_gain_volume(oil_gas_pipelines_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
oil_gas_production=['APA', 'AR', 'BATL', 'BCEI', 'BRN', 'BRY', 'BSM', 'BTE', 'CDEV', 'CEI', 'CEO', 'CHAP', 'CHK', 'CLR', 'CNQ', 'COP', 'CPE', 'CPG', 'CRC', 'CRK', 'CRT', 'CVE', 'CXO', 'DMLP', 'DNR', 'DVN', 'ECT', 'EGY', 'ENSV', 'EOG', 'EPM', 'EPSN', 'EQT', 'ERF', 'FANG', 'FLNG', 'GBR', 'GPOR', 'GPRK', 'GTE', 'HES', 'HPR', 'HUSA', 'KOS', 'LLEX', 'LONE', 'LPI', 'MARPS', 'MCEP', 'MCF', 'MGY', 'MR', 'MRO', 'MTDR', 'MTR', 'MUR', 'MXC', 'NBL', 'NEXT', 'NOG', 'NRT', 'OAS', 'OVV', 'OXY', 'PDCE', 'PE', 'PER', 'PHX', 'PNRG', 'PVAC', 'PXD', 'QEP', 'REI', 'ROYT', 'RRC', 'SBOW', 'SBR', 'SD', 'SM', 'SWN', 'TAT', 'TGA', 'TGC', 'USEG', 'VET', 'VIST', 'VOC', 'WLL', 'WPX', 'WTI', 'XEC', 'XOG', 'ZN']
oil_gas_production_master_file='oil_gas_production_mf.csv'
oil_gas_production_directory = 'oil_gas_production'
get_industry_ticker_data(oil_gas_production ,oil_gas_production_directory)
create_industry_master_list(oil_gas_production ,oil_gas_production_master_file,oil_gas_production_directory)
print("-> Master file created, filename: %s\n" % oil_gas_production_master_file)
purge_industry_low_gain_volume(oil_gas_production_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
oil_refining_marketing=['CVI', 'CZZ', 'DK', 'HFC', 'MPC', 'PARR', 'PBF', 'PSX', 'TGS', 'TREC', 'TRGP', 'UGP', 'VLO', 'WES']
oil_refining_marketing_master_file='oil_refining_marketing_mf.csv'
oil_refining_marketing_directory = 'oil_refining_marketing'
get_industry_ticker_data(oil_refining_marketing ,oil_refining_marketing_directory)
create_industry_master_list(oil_refining_marketing ,oil_refining_marketing_master_file,oil_refining_marketing_directory)
print("-> Master file created, filename: %s\n" % oil_refining_marketing_master_file)
purge_industry_low_gain_volume(oil_refining_marketing_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
oilfield_services_equipment=['AROC', 'BKR', 'CCLP', 'CKH', 'CLB', 'DOV', 'DRQ', 'ENG', 'EXTN', 'FET', 'FI', 'FTI', 'FTK', 'FTSI', 'GEOS', 'HAL', 'HLX', 'KLXE', 'LBRT', 'MIND', 'NCSM', 'NESR', 'NEX', 'NGS', 'NINE', 'NOA', 'NOV', 'NR', 'OII', 'OIS', 'PBFX', 'PED', 'PUMP', 'RES', 'RNGR', 'SAEX', 'SDPI', 'SLB', 'SOI', 'SPN', 'TDW', 'TTI', 'TUSK', 'USAC', 'WHD', 'WTTR']
oilfield_services_equipment_master_file='oilfield_services_equip_mf.csv'
oilfield_services_equipment_directory = 'oilfield_services_equipment'
get_industry_ticker_data(oilfield_services_equipment ,oilfield_services_equipment_directory)
create_industry_master_list(oilfield_services_equipment ,oilfield_services_equipment_master_file,oilfield_services_equipment_directory)
print("-> Master file created, filename: %s\n" % oilfield_services_equipment_master_file)
purge_industry_low_gain_volume(oilfield_services_equipment_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
other_consumer_services=['AFYA', 'AMBO', 'APEI', 'ATGE', 'AYTU', 'BKNG', 'BLNK', 'BXG', 'CLCT', 'CLUB', 'CSV', 'CTAS', 'DESP', 'EBAY', 'EDU', 'EXPE', 'GHC', 'HLG', 'HMHC', 'HRB', 'LIND', 'LOV', 'LQDT', 'LRN', 'LTRPA', 'LTRPB', 'MIN', 'MMYT', 'MRKR', 'OSW', 'PLNT', 'PRDO', 'PTON', 'RGS', 'ROL', 'RST', 'SCI', 'SERV', 'STON', 'STRA', 'TA', 'TAL', 'TCOM', 'TOUR', 'TRCH', 'TRIP', 'UNF', 'VAC', 'WW', 'XSPA', 'YCBD', 'YTRA', 'ZVO']
other_consumer_services_master_file='other_consumer_services_mf.csv'
other_consumer_services_directory = 'other_consumer_services'
get_industry_ticker_data(other_consumer_services ,other_consumer_services_directory)
create_industry_master_list(other_consumer_services ,other_consumer_services_master_file,other_consumer_services_directory)
print("-> Master file created, filename: %s\n" % other_consumer_services_master_file)
purge_industry_low_gain_volume(other_consumer_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
other_consumer_specialties=['ASPU', 'CTHR', 'FIT', 'FORD', 'FOSL', 'HMI', 'KGJI', 'MOV', 'MSA', 'TLF']
other_consumer_specialties_master_file='other_consumer_special_mf.csv'
other_consumer_specialties_directory = 'other_consumer_specialties'
get_industry_ticker_data(other_consumer_specialties ,other_consumer_specialties_directory)
create_industry_master_list(other_consumer_specialties ,other_consumer_specialties_master_file,other_consumer_specialties_directory)
print("-> Master file created, filename: %s\n" % other_consumer_specialties_master_file)
purge_industry_low_gain_volume(other_consumer_specialties_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
other_metals_minerals=['AQMS', 'AUG', 'AXU', 'BBL', 'BHP', 'CCJ', 'CINR', 'CMCL', 'CSTE', 'DNN', 'EQX', 'FCX', 'GMO', 'GSV', 'HBM', 'HCR', 'IPI', 'LEU', 'LODE', 'MMX', 'MTRN', 'NAK', 'NEXA', 'NXE', 'PLM', 'PZG', 'RETO', 'RIO', 'SCCO', 'SLCA', 'TECK', 'TMQ', 'TRQ', 'UEC', 'URG', 'USAS', 'UUUU', 'VEDL', 'WRN', 'WWR', 'XPL']
other_metals_minerals_master_file='other_metals_minerals_mf.csv'
other_metals_minerals_directory = 'other_metals_minerals'
get_industry_ticker_data(other_metals_minerals ,other_metals_minerals_directory)
create_industry_master_list(other_metals_minerals ,other_metals_minerals_master_file,other_metals_minerals_directory)
print("-> Master file created, filename: %s\n" % other_metals_minerals_master_file)
purge_industry_low_gain_volume(other_metals_minerals_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
other_transportation=['ASR', 'CAAP', 'ECHO', 'ERA', 'HMLP', 'MATX', 'OMAB', 'PAC', 'PRSC', 'SMHI', 'VRRM']
other_transportation_master_file='other_transportation_mf.csv'
other_transportation_directory = 'other_transportation'
get_industry_ticker_data(other_transportation ,other_transportation_directory)
create_industry_master_list(other_transportation ,other_transportation_master_file,other_transportation_directory)
print("-> Master file created, filename: %s\n" % other_transportation_master_file)
purge_industry_low_gain_volume(other_transportation_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
packaged_software=['ACIW', 'ADBE', 'ADSK', 'AGYS', 'ALTR', 'AMSWA', 'ANSS', 'APPN', 'ASUR', 'AVID', 'AVLR', 'AYX', 'AZPN', 'BB', 'BILL', 'BKI', 'BL', 'BLKB', 'BNFT', 'BSQR', 'CANG', 'CDNS', 'CHNG', 'CHNGU', 'CLDR', 'CMCM', 'COUP', 'CREX', 'CRM', 'CRNC', 'CRWD', 'CSOD', 'CTK', 'CTXS', 'CVET', 'CVLT', 'CYBR', 'DAO', 'DDOG', 'DOCU', 'DOMO', 'DOYU', 'DT', 'EB', 'ECOM', 'EGHT', 'EH', 'ENV', 'EPAY', 'ESTC', 'EVBG', 'FICO', 'FNJN', 'GAN', 'GEC', 'GSB', 'GSUM', 'GTYH', 'HCAT', 'HSTM', 'HUYA', 'ICAD', 'IIIV', 'INS', 'INSE', 'INTU', 'IO', 'IZEA', 'LAIX', 'LIZI', 'LKCO', 'LN', 'LVGO', 'LYFT', 'MANH', 'MANT', 'MDB', 'MDLA', 'MFGP', 'MGIC', 'MIME', 'MSFT', 'MTLS', 'MWK', 'NATI', 'NCTY', 'NLOK', 'NTNX', 'NUAN', 'NVEC', 'OKTA', 'OPRA', 'ORCL', 'OTEX', 'PAYC', 'PBTS', 'PCTY', 'PD', 'PHR', 'PHUN', 'PING', 'PLAN', 'PRGS', 'PRSP', 'PRTH', 'PT', 'PTC', 'QLYS', 'RDCM', 'RDVT', 'RMNI', 'RNG', 'RNWK', 'SABR', 'SAIL', 'SAP', 'SCPL', 'SCWX', 'SDGR', 'SGLB', 'SLP', 'SMAR', 'SMSI', 'SNPS', 'SPNS', 'SPRT', 'SPSC', 'SPT', 'STNE', 'SWI', 'TEAM', 'TENB', 'TKAT', 'TLND', 'TWLO', 'TWOU', 'UBER', 'UEPS', 'UPWK', 'VEEV', 'VERI', 'WIMI', 'WORK', 'XAIR', 'YVR', 'ZDGE', 'ZM', 'ZS', 'ZUO']
packaged_software_master_file='packaged_software_mf.csv'
packaged_software_directory = 'packaged_software'
get_industry_ticker_data(packaged_software ,packaged_software_directory)
create_industry_master_list(packaged_software ,packaged_software_master_file,packaged_software_directory)
print("-> Master file created, filename: %s\n" % packaged_software_master_file)
purge_industry_low_gain_volume(packaged_software_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
personnel_services=['AMN', 'ASGN', 'BBSI', 'BGSF', 'CCRN', 'DLHC', 'GVP', 'HQI', 'HSII', 'HSON', 'JOB', 'KELYA', 'KELYB', 'KFRC', 'KFY', 'MAN', 'MHH', 'NSP', 'PIXY', 'PRFT', 'RGP', 'RHI', 'STAF', 'TBI', 'TSRI', 'VOLT']
personnel_services_master_file='personnel_services_mf.csv'
personnel_services_directory = 'personnel_services'
get_industry_ticker_data(personnel_services ,personnel_services_directory)
create_industry_master_list(personnel_services ,personnel_services_master_file,personnel_services_directory)
print("-> Master file created, filename: %s\n" % personnel_services_master_file)
purge_industry_low_gain_volume(personnel_services_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
pharmaceuticals_generic=['ABEO', 'BBIO', 'COCP', 'CPHI', 'JAGX', 'KIN', 'LFVN', 'MYL', 'RDY', 'TARO', 'TEVA', 'TXMD', 'ZTS']
pharmaceuticals_generic_master_file='pharmaceuticals_generic_mf.csv'
pharmaceuticals_generic_directory = 'pharmaceuticals_generic'
get_industry_ticker_data(pharmaceuticals_generic ,pharmaceuticals_generic_directory)
create_industry_master_list(pharmaceuticals_generic ,pharmaceuticals_generic_master_file,pharmaceuticals_generic_directory)
print("-> Master file created, filename: %s\n" % pharmaceuticals_generic_master_file)
purge_industry_low_gain_volume(pharmaceuticals_generic_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
pharmaceuticals_major=['ABBV', 'ACER', 'ACIU', 'ADCT', 'ADIL', 'ADMP', 'AGE', 'AGLE', 'AKCA', 'AKRO', 'AKTX', 'ALLK', 'ALNA', 'ALRN', 'AMAG', 'AMRX', 'ANVS', 'APLS', 'APLT', 'APM', 'APRE', 'AQST', 'ARCT', 'ARDS', 'ARPO', 'ARQT', 'ASRT', 'ATNX', 'ATXI', 'AYLA', 'AZN', 'BBI', 'BHVN', 'BMY', 'BSTC', 'BVXV', 'BXRX', 'BYSI', 'CALT', 'CKPT', 'CLSD', 'CNSP', 'CNST', 'CRNX', 'CRVS', 'CTXR', 'CYCN', 'DARE', 'DRRX', 'EDSA', 'ENTX', 'EOLS', 'EPIX', 'ETNB', 'ETON', 'EVLO', 'EYEN', 'EYPT', 'FULC', 'GMDA', 'GNPX', 'GOSS', 'GRFS', 'GRTS', 'GRTX', 'GSK', 'GTHX', 'HCM', 'HOOK', 'HOTH', 'HROW', 'HSTO', 'IDYA', 'IGMS', 'IMAB', 'IMRA', 'IMRN', 'IMUX', 'JNJ', 'KALA', 'KALV', 'KLDO', 'KNSA', 'KRTX', 'KRYS', 'KTOV', 'LCI', 'LEGN', 'LLY', 'LOGC', 'LQDA', 'LYRA', 'MBIO', 'MBRX', 'MDWD', 'MIRM', 'MIST', 'MLND', 'MNLO', 'MRK', 'MTEX', 'NAVB', 'NBSE', 'NKTR', 'NOVN', 'NVO', 'NVS', 'NVUS', 'OBSV', 'OCUL', 'ODT', 'ONCT', 'OPK', 'ORIC', 'ORTX', 'OSMT', 'OVID', 'OYST', 'PAHC', 'PDSB', 'PETQ', 'PFE', 'PHAS', 'PHAT', 'PHGE', 'PHGE.U', 'PRNB', 'PRPH', 'PRVB', 'QTRX', 'RDHL', 'RETA', 'RGEN', 'RLMD', 'SBPH', 'SCPH', 'SIGA', 'SLGL', 'SNY', 'SPRO', 'STRO', 'STSA', 'SWTX', 'SXTC', 'TAK', 'TARA', 'TCDA', 'TCRR', 'TFFP', 'THTX', 'TLGT', 'TMBR', 'TPTX', 'TRVI', 'TTNP', 'TYME', 'UROV', 'VERU', 'VRCA', 'XERS', 'YMAB', 'ZEAL', 'ZYME']
pharmaceuticals_major_master_file='pharmaceuticals_major_mf.csv'
pharmaceuticals_major_directory = 'pharmaceuticals_major'
get_industry_ticker_data(pharmaceuticals_major ,pharmaceuticals_major_directory)
create_industry_master_list(pharmaceuticals_major ,pharmaceuticals_major_master_file,pharmaceuticals_major_directory)
print("-> Master file created, filename: %s\n" % pharmaceuticals_major_master_file)
purge_industry_low_gain_volume(pharmaceuticals_major_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
pharmaceuticals_other=['ACRS', 'ACRX', 'ACST', 'ADMS', 'AERI', 'AGRX', 'ALBO', 'AMPE', 'AMPH', 'AMRN', 'ANIK', 'ANIP', 'ARGX', 'ASMB', 'AXSM', 'BDSI', 'BHC', 'BLRX', 'BRBR', 'CERC', 'COLL', 'CORT', 'CPIX', 'CPRX', 'CTIC', 'CTLT', 'EGRX', 'ENDP', 'EVOK', 'EYEG', 'GLMD', 'GWPH', 'HAPP', 'HEPA', 'HZNP', 'ICLR', 'ICPT', 'INVA', 'IRWD', 'ISR', 'JAZZ', 'LPCN', 'MDGL', 'MEIP', 'MNK', 'MRNS', 'MYOK', 'NAII', 'NEOS', 'NTEC', 'ORMP', 'PBH', 'PCRX', 'PLXP', 'PRGO', 'PTLA', 'QLGN', 'RGLS', 'RVMD', 'RVNC', 'SAGE', 'SAVA', 'SCYX', 'SNOA', 'SUPN', 'TBPH', 'TNXP', 'TTPH', 'UTHR', 'VRNA', 'VVUS', 'XENT', 'ZGNX', 'ZYNE']
pharmaceuticals_other_master_file='pharmaceuticals_other_mf.csv'
pharmaceuticals_other_directory = 'pharmaceuticals_other'
get_industry_ticker_data(pharmaceuticals_other ,pharmaceuticals_other_directory)
create_industry_master_list(pharmaceuticals_other ,pharmaceuticals_other_master_file,pharmaceuticals_other_directory)
print("-> Master file created, filename: %s\n" % pharmaceuticals_other_master_file)
purge_industry_low_gain_volume(pharmaceuticals_other_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
precious_metals=['AAU', 'AEM', 'AG', 'AGI', 'ALO', 'ASM', 'AU', 'AUMN', 'AUY', 'BTG', 'BVN', 'CDE', 'DRD', 'EGO', 'EMX', 'EXK', 'FNV', 'FSM', 'GAU', 'GFI', 'GOLD', 'GORO', 'GPL', 'GSS', 'HL', 'HMY', 'HYMC', 'IAG', 'KGC', 'KL', 'MAG', 'MUX', 'NEM', 'NG', 'NGD', 'OR', 'PAAS', 'PLG', 'PVG', 'RGLD', 'SA', 'SAND', 'SBSW', 'SILV', 'SMTS', 'SSRM', 'SVM', 'TGB', 'THM', 'TRX', 'USAU', 'VGZ', 'WPM']
precious_metals_master_file='precious_metals_mf.csv'
precious_metals_directory = 'precious_metals'
get_industry_ticker_data(precious_metals ,precious_metals_directory)
create_industry_master_list(precious_metals ,precious_metals_master_file,precious_metals_directory)
print("-> Master file created, filename: %s\n" % precious_metals_master_file)
purge_industry_low_gain_volume(precious_metals_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
property_casualty_insurance=['AFG', 'AFH', 'ALL', 'CB', 'CINF', 'CNFR', 'CNFRL', 'ERIE', 'FNHC', 'GBLI', 'GLRE', 'HALL', 'HCI', 'HRTG', 'ICCH', 'JRVR', 'KFS', 'L', 'MCY', 'MHLA', 'MHLD', 'NGHC', 'NMIH', 'NSEC', 'ORI', 'OXBR', 'PGR', 'PIH', 'PLMR', 'PPHI', 'PROS', 'PTVCA', 'PTVCB', 'RE', 'RLI', 'RNR', 'SAFT', 'SIGI', 'STFC', 'THG', 'TPRE', 'UFCS', 'UIHC', 'UNAM', 'UVE', 'WRB', 'Y']
property_casualty_insurance_master_file='property_casualty_insur_mf.csv'
property_casualty_insurance_directory = 'property_casualty_insurance'
get_industry_ticker_data(property_casualty_insurance ,property_casualty_insurance_directory)
create_industry_master_list(property_casualty_insurance ,property_casualty_insurance_master_file,property_casualty_insurance_directory)
print("-> Master file created, filename: %s\n" % property_casualty_insurance_master_file)
purge_industry_low_gain_volume(property_casualty_insurance_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
publishing_books_magazines=['EDUC','JW.A','JB.B','MDP','PSO','SCHL']
publishing_books_magazines_master_file='publishing_books_mags_mf.csv'
publishing_books_magazines_directory = 'publishing_books_magazines'
get_industry_ticker_data(publishing_books_magazines ,publishing_books_magazines_directory)
create_industry_master_list(publishing_books_magazines ,publishing_books_magazines_master_file,publishing_books_magazines_directory)
print("-> Master file created, filename: %s\n" % publishing_books_magazines_master_file)
purge_industry_low_gain_volume(publishing_books_magazines_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
publishing_newspapers=['AHC','DJCO','GCI','LEE','NWS','NWSA','NYT','TPCO']
publishing_newspapers_master_file='publishing_newspapers_mf.csv'
publishing_newspapers_directory = 'publishing_newspapers'
get_industry_ticker_data(publishing_newspapers ,publishing_newspapers_directory)
create_industry_master_list(publishing_newspapers ,publishing_newspapers_master_file,publishing_newspapers_directory)
print("-> Master file created, filename: %s\n" % publishing_newspapers_master_file)
purge_industry_low_gain_volume(publishing_newspapers_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
pulp_paper=['CLW','GLT','ITP','MERC','NP','RFP','SUZ','SWM','UFS','VRS']
pulp_paper_master_file='pulp_paper_mf.csv'
pulp_paper_directory = 'pulp_paper'
get_industry_ticker_data(pulp_paper ,pulp_paper_directory)
create_industry_master_list(pulp_paper ,pulp_paper_master_file,pulp_paper_directory)
print("-> Master file created, filename: %s\n" % pulp_paper_master_file)
purge_industry_low_gain_volume(pulp_paper_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
railroads=['CNI','CP','CSX','GSH','KSU','NSC','UNP']
railroads_master_file='railroads_mf.csv'
railroads_directory = 'railroads'
get_industry_ticker_data(railroads ,railroads_directory)
create_industry_master_list(railroads ,railroads_master_file,railroads_directory)
print("-> Master file created, filename: %s\n" % railroads_master_file)
purge_industry_low_gain_volume(railroads_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
real_estate_development=['ARL', 'AXR', 'BPY', 'CBRE', 'CIGI', 'CRESY', 'CTO', 'CWK', 'EPRT', 'FOR', 'FPH', 'FRPH', 'FRSX', 'FSV', 'GRIF', 'GYRO', 'HGV', 'HHC', 'INTG', 'IRCP', 'IRS', 'JCAP', 'JLL', 'JOE', 'KW', 'LEJU', 'LMRK', 'MAYS', 'MDJH', 'MLP', 'MMI', 'NEN', 'NMRK', 'PICO', 'PINE', 'QK', 'RDFN', 'RDI', 'RDIB', 'RFL', 'RLGY', 'RMAX', 'RMR', 'STRS', 'TPHS', 'TRC', 'VICI', 'XIN']
real_estate_development_master_file='real_estate_development_mf.csv'
real_estate_development_directory = 'real_estate_development'
get_industry_ticker_data(real_estate_development ,real_estate_development_directory)
create_industry_master_list(real_estate_development ,real_estate_development_master_file,real_estate_development_directory)
print("-> Master file created, filename: %s\n" % real_estate_development_master_file)
purge_industry_low_gain_volume(real_estate_development_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
real_estate_investment_trusts=['AAT', 'ABR', 'ACC', 'ADC', 'AFIN', 'AGNC', 'AHH', 'AHT', 'AI', 'AIV', 'AKR', 'ALEX', 'ALX', 'AMH', 'AMT', 'ANH', 'APLE', 'APTS', 'ARE', 'ARI', 'ARR', 'AVB', 'BDN', 'BFS', 'BHR', 'BPYU', 'BRG', 'BRMK', 'BRT', 'BRX', 'BXMT', 'BXP', 'CBL', 'CCI', 'CDOR', 'CDR', 'CHCT', 'CHMI', 'CIM', 'CIO', 'CLDT', 'CLI', 'CLNC', 'CLNY', 'CLPR', 'CMCT', 'CMO', 'COLD', 'CONE', 'COR', 'CPLG', 'CPT', 'CTRE', 'CTT', 'CUBE', 'CUZ', 'CXP', 'CXW', 'DEA', 'DEI', 'DHC', 'DLR', 'DOC', 'DRE', 'DRH', 'DX', 'EARN', 'EGP', 'ELS', 'EPR', 'EQC', 'EQIX', 'EQR', 'ESBA', 'ESRT', 'ESS', 'EXR', 'FCPT', 'FISK', 'FPI', 'FR', 'FRT', 'FSP', 'GEO', 'GLPI', 'GMRE', 'GNL', 'GOOD', 'GPMT', 'GRP.U', 'GTY', 'HASI', 'HCFT', 'HIW', 'HMG', 'HPP', 'HR', 'HST', 'HT', 'HTA', 'IHT', 'IIPR', 'ILPT', 'INN', 'INVH', 'IOR', 'IRET', 'IRM', 'IRT', 'IVR', 'JBGS', 'KIM', 'KRC', 'KREF', 'KRG', 'LAMR', 'LAND', 'LOAN', 'LSI', 'LTC', 'LXP', 'MAA', 'MAC', 'MDRR', 'MFA', 'MGP', 'MITT', 'MNR', 'MPW', 'NHI', 'NLY', 'NNN', 'NREF', 'NRZ', 'NSA', 'NXRT', 'NYMT', 'O', 'OFC', 'OGCP', 'OHI', 'OLP', 'OPI', 'ORC', 'OUT', 'PCH', 'PDM', 'PEAK', 'PEB', 'PEI', 'PGRE', 'PK', 'PLD', 'PLYM', 'PMT', 'PSA', 'PSB', 'PSTL', 'PW', 'QTS', 'RC', 'REG', 'RESI', 'REXR', 'RHE', 'RHP', 'RLJ', 'ROIC', 'RPAI', 'RPT', 'RVI', 'RWT', 'RYN', 'SACH', 'SAFE', 'SBAC', 'SBRA', 'SELF', 'SHO', 'SITC', 'SKT', 'SLG', 'SNR', 'SOHO', 'SPG', 'SRC', 'SRG', 'STAG', 'STAR', 'STOR', 'STWD', 'SUI', 'SVC', 'TCI', 'TCO', 'TRMT', 'TRNO', 'TRTX', 'TWO', 'UBA', 'UBP', 'UDR', 'UE', 'UHT', 'UMH', 'UNIT', 'VER', 'VNO', 'VTR', 'WELL', 'WHLR', 'WMC', 'WPC', 'WPG', 'WRE', 'WRI', 'WSR', 'WY', 'XAN', 'XHR']
real_estate_investment_trusts_master_file='reit_mf.csv'
real_estate_investment_trusts_directory = 'real_estate_investment_trusts'
get_industry_ticker_data(real_estate_investment_trusts ,real_estate_investment_trusts_directory)
create_industry_master_list(real_estate_investment_trusts ,real_estate_investment_trusts_master_file,real_estate_investment_trusts_directory)
print("-> Master file created, filename: %s\n" % real_estate_investment_trusts_master_file)
purge_industry_low_gain_volume(real_estate_investment_trusts_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
recreational_products=['ATVI', 'BC', 'BHAT', 'CLAR', 'DOOO', 'EA', 'ELY', 'ESCA', 'FNKO', 'FOXF', 'GIGM', 'GOLF', 'HAS', 'JAKK', 'JOUT', 'MAT', 'MBUU', 'MCFT', 'MPX', 'NLS', 'PII', 'PTE', 'RGR', 'SLGG', 'THO', 'TTWO', 'VSTO', 'WGO']
recreational_products_master_file='recreational_products_mf.csv'
recreational_products_directory = 'recreational_products'
get_industry_ticker_data(recreational_products ,recreational_products_directory)
create_industry_master_list(recreational_products ,recreational_products_master_file,recreational_products_directory)
print("-> Master file created, filename: %s\n" % recreational_products_master_file)
purge_industry_low_gain_volume(recreational_products_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
regional_banks=['ABCB', 'ABTX', 'ACBI', 'ACNB', 'ALLY', 'AMAL', 'AMRB', 'ASB', 'ASRV', 'ATLO', 'AUB', 'AUBN', 'BANC', 'BANF', 'BANR', 'BAP', 'BBAR', 'BCBP', 'BCH', 'BCML', 'BDGE', 'BFC', 'BFST', 'BHB', 'BHLB', 'BKSC', 'BKU', 'BLX', 'BMA', 'BMRC', 'BMTC', 'BOCH', 'BOH', 'BOKF', 'BOTJ', 'BPFH', 'BPOP', 'BPRN', 'BSAC', 'BSBR', 'BSMX', 'BSRR', 'BSVN', 'BUSE', 'BWB', 'BWFG', 'BXS', 'BY', 'CAC', 'CALB', 'CARE', 'CATC', 'CATY', 'CBAN', 'CBFV', 'CBSH', 'CBU', 'CCB', 'CCBG', 'CCNE', 'CFB', 'CFFI', 'CFG', 'CFR', 'CHMG', 'CIB', 'CIVB', 'CLDB', 'CNBKA', 'CNOB', 'COFS', 'COLB', 'CPF', 'CSFL', 'CSTR', 'CTBI', 'CUBI', 'CVBF', 'CVCY', 'CVLY', 'CWBC', 'CZNC', 'DBCP', 'DFS', 'EBMT', 'EBTC', 'EFSC', 'EGBN', 'EMCF', 'EQBK', 'ESSA', 'ESXB', 'EVBN', 'EWBC', 'FBNC', 'FBP', 'FBSS', 'FCBC', 'FCBP', 'FCCO', 'FCCY', 'FCF', 'FCNCA', 'FDBC', 'FFBC', 'FFIC', 'FFIN', 'FFWM', 'FGBI', 'FHB', 'FHN', 'FIBK', 'FISI', 'FITB', 'FLIC', 'FMBH', 'FMBI', 'FMNB', 'FNB', 'FNCB', 'FRAF', 'FRBA', 'FRBK', 'FRC', 'FRME', 'FSB', 'FSFG', 'FULT', 'FUNC', 'FUSB', 'FVCB', 'FXNC', 'GABC', 'GBCI', 'GGAL', 'GLBZ', 'GNTY', 'GWB', 'HAFC', 'HBAN', 'HBMD', 'HBNC', 'HBT', 'HDB', 'HFWA', 'HMST', 'HOMB', 'HONE', 'HTBI', 'HTBK', 'HTH', 'HTLF', 'HVBC', 'HWBK', 'HWC', 'IBCP', 'IBN', 'IBOC', 'IBTX', 'ICBK', 'INBK', 'INDB', 'ISTR', 'ITCB', 'KB', 'LBAI', 'LCNB', 'LEVL', 'LKFN', 'LMST', 'LOB', 'MBCN', 'MBIN', 'MBWM', 'MCBC', 'MCBS', 'MFNC', 'MNSB', 'MOFG', 'MPB', 'MRBK', 'MSBI', 'MTB', 'MVBF', 'NBHC', 'NBN', 'NBTB', 'NCBS', 'NRIM', 'NTB', 'NTRS', 'NWFL', 'OBNK', 'ONB', 'OPBK', 'OPHC', 'ORRF', 'OSBC', 'OVBC', 'OVLY', 'OZK', 'PACW', 'PB', 'PBFS', 'PCB', 'PCSB', 'PDLB', 'PEBK', 'PFBC', 'PFBI', 'PFHD', 'PFIS', 'PGC', 'PKBK', 'PLBC', 'PNBK', 'PNFP', 'PPBI', 'PUB', 'PWOD', 'QCRH', 'RBB', 'RBCAA', 'RBKB', 'RBNC', 'RIVE', 'RNST', 'RRBI', 'SAL', 'SASR', 'SBBX', 'SBCF', 'SBNY', 'SBSI', 'SFBC', 'SFBS', 'SFST', 'SHBI', 'SI', 'SIVB', 'SLCT', 'SMBK', 'SMMF', 'SNV', 'SONA', 'SPFI', 'SRCE', 'SSB', 'SSBI', 'STBA', 'STL', 'STT', 'STXB', 'SYBT', 'TBBK', 'TBK', 'TCBI', 'TCBK', 'TCFC', 'TFC', 'TMP', 'TOWN', 'TRMK', 'TRST', 'TSC', 'UBCP', 'UBFO', 'UBOH', 'UBSI', 'UCBI', 'UMBF', 'UMPQ', 'UNB', 'UNTY', 'UVSP', 'VBFC', 'VBTX', 'VLY', 'WABC', 'WAL', 'WALA', 'WASH', 'WSBC', 'WTFC', 'ZION']
regional_banks_master_file='regional_banks_mf.csv'
regional_banks_directory = 'regional_banks'
get_industry_ticker_data(regional_banks ,regional_banks_directory)
create_industry_master_list(regional_banks ,regional_banks_master_file,regional_banks_directory)
print("-> Master file created, filename: %s\n" % regional_banks_master_file)
purge_industry_low_gain_volume(regional_banks_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
restaurants=['ARCO', 'ARKR', 'ARMK', 'BBQ', 'BDL', 'BH', 'BH.A', 'BJRI', 'BLMN', 'CAKE', 'CBRL', 'CHUY', 'CMG', 'DENN', 'DIN', 'DNKN', 'DPZ', 'DRI', 'EAT', 'FAT', 'FRGI', 'GRIL', 'GTIM', 'JACK', 'JAX', 'KRUS', 'LOCO', 'LUB', 'MCD', 'NATH', 'NDLS', 'PBPB', 'PZZA', 'QSR', 'RAVE', 'RRGB', 'RUTH', 'SBUX', 'SHAK', 'STKS', 'TACO', 'TAST', 'TXRH', 'WEN', 'WING', 'YUM', 'YUMC']
restaurants_master_file='restaurants_mf.csv'
restaurants_directory = 'restaurants'
get_industry_ticker_data(restaurants ,restaurants_directory)
create_industry_master_list(restaurants ,restaurants_master_file,restaurants_directory)
print("-> Master file created, filename: %s\n" % restaurants_master_file)
purge_industry_low_gain_volume(restaurants_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
savings_banks=['AX', 'BBX', 'BCOW', 'BCTF', 'BRKL', 'BSBK', 'BYFC', 'CARV', 'CBMB', 'CFBI', 'CFBK', 'CFFN', 'CLBK', 'CNNB', 'DCOM', 'EBSB', 'ESBK', 'FCAP', 'FDEF', 'FFBW', 'FFNW', 'FNWB', 'FSEA', 'GCBC', 'HFBL', 'HIFS', 'HMNF', 'IBKC', 'IROQ', 'ISBC', 'KFFB', 'KRNY', 'LBC', 'LSBK', 'MGYR', 'MLVF', 'MSBF', 'MSVB', 'NFBK', 'NWBI', 'NYCB', 'OCFC', 'OTTW', 'PBHC', 'PBIP', 'PFS', 'PROV', 'RNDB', 'RVSB', 'SBT', 'SMBC', 'STND', 'SVBI', 'TFSL', 'TSBK', 'WAFD', 'WBS', 'WNEB', 'WSFS', 'WVFC']
savings_banks_master_file='savings_banks_mf.csv'
savings_banks_directory = 'savings_banks'
get_industry_ticker_data(savings_banks ,savings_banks_directory)
create_industry_master_list(savings_banks ,savings_banks_master_file,savings_banks_directory)
print("-> Master file created, filename: %s\n" % savings_banks_master_file)
purge_industry_low_gain_volume(savings_banks_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
semiconductors=['ADI', 'AMBA', 'AMD', 'AMKR', 'AOSL', 'ASX', 'AVGO', 'BOSC', 'CAN', 'CCMP', 'CEVA', 'COHR', 'CRUS', 'DIOD', 'DSPG', 'ENPH', 'ENTG', 'FN', 'GSIT', 'HIMX', 'ICHR', 'IMOS', 'INTC', 'INVE', 'IOTS', 'IPGP', 'IPHI', 'KLIC', 'KOPN', 'LASR', 'LSCC', 'MCHP', 'MOSY', 'MPWR', 'MRAM', 'MRVL', 'MTSI', 'MU', 'MX', 'MXIM', 'MXL', 'NVDA', 'NXPI', 'OIIM', 'ON', 'OSIS', 'POWI', 'PXLW', 'QRVO', 'QUIK', 'RBCN', 'RMBS', 'SCON', 'SGH', 'SIMO', 'SITM', 'SLAB', 'SMTC', 'SOL', 'SPI', 'SQNS', 'STM', 'SWKS', 'TRT', 'TSEM', 'TSM', 'TXN', 'UCTT', 'UMC', 'WISA', 'XLNX', 'XPER']
semiconductors_master_file='semiconductors_mf.csv'
semiconductors_directory = 'semiconductors'
get_industry_ticker_data(semiconductors ,semiconductors_directory)
create_industry_master_list(semiconductors ,semiconductors_master_file,semiconductors_directory)
print("-> Master file created, filename: %s\n" % semiconductors_master_file)
purge_industry_low_gain_volume(semiconductors_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
services_to_the_health_industry=['AMEH', 'BASI', 'DGX', 'IDXG', 'IQV', 'LH', 'NRC', 'PLX', 'PMD', 'SYNH']
services_to_the_health_industry_master_file='services_to_hlth_ind_mf.csv'
services_to_the_health_industry_directory = 'services_to_the_health_industry'
get_industry_ticker_data(services_to_the_health_industry ,services_to_the_health_industry_directory)
create_industry_master_list(services_to_the_health_industry ,services_to_the_health_industry_master_file,services_to_the_health_industry_directory)
print("-> Master file created, filename: %s\n" % services_to_the_health_industry_master_file)
purge_industry_low_gain_volume(services_to_the_health_industry_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
specialty_insurance=['AGO', 'AMBC', 'ATH', 'ESNT', 'FAF', 'FNF', 'ITIC', 'MBI', 'MKL', 'MTG', 'PRA', 'RDN', 'STC', 'TIPT', 'TRUP']
specialty_insurance_master_file='specialty_insurance_mf.csv'
specialty_insurance_directory = 'specialty_insurance'
get_industry_ticker_data(specialty_insurance ,specialty_insurance_directory)
create_industry_master_list(specialty_insurance ,specialty_insurance_master_file,specialty_insurance_directory)
print("-> Master file created, filename: %s\n" % specialty_insurance_master_file)
purge_industry_low_gain_volume(specialty_insurance_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
specialty_stores=['AAP', 'ABG', 'AN', 'AZO', 'BBBY', 'BBW', 'BGFV', 'BGI', 'BJ', 'BNED', 'BWMX', 'CASY', 'COST', 'CRMT', 'CVNA', 'CWH', 'DKS', 'ELA', 'EYE', 'GME', 'GNC', 'GPI', 'GRUB', 'HIBB', 'HOME', 'HUD', 'HVT', 'HVT.A', 'HZO', 'IAA', 'KIRK', 'KMX', 'KXIN', 'LAD', 'LAZY', 'LL', 'LOVE', 'MED', 'MIK', 'MNRO', 'MUSA', 'ODP', 'ONEW', 'ORLY', 'PAG', 'PRTY', 'RH', 'SAH', 'SBH', 'SIG', 'SPWH', 'TCS', 'TGT', 'TIF', 'TSCO', 'TUES', 'ULTA', 'WSM']
specialty_stores_master_file='specialty_stores_mf.csv'
specialty_stores_directory = 'specialty_stores'
get_industry_ticker_data(specialty_stores ,specialty_stores_directory)
create_industry_master_list(specialty_stores ,specialty_stores_master_file,specialty_stores_directory)
print("-> Master file created, filename: %s\n" % specialty_stores_master_file)
purge_industry_low_gain_volume(specialty_stores_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
specialty_telecommunications=['CBB', 'CCOI', 'CPAH', 'CTBB', 'CTDD', 'CTL', 'CTV', 'CTZ', 'ENT', 'GLIBA', 'GTT', 'IDT', 'KLR', 'LORL', 'MARK', 'OTEL', 'PCTI', 'PHI', 'RNET', 'TEF', 'VG', 'WOW']
specialty_telecommunications_master_file='specialty_telecomm_mf.csv'
specialty_telecommunications_directory = 'specialty_telecommunications'
get_industry_ticker_data(specialty_telecommunications ,specialty_telecommunications_directory)
create_industry_master_list(specialty_telecommunications ,specialty_telecommunications_master_file,specialty_telecommunications_directory)
print("-> Master file created, filename: %s\n" % specialty_telecommunications_master_file)
purge_industry_low_gain_volume(specialty_telecommunications_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
steel=['AP', 'ATI', 'CLF', 'CMC', 'CRS', 'GGB', 'HAYN', 'MSB', 'MT', 'MTL', 'NUE', 'OSN', 'PKX', 'RS', 'RYI', 'SCHN', 'SID', 'STLD', 'SYNL', 'TMST', 'TS', 'TX', 'USAP', 'VALE', 'WOR', 'X', 'ZEUS']
steel_master_file='steel_mf.csv'
steel_directory = 'steel'
get_industry_ticker_data(steel ,steel_directory)
create_industry_master_list(steel ,steel_master_file,steel_directory)
print("-> Master file created, filename: %s\n" % steel_master_file)
purge_industry_low_gain_volume(steel_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
telecommunications_equipment=['AAPL', 'ADTN', 'AIRG', 'AKTS', 'APWC', 'AUDC', 'AVNW', 'AVYA', 'BDR', 'BKTI', 'CALX', 'CAMP', 'CASA', 'CIEN', 'CLFD', 'CLRO', 'CMBM', 'CMTL', 'COMM', 'CRNT', 'DGLY', 'DZSI', 'ERIC', 'EXFO', 'GILT', 'GRMN', 'HLIT', 'IDCC', 'INFN', 'INSG', 'ITRN', 'JCS', 'NOK', 'OCC', 'PLT', 'PWFL', 'QCOM', 'RBBN', 'SATS', 'SONM', 'SWIR', 'TCCO', 'TESS', 'UI', 'UTSI', 'VISL', 'VOXX', 'VSAT', 'WATT', 'WSTL', 'WTT']
telecommunications_equipment_master_file='telecommunications_equipment_mf.csv'
telecommunications_equipment_directory = 'telecommunications_equipment'
get_industry_ticker_data(telecommunications_equipment ,telecommunications_equipment_directory)
create_industry_master_list(telecommunications_equipment ,telecommunications_equipment_master_file,telecommunications_equipment_directory)
print("-> Master file created, filename: %s\n" % telecommunications_equipment_master_file)
purge_industry_low_gain_volume(telecommunications_equipment_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
textiles=['AIN','CULP','UFI']
textiles_master_file='textiles_mf.csv'
textiles_directory = 'textiles'
get_industry_ticker_data(textiles ,textiles_directory)
create_industry_master_list(textiles ,textiles_master_file,textiles_directory)
print("-> Master file created, filename: %s\n" % textiles_master_file)
purge_industry_low_gain_volume(textiles_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
tobacco=['BTI','MO','PM','PYX','SDI','TPB','UVV','VGR']
tobacco_master_file='tobacco_mf.csv'
tobacco_directory = 'tobacco'
get_industry_ticker_data(tobacco ,tobacco_directory)
create_industry_master_list(tobacco ,tobacco_master_file,tobacco_directory)
print("-> Master file created, filename: %s\n" % tobacco_master_file)
purge_industry_low_gain_volume(tobacco_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
tools_hardware=['SCX','SNA','SWK']
tools_hardware_master_file='tools_hardware_mf.csv'
tools_hardware_directory = 'tools_hardware'
get_industry_ticker_data(tools_hardware ,tools_hardware_directory)
create_industry_master_list(tools_hardware ,tools_hardware_master_file,tools_hardware_directory)
print("-> Master file created, filename: %s\n" % tools_hardware_master_file)
purge_industry_low_gain_volume(tools_hardware_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
trucking=['ARCB', 'CVTI', 'DSKE', 'GFL', 'GFLU', 'GPP', 'HTLD', 'JBHT', 'KNX', 'LSTR', 'MRTN', 'ODFL', 'PATI', 'PTSI', 'SAIA', 'SNDR', 'TFII', 'ULH', 'USAK', 'USX', 'WERN', 'XPO', 'YRCW']
trucking_master_file='trucking_mf.csv'
trucking_directory = 'trucking'
get_industry_ticker_data(trucking ,trucking_directory)
create_industry_master_list(trucking ,trucking_master_file,trucking_directory)
print("-> Master file created, filename: %s\n" % trucking_master_file)
purge_industry_low_gain_volume(trucking_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
trucks_construction_farm_machinery=['ACA', 'AGCO', 'ALG', 'ALSN', 'ARTW', 'ASTE', 'BGG', 'BLBD', 'CAT', 'CMCO', 'CMI', 'CNHI', 'CVGI', 'CYD', 'DE', 'FSS', 'FSTR', 'GBX', 'HY', 'LNN', 'MLR', 'MNTX', 'MTW', 'NAV', 'OSK', 'PCAR', 'PLOW', 'RAIL', 'SHYF', 'TEX', 'TRN', 'TTC', 'TWI', 'TWIN', 'WAB', 'WNC', 'WPRT']
trucks_construction_farm_machinery_master_file='tcfm.csv'
trucks_construction_farm_machinery_directory = 'trucks_construction_farm_machinery'
get_industry_ticker_data(trucks_construction_farm_machinery ,trucks_construction_farm_machinery_directory)
create_industry_master_list(trucks_construction_farm_machinery ,trucks_construction_farm_machinery_master_file,trucks_construction_farm_machinery_directory)
print("-> Master file created, filename: %s\n" % trucks_construction_farm_machinery_master_file)
purge_industry_low_gain_volume(trucks_construction_farm_machinery_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
water_utilities=['ARTNA', 'AWK', 'AWR', 'CDZI', 'CWCO', 'CWT', 'GWRS', 'MSEX', 'NES', 'PCYO', 'SBS', 'SJW', 'WTRG', 'WTRU', 'YORW']
water_utilities_master_file='water_utilities_mf.csv'
water_utilities_directory = 'water_utilities'
get_industry_ticker_data(water_utilities ,water_utilities_directory)
create_industry_master_list(water_utilities ,water_utilities_master_file,water_utilities_directory)
print("-> Master file created, filename: %s\n" % water_utilities_master_file)
purge_industry_low_gain_volume(water_utilities_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

print("-> Running Sympathy Sniffer to get you your sympathy play candidates\n")
wholesale_distributors=['ACU', 'AE', 'AEY', 'AIT', 'ALTG', 'AXE', 'BECN', 'BXC', 'CAPL', 'CHNR', 'CLBS', 'CORE', 'DIT', 'DMPI', 'DNOW', 'DXPE', 'EVI', 'FAST', 'FBM', 'GCO', 'GFN', 'GLP', 'GMS', 'GNLN', 'GPC', 'GRWG', 'GWW', 'HBP', 'HDS', 'HWKN', 'IBP', 'INT', 'JCTCF', 'LAWS', 'MRC', 'MSM', 'NOVA', 'NS', 'POOL', 'RUSHA', 'RUSHB', 'SITE', 'SUMR', 'SUN', 'SXT', 'TITN', 'UNVR', 'VRTV', 'WCC', 'WEYS']
wholesale_distributors_master_file='wholesale_distributors_mf.csv'
wholesale_distributors_directory = 'wholesale_distributors'
get_industry_ticker_data(wholesale_distributors ,wholesale_distributors_directory)
create_industry_master_list(wholesale_distributors ,wholesale_distributors_master_file,wholesale_distributors_directory)
print("-> Master file created, filename: %s\n" % wholesale_distributors_master_file)
purge_industry_low_gain_volume(wholesale_distributors_master_file)
print('\n\n-> And our Lord J Powell said... let there be no rest and lots of BRRRRRRR to your bottom line !')
print('-> Process completed, enjoy the potential gains !')

combine_masters_into_one('industry','master_industry_file.xlsx')


